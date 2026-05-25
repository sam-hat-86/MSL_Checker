import polars as pl
import tkinter as tk
from tkinter import filedialog, messagebox
import xlsxwriter
import openpyxl
import os
from datetime import datetime
from typing import Any, cast
import re
import unicodedata
import importlib
import multiprocessing as mp
import sys
import ctypes
import tkinter.font as tkfont
import traceback
import time
import math
import unicodedata


# 画面をきれいに表示する
def enable_windows_dpi_awareness() -> None:
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


# TkinterのスケーリングをDPIに合わせて調整し、UDフォントを設定する
def configure_tk_scaling(root: tk.Tk) -> None:
    # 画面のDPIを取得
    try:
        dpi = root.winfo_fpixels("1i")
        scale = float(dpi) / 96.0
        if scale <= 0:
            scale = 1.0
    except Exception:
        try:
            scale = float(root.tk.call("tk", "scaling"))
        except Exception:
            scale = 1.0

    # DPIをもとに画面の倍率を設定
    try:
        root.tk.call("tk", "scaling", scale)
    except Exception:
        pass

    try:
        # GUI画面のフォントも見やすいUDフォントに変更
        base = tkfont.nametofont("TkDefaultFont")
        size = int(max(base.cget("size") * scale, 10))
        base.configure(family="Noto Sans JP", size=size)
    except Exception:
        pass


# 進捗表示用の関数
def progress(msg: str) -> None:
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


# 宿題名に含まれるとカウントから除外する語のサンプル一覧。
# ここはユーザーが運用に合わせて編集してください（部分一致で判定されます）。
EXCLUDED_HW_TERMS: list[str] = [
    "たんご","単語",
    "しすたん","シスタン","シス単",
    "たーげっと","ターゲット",
    "りーぷ","リープ","Leap","LEAP",
    "まどんな","マドンナ",
]

MAX_HW_PAGES = 30

# ベンチ結果に基づく正規化並列化の推奨デフォルト値
# ユーザーは環境変数で上書き可能（NORMALIZE_WORKERS, NORMALIZE_CHUNKSIZE）。
DEFAULT_NORMALIZE_WORKERS = 6
DEFAULT_NORMALIZE_CHUNKSIZE = 100


def normalize_text_for_matching(s: object) -> str:
    """
    文字列を検索用に正規化するユーティリティ。

    - Unicode NFKC 正規化
    - 半角全角の整理（可能なら mojimoji を使用）
    - カタカナ→ひらがな化（可能なら jaconv を使用）
    - ASCII を小文字化
    - 空白類を除去

    この関数は外部ライブラリが利用可能な場合はそれらを利用し、
    利用不可でもフォールバックして最低限の正規化を行います。
    """
    if s is None:
        return ""
    text = str(s)
    # Unicode 正規化
    text = unicodedata.normalize("NFKC", text)

    # 全角→半角やカナ正規化は jaconv を使う
    try:
        jaconv = importlib.import_module("jaconv")
        # 全角英数字を半角に（カナは変換しない）、その後カタカナ→ひらがな
        text = jaconv.z2h(text, kana=False, ascii=True, digit=True)
        text = jaconv.kata2hira(text)
    except Exception:
        # フォールバック: jaconv が無ければ NFKC のみで進める
        pass

    # 英字は小文字化、空白は削除して比較用にする
    text = text.lower()
    text = re.sub(r"\s+", "", text)
    return text


def build_hw_pages_expr(
    hw_name_column: str,
    hw_start_column: str,
    hw_end_column: str,
    excluded_terms: list[str],
    page_limit: int = MAX_HW_PAGES,
) -> pl.Expr:
    """
    指定された宿題の3列（名前, 開始ページ, 終了ページ）を元に
    1行分の宿題ページ数を表すPolars式（`pl.Expr`）を生成します。

    振る舞い:
    - 宿題名が空文字列の場合は0を返します。
    - `excluded_terms` のいずれかが宿題名に部分一致する場合は0を返します。
    - 開始/終了ページからページ数を計算し、負値やNULLは0にします。
    - ページ数が `page_limit` を超える場合は `page_limit` に切り捨てます。

    注意:
    - 開始/終了ページの数値抽出は正規表現 `r"(\\d+)"` により最初の数値を取り出します。
    """
    # 宿題名列を文字列化してトリム。NULL は空文字に置換。
    hw_name_expr = pl.col(hw_name_column).cast(pl.Utf8).fill_null("").str.strip_chars()
    # 除外語リストから空文字列を除去して有効な語のみを扱う
    # 除外語リストから空文字を除外して有効語のみを扱う
    active_excluded_hw_terms = [term for term in excluded_terms if str(term).strip()]

    # 宿題名が除外語のいずれかを含むかどうか（部分一致・リテラル）を示す式
    # 除外語が空の場合は常に False を返すリテラル式を使う
    # 各除外語について宿題名に含まれるかをチェックする式を作り、
    # 横持ちで "どれか一つでも含む" なら True を返す式にまとめる。
    # 除外語リストが空なら常に False を返すリテラル式を使う。
    is_excluded_hw_expr = (
        pl.any_horizontal(
            [
                hw_name_expr.str.contains(term, literal=True)
                for term in active_excluded_hw_terms
            ]
        )
        if active_excluded_hw_terms
        else pl.lit(False)
    )

    # 開始/終了ページ列から最初の数値を抽出して整数化する
    # （例: "10" や "10ページ" -> 10）。複雑な範囲表記は別途正規化が必要。
    # 開始/終了ページ列から最初の数値を抽出して整数化する
    # 例: "10ページ" -> 10
    hw_start_page_expr = (
        pl.col(hw_start_column)
        .cast(pl.Utf8)
        .str.extract(r"(\\d+)")
        .cast(pl.Int64, strict=False)
    )
    hw_end_page_expr = (
        pl.col(hw_end_column)
        .cast(pl.Utf8)
        .str.extract(r"(\\d+)")
        .cast(pl.Int64, strict=False)
    )

    # ページ数は (end - start + 1) とする（例: start=1,end=2 -> 2ページ）
    # 実際の数値が無い場合は NULL になり、下で 0 に丸める
    hw_page_count_expr = hw_end_page_expr - hw_start_page_expr + 1
    # 開始/終了から計算したページ数を適切に丸める
    # - NULL や負数は 0
    # - 上限は page_limit
    # NULL や負数を 0 に、page_limit を超える場合は page_limit に切り詰める
    capped_hw_page_count_expr = (
        pl.when(hw_page_count_expr.is_null())
        .then(0)
        .when(hw_page_count_expr < 0)
        .then(0)
        .when(hw_page_count_expr > page_limit)
        .then(page_limit)
        .otherwise(hw_page_count_expr)
    )

    # 宿題名が空でない場合のみページ数を評価する。除外語に該当する場合は 0。
    # 戻り値の意味: この pl.Expr は1行分の宿題ページ数を返す（整数）。
    return (
        pl.when(hw_name_expr != "")
        .then(
            pl.when(~is_excluded_hw_expr).then(capped_hw_page_count_expr).otherwise(0)
        )
        .otherwise(0)
    )


def build_test_presence_exprs(
    lap_column_names: list[str], kaku_column_names: list[str]
) -> tuple[pl.Expr, pl.Expr, pl.Expr]:
    """
    ラップ／確認（kaku）列群を受け取り、それぞれ「該当する列のいずれかに値があるか」を示す
    `pl.Expr` を返します。

    - 空文字やNULLは存在しないものとみなします。
    - 戻り値は `(has_lap_expr, has_kaku_expr, has_test_expr)` の3つの式。
    """

    def _presence_expr(column_names: list[str]) -> pl.Expr:
        # 各列について:
        # - NULL ではない
        # - 文字列としてトリム後に空でない
        # という条件を作る（存在判定）。
        presence_checks = [
            pl.col(column_name).is_not_null()
            & (pl.col(column_name).cast(pl.Utf8).str.strip_chars() != "")
            for column_name in column_names
        ]

        # 列群のどれか一つでも存在すれば True とする（横持ち判定）。
        # 列が一つも指定されていない場合は常に False を返す。
        return (
            pl.any_horizontal(presence_checks).fill_null(False)
            if presence_checks
            else pl.lit(False)
        )

    has_lap_expr = _presence_expr(lap_column_names)
    has_kaku_expr = _presence_expr(kaku_column_names)
    # ラップまたは確認のどちらかが存在すればテストありとみなす
    has_test_expr = (has_lap_expr | has_kaku_expr).fill_null(False)
    return has_lap_expr, has_kaku_expr, has_test_expr


# Excelシートを読み込む関数。fastexcelで高速に読み込み、ヘッダー処理も行う。
# 失敗した場合はopenpyxlでフォールバックする。
def load_sheet_preserve_extra_columns(path: str, sheet_index: int = 1) -> pl.DataFrame:
    # fastexcelで読み込み。Polarsネイティブでヘッダー処理を行う。
    try:
        import fastexcel

        excel_reader = fastexcel.read_excel(path)
        sheet = excel_reader.load_sheet(sheet_index - 1, header_row=None)
        df_raw = sheet.to_polars()

        if df_raw.height == 0:
            return pl.DataFrame()

        progress(
            "fastexcelエンジンによる読み込み完了。Polarsネイティブでヘッダー処理を開始します。"
        )

        # 最初の行をヘッダーとして扱う（Noneは空文字に変換）
        header = [str(x) if x is not None else "" for x in df_raw.row(0)]

        # "宿題名" を見つけたら、その位置から以降の列を
        # "宿題名", "宿題開始ページ", "宿題終了ページ" のパターンで補完する。
        # これは、Excelの列が3列セットで繰り返される仕様を想定した処理です。
        hw_base_idx = -1
        for i, h in enumerate(header):
            if "宿題名" in h:
                hw_base_idx = i
                break

        if hw_base_idx != -1:
            # 以降の空白列をパターン名で埋め、3列ごとに番号を増やす
            pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
            k = 2
            pattern_idx = 0
            for i in range(hw_base_idx + 1, len(header)):
                if header[i].strip() == "":
                    header[i] = f"{pattern[pattern_idx]}_{k}"
                    pattern_idx += 1
                    if pattern_idx >= 3:
                        pattern_idx = 0
                        k += 1
        else:
            # 宿題群が見つからない場合は単純に空白列に追加列_N を割り当てる
            idx = 1
            for i in range(len(header)):
                if header[i].strip() == "":
                    header[i] = f"追加列_{idx}"
                    idx += 1

        unique_headers = []
        seen = set()
        for h in header:
            h = h.strip()
            if h == "":
                h = "無名列"
            original_h = h
            counter = 1
            while h in seen:
                h = f"{original_h}_{counter}"
                counter += 1
            seen.add(h)
            unique_headers.append(h)

        df_data = df_raw.slice(1)
        df_data = df_data.rename(dict(zip(df_data.columns, unique_headers)))

        progress(f"データ構築完了: 行数={df_data.height}, 列数={df_data.width}")
        return df_data

    # openpyxlで読み込み。
    except Exception as e:
        progress(f"高速エンジンでの読み込み失敗。通常モードに切り替えます: {e}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[sheet_index - 1]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not rows:
            return pl.DataFrame()

        header = [str(c) if c is not None else "" for c in rows[0]]
        max_cols = max(len(r) for r in rows)

        if max_cols > len(header):
            try:
                base = header.index("宿題名")
                pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
                k = 2
                while len(header) < max_cols:
                    for p in pattern:
                        header.append(f"{p}_{k}")
                        if len(header) >= max_cols:
                            break
                    k += 1
            except ValueError:
                idx = 1
                while len(header) < max_cols:
                    header.append(f"追加列_{idx}")
                    idx += 1

        header_len = len(header)
        padded_rows = [
            (
                tuple(row) + (None,) * (header_len - len(row))
                if len(row) < header_len
                else tuple(row)
            )
            for row in rows[1:]
        ]

        df = pl.DataFrame(padded_rows, schema=header, orient="row")
        return df


# 日付から日本語の曜日を取得する関数
def get_jp_weekday(dt_obj):
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    return weekdays[dt_obj.weekday()]


# メインのデータ処理関数。
# 入力ファイルを読み込み、必要な指標を計算し、集計用のDataFrameを返す。
def process_attendance_data(
    input_file: str,
) -> tuple[
    pl.DataFrame,
    pl.DataFrame,
    pl.DataFrame,
    pl.DataFrame,
    str,
    str,
    str,
    list[str],
    list[str],
]:
    df = load_sheet_preserve_extra_columns(input_file, sheet_index=1)
    cols = df.columns

    # ----------------------------
    # 宿題名列の一括正規化（外部ライブラリ許可済み）
    # - 各 "宿題名" 列について正規化された新しい列 "<元名>__norm" を追加する
    # - 正規化は Python レベルで一括処理してから DataFrame に戻す（Polars の UDF より高速な場合が多い）
    hw_name_cols = [c for c in cols if c and "宿題名" in c]
    if hw_name_cols:
        norm_series_list = []
        for c in hw_name_cols:
            try:
                orig_vals = df.get_column(c).to_list()
            except Exception:
                orig_vals = []

            # 並列化して正規化を実行（ワーカーは CPU-1 をデフォルト）
            try:
                # ワーカー数 / チャンクサイズは環境変数でオーバーライド可能
                # NORMALIZE_WORKERS: 正の整数でワーカー数（0 または未指定で自動設定）
                # NORMALIZE_CHUNKSIZE: 正の整数で chunksize（0 または未指定で自動計算）
                env_workers = os.environ.get("NORMALIZE_WORKERS")
                env_chunks = os.environ.get("NORMALIZE_CHUNKSIZE")

                if env_workers and env_workers.isdigit():
                    cpu = max(1, int(env_workers))
                else:
                    cpu = max(1, DEFAULT_NORMALIZE_WORKERS)

                if env_chunks and env_chunks.isdigit():
                    chunksize = max(1, int(env_chunks))
                else:
                    # デフォルトはベンチで得られた固定チャンクサイズ
                    chunksize = max(1, DEFAULT_NORMALIZE_CHUNKSIZE)

                with mp.Pool(processes=cpu) as pool:
                    norm_vals = pool.map(
                        normalize_text_for_matching, orig_vals, chunksize
                    )
            except Exception:
                # フォールバック: シリアル処理
                norm_vals = [normalize_text_for_matching(v) for v in orig_vals]

            norm_col_name = f"{c}__norm"
            norm_series_list.append(pl.Series(norm_col_name, norm_vals))
        if norm_series_list:
            df = df.with_columns(norm_series_list)
            cols = df.columns
    # ----------------------------

    possible_date_cols = ["日付", "授業日"]
    date_col: str | None = None
    for dc in possible_date_cols:
        if dc in cols:
            date_col = dc
            break

    progress("日付情報からファイル名・シート名を生成中...")
    if date_col is not None:
        valid_dates = df.select(
            pl.col(date_col).cast(pl.Date, strict=False)
        ).drop_nulls()
    else:
        valid_dates = pl.DataFrame()

    if valid_dates.height > 0:
        min_dt = valid_dates.min().item()
        max_dt = valid_dates.max().item()

        wd_min = get_jp_weekday(min_dt)
        wd_max = get_jp_weekday(max_dt)

        min_date_file = min_dt.strftime("%Y%m%d")
        max_date_file = max_dt.strftime("%Y%m%d")

        date_range_str = f"集計期間: {min_dt.strftime('%Y/%m/%d')}({wd_min}) ～ {max_dt.strftime('%Y/%m/%d')}({wd_max})"
    else:
        min_date_file, max_date_file = "不明", "不明"
        date_range_str = "集計期間: 不明"

    sheet_name = "集計結果"
    now_str = datetime.now().strftime("%y%m%d-%H%M%S")
    input_dir = os.path.dirname(input_file)
    # 出力ファイル名は集計期間と現在時刻を含めて一意化する
    # 例: 集計結果_[20230101-20230107]_230101-120000.xlsx
    output_filename = f"集計結果_[{min_date_file}-{max_date_file}]_{now_str}.xlsx"
    output_file = os.path.join(input_dir, output_filename)

    # ここから列名を走査して各セクションの開始インデックスを検出する
    progress("列インデックスの特定中...")
    if date_col is not None:
        dt_expr = pl.col(date_col).cast(pl.Date).dt
        dt_iso = cast(Any, dt_expr).week()
    else:
        dt_iso = pl.lit(None)

    # ラップ関連の列群の先頭位置を探す（存在しなければ 0 のまま）
    lap_section_start_index = 0
    for i, c in enumerate(cols):
        if c and ("ラップテスト" in c or "ラップ" in c):
            # 最初にラップ関連の列が現れた位置を記録してループ終了
            lap_section_start_index = i
            break

    # 宿題セクションの先頭列を探す。見つからなければ末尾扱いにする。
    hw_section_start_index = len(cols)
    for i, c in enumerate(cols):
        if c and "宿題名" in c:
            hw_section_start_index = i
            break

    # ここまでで各セクション（ラップ開始位置、宿題開始位置）を検出しました。
    # 以降、宿題は 3 列セット（宿題名, 開始, 終了）が繰り返されている想定で処理します。
    progress(
        f"インデックス特定完了: ラップ開始={lap_section_start_index}, 宿題開始={hw_section_start_index}"
    )

    # テスト関連の列群はラップ開始位置から宿題開始位置の手前まで
    test_section_column_names = cols[lap_section_start_index:hw_section_start_index]
    # ラップ回数をカウントする列（分子表記の列）を抽出
    lap_count_column_names = [
        c for c in test_section_column_names if "ラップ" in c and "（分子）" in c
    ]
    # 確認（kaku）をカウントする列（分子表記の列）を抽出
    confirmation_count_column_names = [
        c for c in test_section_column_names if "確認" in c and "（分子）" in c
    ]

    hw_page_exprs = []
    for i in range(hw_section_start_index, len(cols), 3):
        if i + 2 < len(cols):
            hw_name_col = cols[i]
            hw_start_col = cols[i + 1]
            hw_end_col = cols[i + 2]

            # 各宿題トリプルについて、行単位でのページ数を返す pl.Expr を生成しリストに追加
            # 生成された式群は後で横持ちで合計され、1行の宿題合計ページ数（hw_pages）となる
            # 正規化済みの列があればそちらを優先して式に渡す
            norm_col = f"{hw_name_col}__norm"
            hw_name_for_expr = norm_col if norm_col in cols else hw_name_col

            hw_page_exprs.append(
                build_hw_pages_expr(
                    hw_name_for_expr,
                    hw_start_col,
                    hw_end_col,
                    EXCLUDED_HW_TERMS,
                )
            )

    progress("データのクレンジングと指標の計算中...")

    missing_required_columns = []
    for req_col in ["出欠", "担当講師名", "担当講師N0", "教室"]:
        # 必須列が欠けていると後続処理で失敗するため事前にチェック
        if req_col not in cols:
            missing_required_columns.append(req_col)

    if missing_required_columns:
        raise ValueError(
            f"必須の列が見つかりません: {', '.join(missing_required_columns)}"
        )

    # 出席のみ抽出して指標列を追加する
    attendance_metrics_df = (
        df.filter(pl.col("出欠") == "出席").with_columns(
            # 担当講師名から全角英大文字およびアンダースコアを除去（表示用のクリーンアップ）
            pl.col("担当講師名")
            .cast(pl.Utf8)
            .str.replace_all(r"[Ａ-Ｚ_]", "")
            .alias("担当講師名"),
            # 週番号を示す式（Polars の dt.week() を事前に dt_iso として構築）
            dt_iso.alias("week_num"),
            # 元の日付列を Date 型に変換して保持（週範囲の判定に使う）
            (
                pl.col(date_col).cast(pl.Date, strict=False)
                if date_col
                else pl.lit(None)
            ).alias("date_val"),
            # 各宿題ごとに作成した pl.Expr を横に合計してその日の宿題ページ合計を得る
            (pl.sum_horizontal(hw_page_exprs) if hw_page_exprs else pl.lit(0)).alias(
                "hw_pages"
            ),
        )
        # ラップ／確認（kaku）列群から存在判定式（has_lap, has_kaku, has_test）を追加
        .with_columns(
            *build_test_presence_exprs(
                lap_count_column_names, confirmation_count_column_names
            )
        )
    )

    week_period_by_week_number = {}
    if "date_val" in attendance_metrics_df.columns:
        week_ranges_df = (
            attendance_metrics_df.filter(pl.col("date_val").is_not_null())
            .group_by("week_num")
            .agg(
                pl.col("date_val").min().alias("min_dt"),
                pl.col("date_val").max().alias("max_dt"),
            )
        )
        # 各週番号ごとに該当週の開始日・終了日を文字列化して辞書に保持
        for row in week_ranges_df.iter_rows(named=True):
            w_num = row["week_num"]
            md = row["min_dt"]
            xd = row["max_dt"]
            if md and xd:
                week_period_by_week_number[str(w_num)] = (
                    f"{md.strftime('%Y/%m/%d')}({get_jp_weekday(md)}) ～ {xd.strftime('%Y/%m/%d')}({get_jp_weekday(xd)})"
                )

    # 共通処理: 集計→ピボットを行う内部関数
    def aggregate_and_pivot(attendance_metrics_local_df: pl.DataFrame):
        progress("週次集計中...")
        week_period_by_week_number_local = {}
        if "date_val" in attendance_metrics_local_df.columns:
            week_ranges_df_local = (
                attendance_metrics_local_df.filter(pl.col("date_val").is_not_null())
                .group_by("week_num")
                .agg(
                    pl.col("date_val").min().alias("min_dt"),
                    pl.col("date_val").max().alias("max_dt"),
                )
            )
            for row in week_ranges_df_local.iter_rows(named=True):
                w_num = row["week_num"]
                md = row["min_dt"]
                xd = row["max_dt"]
                if md and xd:
                    week_period_by_week_number_local[str(w_num)] = (
                        f"{md.strftime('%Y/%m/%d')}({get_jp_weekday(md)}) ～ {xd.strftime('%Y/%m/%d')}({get_jp_weekday(xd)})"
                    )

        # グルーピング: 講師N0/講師名/教室/週ごとに集計し、授業数・宿題合計・Lap/テスト回数を求める
        base_aggregate_df_local = attendance_metrics_local_df.group_by(
            ["担当講師N0", "担当講師名", "教室", "week_num"]
        ).agg(
            pl.len().alias("授業数"),
            pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"),
            pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数"),
        )

        aggregated_df_local = base_aggregate_df_local.with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率"),
        ).select(
            [
                "担当講師N0",
                "担当講師名",
                "教室",
                "week_num",
                "授業数",
                "宿題平均",
                "ラップ回数",
                "ラップ率",
                "テスト回数",
                "テスト率",
            ]
        )

        progress("横持ちへのピボット処理...")
        # 横持ち変換: week_num を列ヘッダーにして週ごとの値を横に並べる
        pivot_df_local = aggregated_df_local.pivot(
            on="week_num",
            index=["担当講師N0", "担当講師名", "教室"],
            values=[
                "授業数",
                "宿題平均",
                "ラップ回数",
                "ラップ率",
                "テスト回数",
                "テスト率",
            ],
        )

        progress("希望の列順への並び替えとソート処理中...")
        # pivot 後の列名から週番号部分を取り出してソートしたリストを作成
        week_numbers_local = sorted(
            list(set([c.split("_")[-1] for c in pivot_df_local.columns if "_" in c]))
        )

        # 総授業数を計算
        # 各週の授業数列を横合計して総授業数列を作る
        class_count_columns_local = [f"授業数_{w}" for w in week_numbers_local]
        total_class_expr_local = pl.sum_horizontal(
            [
                pl.col(c)
                for c in class_count_columns_local
                if c in pivot_df_local.columns
            ]
        ).fill_null(0)
        pivot_df_local = pivot_df_local.with_columns(
            total_class_expr_local.alias("総授業数")
        )

        desired_column_order_local = ["担当講師N0", "担当講師名", "教室", "総授業数"]
        rename_map_local = {
            "担当講師N0": "No",
            "担当講師名": "氏名",
            "教室": "教室",
            "総授業数": "総授業数",
        }

        week_period_labels_local = []
        # 各週番号について、出力列名（授業1, 平均HW1 等）とラベル（週期間文字列）を作る
        for i, w in enumerate(week_numbers_local, start=1):
            c_class = f"授業数_{w}"
            c_hw = f"宿題平均_{w}"
            c_lap_cnt = f"ラップ回数_{w}"
            c_lap_rate = f"ラップ率_{w}"
            c_test_cnt = f"テスト回数_{w}"
            c_test_rate = f"テスト率_{w}"

            desired_column_order_local.extend(
                [c_class, c_hw, c_lap_cnt, c_lap_rate, c_test_cnt, c_test_rate]
            )

            rename_map_local[c_class] = f"授業{i}"
            rename_map_local[c_hw] = f"平均HW{i}"
            rename_map_local[c_lap_cnt] = f"Lap数{i}"
            rename_map_local[c_lap_rate] = f"Lap％{i}"
            rename_map_local[c_test_cnt] = f"テスト数{i}"
            rename_map_local[c_test_rate] = f"テスト％{i}"

            # 週期間ラベルは存在しない場合はフォールバック文言を使う
            week_period_labels_local.append(
                week_period_by_week_number_local.get(str(w), f"Week {i}")
            )

        existing_desired_columns_local = [
            c for c in desired_column_order_local if c in pivot_df_local.columns
        ]

        pivot_df_local = (
            pivot_df_local.select(existing_desired_columns_local)
            .rename(
                {
                    k: v
                    for k, v in rename_map_local.items()
                    if k in existing_desired_columns_local
                }
            )
            .sort(["教室", "No"])
        )

        return pivot_df_local, week_period_labels_local

    # フル集計
    pivot_df_all, week_period_labels_all = aggregate_and_pivot(attendance_metrics_df)

    # 除外条件: 教室に高3/高３/高卒 を含む、または 科目欄(もしあれば)が国語 を含む行を除外
    subject_column_name = None
    for c in cols:
        if c in ("科目", "教科"):
            subject_column_name = c
            break

    classroom_exclusion_condition = (
        pl.col("教室").cast(pl.Utf8).str.contains(r"(高3|高３|高卒)").fill_null(False)
    )
    subject_exclusion_condition = (
        (
            pl.col(subject_column_name)
            .cast(pl.Utf8)
            .str.contains(r"国語")
            .fill_null(False)
        )
        if subject_column_name
        else pl.lit(False)
    )

    excluded_metrics_df = attendance_metrics_df.filter(
        ~classroom_exclusion_condition & ~subject_exclusion_condition
    )

    excluded_pivot_df, week_period_labels_excl = aggregate_and_pivot(
        excluded_metrics_df
    )

    # 教室ごとの総括（教室レベル集計）
    classroom_summary_df = (
        attendance_metrics_df.group_by("教室")
        .agg(
            pl.len().alias("授業数"),
            pl.col("担当講師N0").n_unique().alias("講師数"),
            pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"),
            pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数"),
        )
        .with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率"),
        )
        .sort("教室")
    )

    # 除外版の教室サマリー（Lap除外など）
    excluded_classroom_summary_df = (
        excluded_metrics_df.group_by("教室")
        .agg(
            pl.len().alias("授業数"),
            pl.col("担当講師N0").n_unique().alias("講師数"),
            pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"),
            pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数"),
        )
        .with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率"),
        )
        .sort("教室")
    )

    return (
        pivot_df_all,
        excluded_pivot_df,
        classroom_summary_df,
        excluded_classroom_summary_df,
        output_file,
        date_range_str,
        sheet_name,
        week_period_labels_all,
        week_period_labels_excl,
    )


# Excelシートへの書き出し関数。xlsxwriterで高速に書き出し、見やすい書式を適用する。
def format_excel_fast(
    all_pivot_df: pl.DataFrame,
    excluded_pivot_df: pl.DataFrame,
    classroom_summary_df: pl.DataFrame,
    excluded_classroom_summary_df: pl.DataFrame,
    output_file: str,
    date_range_str: str,
    sheet_name: str,
    week_period_labels_all: list[str],
    week_period_labels_excl: list[str],
):
    progress("xlsxwriterによる超高速書き出し・書式設定中...")

    workbook = xlsxwriter.Workbook(output_file)

    # Windows標準の見やすいUDフォントを設定
    ud_font = "Noto Sans JP"

    # 全ての書式に UDフォント を適用
    fmt_title = workbook.add_format(
        {
            "font_name": ud_font,
            "bold": True,
            "font_size": 11,
            "align": "left",
            "valign": "vcenter",
        }
    )
    fmt_title_center = workbook.add_format(
        {
            "font_name": ud_font,
            "bold": True,
            "font_size": 11,
            "align": "center",
            "valign": "vcenter",
            "bg_color": "#E8F6F3",
            "border": 1,
        }
    )

    fmt_header = workbook.add_format(
        {"font_name": ud_font, "border": 1, "bg_color": "#D3D3D3", "bold": True}
    )
    fmt_white = workbook.add_format({"font_name": ud_font, "border": 1})
    fmt_gray = workbook.add_format(
        {"font_name": ud_font, "border": 1, "bg_color": "#F5F5F5"}
    )
    fmt_khaki = workbook.add_format(
        {"font_name": ud_font, "border": 1, "bg_color": "#F0E68C"}
    )
    fmt_lavender = workbook.add_format(
        {"font_name": ud_font, "border": 1, "bg_color": "#E6E6FA"}
    )

    fmt_white_pct = workbook.add_format(
        {"font_name": ud_font, "border": 1, "num_format": "0.0%"}
    )
    fmt_gray_pct = workbook.add_format(
        {"font_name": ud_font, "border": 1, "bg_color": "#F5F5F5", "num_format": "0.0%"}
    )

    fmt_white_center = workbook.add_format(
        {"font_name": ud_font, "border": 1, "align": "center"}
    )
    fmt_gray_center = workbook.add_format(
        {"font_name": ud_font, "border": 1, "bg_color": "#F5F5F5", "align": "center"}
    )

    fmt_cond_orange = workbook.add_format(
        {"font_name": ud_font, "bg_color": "#FFA500", "font_color": "#000000"}
    )
    fmt_cond_firebrick = workbook.add_format(
        {"font_name": ud_font, "bg_color": "#B22222", "font_color": "#FFFFFF"}
    )

    def make_safe_sheet_name(raw_name: str, used_names: set[str]) -> str:
        name = str(raw_name).strip() or "無名教室"
        translate_table = str.maketrans(
            {
                "/": "／",
                "\\": "＼",
                "?": "？",
                "*": "＊",
                "[": "［",
                "]": "］",
                ":": "：",
            }
        )
        name = name.translate(translate_table)
        if name.startswith("'"):
            name = name[1:]
        if name.endswith("'"):
            name = name[:-1]
        name = name[:31] or "無名教室"

        base_name = name
        counter = 2
        while name in used_names:
            suffix = f"_{counter}"
            name = f"{base_name[:31 - len(suffix)]}{suffix}"
            counter += 1

        used_names.add(name)
        return name

    # シートごとに書き込む関数。
    # ヘッダーの書式設定、データの書式設定、カスタムオートフィットを行う。
    def write_sheet(
        worksheet,
        sheet_df: pl.DataFrame,
        week_period_labels: list[str],
        sheet_title: str,
        summary_text: str | None = None,
    ):
        progress(f"{sheet_title} の書き込みを開始します...")
        # 出力に使うヘッダーとデータ行ジェネレータを取得
        headers = sheet_df.columns
        max_col = len(headers) - 1
        # .rows() はイテレータではなくリスト状の反復可能オブジェクトを返す
        data_rows = sheet_df.rows()

        # header_row / data_start を summaryの有無で調整
        header_row = 1
        if summary_text:
            # 日付タイトルが row 0 に入るため、その下に summary を入れる
            header_row = 2

        data_start = header_row + 1

        # 1. summary（ある場合）は header_row-1 に表示
        if summary_text:
            try:
                worksheet.merge_range(
                    header_row - 1, 0, header_row - 1, max_col, summary_text, fmt_title
                )
            except Exception:
                worksheet.write(header_row - 1, 0, summary_text, fmt_title)

        # 2. ヘッダーを書き込む
        for col_num, col_name in enumerate(headers):
            worksheet.write(header_row, col_num, col_name, fmt_header)

        # カスタムオートフィット用の幅計算変数
        max_a_width = 4  # Noの最低幅
        max_b_width = 12  # 氏名の最低幅
        max_c_width = 10  # 教室の最低幅

        # 3. data_start 行からデータを書き込む
        #    同時に A/B/C 列の幅（No/氏名/教室）を文字幅で自動調整する
        for ri, row_data in enumerate(data_rows, start=0):
            row_num = data_start + ri
            # 進捗ログ: 最初の行と区切り毎に報告
            if ri == 0 or (ri + 1) % 1000 == 0:
                progress(f"{sheet_title}: {ri + 1}/{len(data_rows)} 行を書き込み中...")

            # Excel 偶数行かどうか（行の背景色決定に使用）
            is_even_excel_row = (row_num + 1) % 2 == 0

            # A列(No), B列(氏名), C列(教室) の文字幅を計算して記録
            # 日本語は幅が大きくなるため east_asian_width を使っておおよその幅を計算
            val_a = row_data[0]
            if val_a is not None:
                w_a = sum(
                    2.2 if unicodedata.east_asian_width(c) in ("F", "W", "A") else 1.2
                    for c in str(val_a)
                )
                if w_a > max_a_width:
                    max_a_width = w_a

            val_b = row_data[1]
            if val_b is not None:
                w_b = sum(
                    2.2 if unicodedata.east_asian_width(c) in ("F", "W", "A") else 1.2
                    for c in str(val_b)
                )
                if w_b > max_b_width:
                    max_b_width = w_b

            val_c = row_data[2]
            if val_c is not None:
                w_c = sum(
                    2.2 if unicodedata.east_asian_width(c) in ("F", "W", "A") else 1.2
                    for c in str(val_c)
                )
                if w_c > max_c_width:
                    max_c_width = w_c

            # 各セルを書き込むループ
            for col_num, cell_value in enumerate(row_data):
                # A〜C列は固定フォーマットで出力
                if col_num < 3:
                    base_fmt = fmt_gray if is_even_excel_row else fmt_white
                    if cell_value is None or (
                        isinstance(cell_value, float) and math.isnan(cell_value)
                    ):
                        worksheet.write_blank(row_num, col_num, "", base_fmt)
                    else:
                        worksheet.write(row_num, col_num, cell_value, base_fmt)
                    continue

                # D列（総授業数）
                if col_num == 3:
                    if cell_value is None or (
                        isinstance(cell_value, float) and math.isnan(cell_value)
                    ):
                        worksheet.write_blank(row_num, col_num, "", fmt_lavender)
                    else:
                        worksheet.write(row_num, col_num, cell_value, fmt_lavender)
                    continue

                offset = (col_num - 4) % 6
                c_base = col_num - offset

                class_count = row_data[c_base]
                has_no_class = (
                    (class_count is None)
                    or (isinstance(class_count, float) and math.isnan(class_count))
                    or (class_count == 0)
                )

                if offset == 0:
                    base_fmt = fmt_khaki
                elif offset == 3 or offset == 5:
                    base_fmt = fmt_gray_pct if is_even_excel_row else fmt_white_pct
                else:
                    base_fmt = fmt_gray if is_even_excel_row else fmt_white

                if has_no_class:
                    if offset == 0:
                        worksheet.write(row_num, col_num, 0, base_fmt)
                    else:
                        hyphen_fmt = (
                            fmt_gray_center if is_even_excel_row else fmt_white_center
                        )
                        worksheet.write(row_num, col_num, "-", hyphen_fmt)
                else:
                    if cell_value is None or (
                        isinstance(cell_value, float) and math.isnan(cell_value)
                    ):
                        worksheet.write_blank(row_num, col_num, "", base_fmt)
                    else:
                        worksheet.write(row_num, col_num, cell_value, base_fmt)

        # 4. カスタムオートフィットの適用（UDフォント向けに少し余白を広めにとる）
        worksheet.set_column(0, 0, max_a_width + 3)  # No (計算した幅+余白)
        worksheet.set_column(1, 1, max_b_width + 3)  # 氏名 (計算した幅+余白)
        worksheet.set_column(2, 2, max_c_width + 3)  # 教室 (計算した幅+余白)
        worksheet.set_column(3, 3, 9)  # 総授業数
        if max_col >= 4:
            worksheet.set_column(4, max_col, 9)  # 授業数等の列はスッキリと固定幅

        # 5. 日付などのタイトルを入力（幅設定後に行うことで安全に結合可能）
        worksheet.merge_range(0, 0, 0, 2, date_range_str, fmt_title)
        worksheet.set_row(0, 20)

        for i, period_str in enumerate(week_period_labels):
            start_col = 4 + i * 6
            if start_col + 5 <= max_col:
                worksheet.merge_range(
                    0, start_col, 0, start_col + 5, period_str, fmt_title_center
                )

        # 6. オートフィルターの設定
        last_row = data_start + len(data_rows) - 1
        worksheet.autofilter(header_row, 0, last_row, max_col)

        # 7. 条件付き書式
        for c_base in range(4, len(headers), 6):
            if c_base + 5 > max_col:
                break

            col_hw = c_base + 1
            col_lap = c_base + 3
            col_test = c_base + 5

            # ここでは週次の列群が6列単位で並んでいる想定（授業数, 宿題平均, ..., テスト％）
            # そのため基準列を c_base を基に決めて条件付き書式を適用している
            worksheet.conditional_format(
                data_start,
                col_hw,
                last_row,
                col_hw,
                {
                    "type": "cell",
                    "criteria": "<",
                    "value": 6,
                    "format": fmt_cond_orange,
                },
            )

            # Lap率・テスト率の閾値を下回る場合に目立たせる
            worksheet.conditional_format(
                data_start,
                col_lap,
                last_row,
                col_lap,
                {
                    "type": "cell",
                    "criteria": "<",
                    "value": 0.7,
                    "format": fmt_cond_firebrick,
                },
            )
            worksheet.conditional_format(
                data_start,
                col_test,
                last_row,
                col_test,
                {
                    "type": "cell",
                    "criteria": "<",
                    "value": 0.7,
                    "format": fmt_cond_firebrick,
                },
            )

        worksheet.freeze_panes(data_start, 3)
        progress(f"{sheet_title} の書き込みが完了しました。")

    # 全集計シート
    ws_all = workbook.add_worksheet("全集計")
    write_sheet(ws_all, all_pivot_df, week_period_labels_all, "全集計")

    # 除外集計シート
    ws_excl = workbook.add_worksheet("除外集計")
    write_sheet(ws_excl, excluded_pivot_df, week_period_labels_excl, "除外集計")

    # 全教室サマリーシート（通常版と除外版を結合して表示）
    ws_allclass = workbook.add_worksheet("全教室")

    # classroom summary は classroom_summary_df / excluded_classroom_summary_df
    def write_simple_sheet(
        worksheet, summary_df: pl.DataFrame, sheet_title: str, date_title: str
    ):
        # タイトル
        try:
            worksheet.merge_range(
                0, 0, 0, max(0, summary_df.width - 1), date_title, fmt_title
            )
        except Exception:
            worksheet.write(0, 0, date_title, fmt_title)
        worksheet.set_row(0, 20)

        # ヘッダー
        for c, col_name in enumerate(summary_df.columns):
            worksheet.write(1, c, col_name, fmt_header)

        # データ行
        for ri, row in enumerate(summary_df.rows(), start=2):
            is_even = ri % 2 == 0
            for c, v in enumerate(row):
                if isinstance(v, float) and math.isnan(v):
                    v = None
                # パーセント列名を判定
                colname = summary_df.columns[c]
                if "率" in colname:
                    fmt = fmt_gray_pct if is_even else fmt_white_pct
                    if v is None:
                        worksheet.write_blank(ri, c, "", fmt)
                    else:
                        worksheet.write(ri, c, v, fmt)
                else:
                    fmt = fmt_gray if is_even else fmt_white
                    if v is None:
                        worksheet.write_blank(ri, c, "", fmt)
                    else:
                        worksheet.write(ri, c, v, fmt)

        # 幅調整: 簡易的
        for c in range(summary_df.width):
            worksheet.set_column(c, c, 12)

        worksheet.autofilter(1, 0, 1 + summary_df.height, max(0, summary_df.width - 1))
        worksheet.freeze_panes(2, 0)

    # マージして左右に除外版の列を並べる
    try:
        merged = classroom_summary_df.join(
            excluded_classroom_summary_df, on="教室", how="outer", suffix="_除外"
        )
    except Exception:
        # フォールバック: 単独表示
        merged = classroom_summary_df

    # 列表示順: 教室, 授業数, 講師数, 宿題平均, ラップ率, テスト率, （除外版）...
    preferred = [
        "教室",
        "授業数",
        "講師数",
        "宿題平均",
        "ラップ率",
        "テスト率",
        "授業数_除外",
        "講師数_除外",
        "宿題平均_除外",
        "ラップ率_除外",
        "テスト率_除外",
    ]
    cols_to_show = [c for c in preferred if c in merged.columns]
    if "教室" not in cols_to_show:
        cols_to_show = merged.columns

    merged = merged.select(cols_to_show)

    write_simple_sheet(ws_allclass, merged, "全教室", date_range_str)

    # 教室別サマリーシート
    used_sheet_names = {"全集計", "除外集計"}
    classroom_values = []
    seen_classrooms = set()
    for value in all_pivot_df.get_column("教室").to_list():
        key = "" if value is None else str(value)
        if key in seen_classrooms:
            continue
        seen_classrooms.add(key)
        classroom_values.append(value)

    progress(f"教室別シートを作成中... 対象教室数={len(classroom_values)}")
    for classroom_value in classroom_values:
        classroom_label = (
            "無名教室"
            if classroom_value is None
            else str(classroom_value).strip() or "無名教室"
        )
        sheet_label = make_safe_sheet_name(classroom_label, used_sheet_names)

        if classroom_value is None:
            classroom_df = all_pivot_df.filter(pl.col("教室").is_null())
        else:
            classroom_df = all_pivot_df.filter(
                pl.col("教室").cast(pl.Utf8).str.strip_chars() == classroom_label
            )

        if classroom_df.height == 0:
            continue

        progress(f"教室シート作成: {classroom_label} -> {sheet_label}")
        ws_classroom = workbook.add_worksheet(sheet_label)
        # 教室の集計値を取得してトップに表示する文字列を作る
        summary_row = None
        try:
            if classroom_value is None:
                summary_row = classroom_summary_df.filter(pl.col("教室").is_null())
            else:
                summary_row = classroom_summary_df.filter(
                    pl.col("教室").cast(pl.Utf8).str.strip_chars() == classroom_label
                )
            if summary_row.height == 1:
                r = summary_row.row(0)
                total_classes = int(r[1]) if r[1] is not None else 0
                instructor_count = int(r[2]) if r[2] is not None else 0
                hw_avg = float(r[6]) if r[6] is not None else 0.0
                lap_rate = float(r[7]) if r[7] is not None else 0.0
                test_rate = float(r[8]) if r[8] is not None else 0.0
                top_summary = f"教室: {classroom_label}  総授業数: {total_classes}  講師数: {instructor_count}  宿題平均: {hw_avg:.1f}  Lap率: {lap_rate:.1%}  テスト率: {test_rate:.1%}"
            else:
                top_summary = f"教室: {classroom_label}"
        except Exception:
            top_summary = f"教室: {classroom_label}"

        write_sheet(
            ws_classroom,
            classroom_df,
            week_period_labels_all,
            classroom_label,
            top_summary,
        )

    workbook.close()
    progress("Excelファイルの保存が完了しました！")


# エラー表示用のダイアログ関数。
# Tkinterでシンプルなウィンドウを作り、エラーメッセージを表示する。
def show_error_dialog(title: str, message: str, parent: tk.Tk | None = None) -> None:
    dlg = tk.Toplevel(parent) if parent else tk.Toplevel()
    dlg.title(title)

    try:
        if parent:
            scale = float(parent.tk.call("tk", "scaling"))
        else:
            scale = float(dlg.tk.call("tk", "scaling"))
    except Exception:
        scale = 1.0

    base_font = tkfont.nametofont("TkDefaultFont")
    try:
        fsize = max(int(base_font.cget("size") * scale), 10)
    except Exception:
        fsize = 11
    dlg_font = tkfont.Font(family=base_font.cget("family"), size=fsize)

    dlg.geometry(f"{int(700*scale)}x{int(360*scale)}")

    txt = tk.Text(dlg, wrap="word", font=dlg_font)
    txt.insert("1.0", message)
    txt.configure(state="disabled")
    txt.pack(expand=True, fill="both", padx=int(6 * scale), pady=int(6 * scale))

    frm = tk.Frame(dlg)
    frm.pack(fill="x", padx=int(6 * scale), pady=int(6 * scale))

    def _copy():
        try:
            dlg.clipboard_clear()
            dlg.clipboard_append(message)
        except Exception:
            pass

    btn_copy = tk.Button(
        frm, text="エラーメッセージをコピー", command=_copy, font=dlg_font
    )
    btn_copy.pack(side="left")

    btn_close = tk.Button(frm, text="閉じる", command=dlg.destroy, font=dlg_font)
    btn_close.pack(side="right")

    try:
        dlg.transient(parent)
        dlg.grab_set()
        dlg.wait_window()
    except Exception:
        pass


def main():
    enable_windows_dpi_awareness()
    root = tk.Tk()
    configure_tk_scaling(root)
    root.withdraw()
    try:
        # ダイアログが背面に隠れないように強制的に最前面に持ってくる
        root.attributes("-topmost", True)
        root.lift()
        root.focus_force()

        print("集計元のExcelファイルを選択してください...")
        input_file = filedialog.askopenfilename(
            title="集計元のExcelファイルを選択してください",
            filetypes=[
                ("Excelファイル", "*.xlsx *.xls *.xlsm"),
                ("すべてのファイル", "*.*"),
            ],
            parent=root,
        )

        root.attributes("-topmost", False)

        if not input_file:
            print("ファイルの選択がキャンセルされました。")
            return

        total_start_time = time.perf_counter()

        (
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
        ) = process_attendance_data(input_file)

        format_excel_fast(
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
        )

        total_elapsed = time.perf_counter() - total_start_time

        result_msg = (
            f"処理が完了しました！\n"
            f"実行時間: {total_elapsed:.2f}秒\n\n"
            f"出力先:\n{output_file}"
        )

        root.attributes("-topmost", True)
        messagebox.showinfo("集計完了", result_msg, parent=root)
        root.attributes("-topmost", False)

        print(f"すべての処理が完了しました！ (実行時間: {total_elapsed:.2f}秒)")

    except KeyboardInterrupt:
        print("処理を中断しました。")
    except Exception as e:
        tb = traceback.format_exc()
        print("\n=== エラー詳細 ===")
        print(tb)
        print("==================\n")
        try:
            root.attributes("-topmost", True)
            show_error_dialog("処理中にエラーが発生しました", tb, parent=root)
        except Exception:
            messagebox.showerror(
                "エラー", f"処理中にエラーが発生しました:\n{e}", parent=root
            )
    finally:
        root.destroy()


if __name__ == "__main__":
    main()
