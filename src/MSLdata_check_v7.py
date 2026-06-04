import polars as pl
import tkinter as tk
from tkinter import filedialog, messagebox
import xlsxwriter
import openpyxl
from openpyxl.worksheet.pagebreak import ColBreak
import os
import pandas as pd
from datetime import datetime
from typing import Any, cast
import re
import unicodedata
import sys
import ctypes
import tkinter.font as tkfont
import traceback
import time
import gc


def progress(msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    print(f"[{ts}] {msg}")

EXCLUDED_HW_TERMS = [
    "タンゴ",
    "単語",
    "シスタン",
    "シス単",
    "ターゲット",
    "リープ",
    "leap",
    "ゲットスルー",
    "パスタン",
    "パス単",
    "マドンナ",
]

FONT_PREFERRED = [
    "Noto Sans JP",
    "Arial Unicode MS",
    "Arial",
]


def select_font(preferred: list[str] | None = None) -> str:
    prefs = preferred or FONT_PREFERRED
    try:
        root_tmp = tk.Tk()
        root_tmp.withdraw()
        fams = set(tkfont.families())
        root_tmp.destroy()
        for f in prefs:
            if f in fams:
                return f
    except Exception:
        pass

    for f in prefs:
        return f
    return "Calibri"

def normalize_text_for_matching(s: object) -> str:
    if s is None:
        return ""
    t = str(s)
    t = unicodedata.normalize("NFKC", t)
    t = t.lower()
    t = re.sub(r"\s+", "", t)
    t = "".join(chr(ord(c) + 96) if "ぁ" <= c <= "ん" else c for c in t)
    return t


def enable_windows_dpi_awareness() -> None:
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def build_hw_pages_expr(
    hw_name_column: str,
    hw_start_column: str,
    hw_end_column: str,
    excluded_terms: list[str],
    page_limit: int = 30,
) -> pl.Expr:
    hw_name_expr = pl.col(hw_name_column).cast(pl.Utf8).fill_null("").str.strip_chars()
    active_excluded_hw_terms = [term for term in excluded_terms if str(term).strip()]

    is_excluded_hw_expr = (
        pl.any_horizontal(
            [hw_name_expr.str.contains(term, literal=True) for term in active_excluded_hw_terms]
        )
        if active_excluded_hw_terms
        else pl.lit(False)
    )

    hw_start_page_expr = (
        pl.col(hw_start_column).cast(pl.Utf8).str.extract(r"(\d+)").cast(pl.Int64, strict=False)
    )
    hw_end_page_expr = (
        pl.col(hw_end_column).cast(pl.Utf8).str.extract(r"(\d+)").cast(pl.Int64, strict=False)
    )

    hw_page_count_expr = hw_end_page_expr - hw_start_page_expr + 1

    capped_hw_page_count_expr = (
        pl.when(hw_page_count_expr.is_null())
        .then(0)
        .when(hw_page_count_expr < 0)
        .then(0)
        .when(hw_page_count_expr > page_limit)
        .then(page_limit)
        .otherwise(hw_page_count_expr)
    )

    return (
        pl.when(hw_name_expr != "")
        .then(pl.when(~is_excluded_hw_expr).then(capped_hw_page_count_expr).otherwise(0))
        .otherwise(0)
    )


def build_test_presence_exprs(
    lap_column_names: list[str], kaku_column_names: list[str]
) -> tuple[pl.Expr, pl.Expr, pl.Expr]:
    def _presence_expr(column_names: list[str]) -> pl.Expr:
        presence_checks = [
            pl.col(column_name).is_not_null()
            & (pl.col(column_name).cast(pl.Utf8).str.strip_chars() != "")
            for column_name in column_names
        ]
        return (
            pl.any_horizontal(presence_checks).fill_null(False)
            if presence_checks
            else pl.lit(False)
        )

    lap_expr = _presence_expr(lap_column_names)
    kaku_expr = _presence_expr(kaku_column_names)
    has_lap_expr = lap_expr.alias("has_lap")
    has_kaku_expr = kaku_expr.alias("has_kaku")
    has_test_expr = (lap_expr | kaku_expr).fill_null(False).alias("has_test")
    return has_lap_expr, has_kaku_expr, has_test_expr


def configure_tk_scaling(root: tk.Tk) -> None:
    try:
        scale = float(root.tk.call("tk", "scaling"))
        root.tk.call("tk", "scaling", scale)
    except Exception:
        pass

def load_sheet_preserve_extra_columns(path: str, sheet_index: int = 1) -> pl.DataFrame:
    try:
        import fastexcel
        excel_reader = fastexcel.read_excel(path)
        sheet = excel_reader.load_sheet(sheet_index - 1, header_row=None)
        df_raw = sheet.to_polars()

        if df_raw.height == 0:
            return pl.DataFrame()

        progress("fastexcel読み込みモード")
        header = [str(x) if x is not None else "" for x in df_raw.row(0)]

        hw_base_idx = -1
        for i, h in enumerate(header):
            if "宿題名" in h:
                hw_base_idx = i
                break

        if hw_base_idx != -1:
            pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
            k = 2
            pattern_idx = 0
            for i in range(hw_base_idx + 1, len(header)):
                if header[i].strip() == "":
                    header[i] = f"{pattern[pattern_idx]}_{k}"
                    pattern_idx += 1
                    if pattern_idx >= 3:
                        pattern_idx = 0
                        k += 1
        else:
            idx = 1
            for i in range(len(header)):
                if header[i].strip() == "":
                    header[i] = f"追加列_{idx}"
                    idx += 1

        unique_headers = []
        seen = set()
        for h in header:
            h = h.strip()
            if h == "":
                h = "無名列"
            original_h = h
            counter = 1
            while h in seen:
                h = f"{original_h}_{counter}"
                counter += 1
            seen.add(h)
            unique_headers.append(h)

        df_data = df_raw.slice(1)
        df_data = df_data.rename(dict(zip(df_data.columns, unique_headers)))
        return df_data

    except Exception as e:
        progress(f"本番Excelフォールバック読み込み: {e}")
        # openpyxlを読み込みフェーズの最小限のフォールバック手段としてのみ残す
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[sheet_index - 1]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not rows:
            return pl.DataFrame()

        header = [str(c) if c is not None else "" for c in rows[0]]
        max_cols = max(len(r) for r in rows)

        if max_cols > len(header):
            try:
                base = header.index("宿題名")
                pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
                k = 2
                while len(header) < max_cols:
                    for p in pattern:
                        header.append(f"{p}_{k}")
                        if len(header) >= max_cols:
                            break
                    k += 1
            except ValueError:
                idx = 1
                while len(header) < max_cols:
                    header.append(f"追加列_{idx}")
                    idx += 1

        header_len = len(header)
        padded_rows = [
            (tuple(row) + (None,) * (header_len - len(row)) if len(row) < header_len else tuple(row))
            for row in rows[1:]
        ]
        return pl.DataFrame(padded_rows, schema=header, orient="row")


def get_jp_weekday(dt_obj):
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    return weekdays[dt_obj.weekday()]


def process_attendance_data(
    input_file: str,
) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame, pl.DataFrame, str, str, str, list[str], list[str]]:
    progress("出欠データの解析開始")
    df = load_sheet_preserve_extra_columns(input_file, sheet_index=1)
    cols = df.columns

    if "教室" in cols:
        progress("教室名から'FS'の除去")
        df = df.with_columns(
            pl.col("教室").cast(pl.Utf8).str.replace_all(r"(ＦＳ|FS|ｆｓ|fs)", "").str.strip_chars().alias("教室")
        )

    hw_name_cols = [c for c in cols if c and "宿題名" in c]
    if hw_name_cols:
        progress(f"宿題名列(計{len(hw_name_cols)}列)の正規化を実行")
        norm_series_list = []
        for c in hw_name_cols:
            try:
                orig_vals = df.get_column(c).to_list()
            except Exception:
                orig_vals = []

            norm_vals = [normalize_text_for_matching(v) for v in orig_vals]
            norm_col_name = f"{c}__norm"
            norm_series_list.append(pl.Series(norm_col_name, norm_vals))
        if norm_series_list:
            df = df.with_columns(norm_series_list)
            cols = df.columns

    possible_date_cols = ["日付", "授業日"]
    date_col: str | None = None
    for dc in possible_date_cols:
        if dc in cols:
            date_col = dc
            break

    if date_col is not None:
        valid_dates = df.select(pl.col(date_col).cast(pl.Date, strict=False)).drop_nulls()
    else:
        valid_dates = pl.DataFrame()

    if valid_dates.height > 0:
        min_dt = valid_dates.min().item()
        max_dt = valid_dates.max().item()
        wd_min = get_jp_weekday(min_dt)
        wd_max = get_jp_weekday(max_dt)
        min_date_file = min_dt.strftime("%Y%m%d")
        max_date_file = max_dt.strftime("%Y%m%d")
        date_range_str = f"集計期間: {min_dt.strftime('%Y/%m/%d')}({wd_min}) ～ {max_dt.strftime('%Y/%m/%d')}({wd_max})"
    else:
        min_date_file, max_date_file = "不明", "不明"
        date_range_str = "集計期間: 不明"

    sheet_name = "集計結果"
    now_str = datetime.now().strftime("%y%m%d-%H%M%S")
    input_dir = os.path.dirname(input_file)
    output_filename = f"集計結果_[{min_date_file}-{max_date_file}]_{now_str}.xlsx"
    output_file = os.path.join(input_dir, output_filename)

    dt_iso = pl.col(date_col).cast(pl.Date, strict=False).dt.week() if date_col is not None else pl.lit(None)

    lap_section_start_index = 0
    for i, c in enumerate(cols):
        if c and ("ラップテスト" in c or "ラップ" in c):
            lap_section_start_index = i
            break

    hw_section_start_index = len(cols)
    for i, c in enumerate(cols):
        if c and "宿題名" in c:
            hw_section_start_index = i
            break

    test_section_column_names = cols[lap_section_start_index:hw_section_start_index]
    lap_count_column_names = [c for c in test_section_column_names if "ラップ" in c and "（分子）" in c]
    confirmation_count_column_names = [c for c in test_section_column_names if "確認" in c and "（分子）" in c]

    hw_page_exprs = []
    for i in range(hw_section_start_index, len(cols), 3):
        if i + 2 < len(cols):
            hw_name_col = cols[i]
            hw_start_col = cols[i + 1]
            hw_end_col = cols[i + 2]
            norm_col = f"{hw_name_col}__norm"
            hw_name_for_expr = norm_col if norm_col in cols else hw_name_col
            hw_page_exprs.append(build_hw_pages_expr(hw_name_for_expr, hw_start_col, hw_end_col, EXCLUDED_HW_TERMS))

    missing_required_columns = [req_col for req_col in ["出欠", "担当講師名", "担当講師N0", "教室"] if req_col not in cols]
    if missing_required_columns:
        raise ValueError(f"必須の列が見つかりません: {', '.join(missing_required_columns)}")

    progress("出席データの抽出および宿題・テスト集計ロジックを構築中")
    attendance_metrics_df = (
        df.filter(pl.col("出欠") == "出席").with_columns(
            pl.col("担当講師名").cast(pl.Utf8).str.replace_all(r"[Ａ-Ｚ_]", "").alias("担当講師名"),
            dt_iso.cast(pl.Utf8).alias("week_num"),
            (pl.col(date_col).cast(pl.Date, strict=False) if date_col else pl.lit(None)).alias("date_val"),
            (pl.sum_horizontal(hw_page_exprs) if hw_page_exprs else pl.lit(0)).alias("hw_pages"),
        )
        .with_columns(*build_test_presence_exprs(lap_count_column_names, confirmation_count_column_names))
    )

    week_period_by_week_number = {}
    if "date_val" in attendance_metrics_df.columns:
        week_ranges_df = (
            attendance_metrics_df.filter(pl.col("date_val").is_not_null())
            .group_by("week_num")
            .agg(pl.col("date_val").min().alias("min_dt"), pl.col("date_val").max().alias("max_dt"))
        )
        for row in week_ranges_df.iter_rows(named=True):
            w_num = row["week_num"]
            md = row["min_dt"]
            xd = row["max_dt"]
            if md and xd:
                week_period_by_week_number[str(w_num)] = f"{md.strftime('%Y/%m/%d')}({get_jp_weekday(md)}) ～ {xd.strftime('%Y/%m/%d')}({get_jp_weekday(xd)})"

    def aggregate_and_pivot(attendance_metrics_local_df: pl.DataFrame):
        base_aggregate_df_local = attendance_metrics_local_df.group_by(
            ["担当講師N0", "担当講師名", "教室", "week_num"]
        ).agg(
            pl.len().alias("授業数"),
            pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"),
            pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数"),
        )

        aggregated_df_local = base_aggregate_df_local.with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率"),
        ).select(["担当講師N0", "担当講師名", "教室", "week_num", "授業数", "宿題平均", "ラップ回数", "ラップ率", "テスト回数", "テスト率"])

        pivot_df_local = aggregated_df_local.pivot(
            on="week_num",
            index=["担当講師N0", "担当講師名", "教室"],
            values=["授業数", "宿題平均", "ラップ回数", "ラップ率", "テスト回数", "テスト率"],
        )

        week_numbers_local = sorted(list(set([c.split("_")[-1] for c in pivot_df_local.columns if "_" in c])))
        class_count_columns_local = [f"授業数_{w}" for w in week_numbers_local]
        total_class_expr_local = pl.sum_horizontal([pl.col(c) for c in class_count_columns_local if c in pivot_df_local.columns]).fill_null(0)
        pivot_df_local = pivot_df_local.with_columns(total_class_expr_local.alias("総授業数"))

        for w in week_numbers_local:
            c_class, c_hw, c_lap_cnt, c_lap_rate, c_test_cnt, c_test_rate = f"授業数_{w}", f"宿題平均_{w}", f"ラップ回数_{w}", f"ラップ率_{w}", f"テスト回数_{w}", f"テスト率_{w}"
            if c_class in pivot_df_local.columns:     pivot_df_local = pivot_df_local.with_columns(pl.col(c_class).fill_null(0).alias(c_class))
            if c_lap_cnt in pivot_df_local.columns:   pivot_df_local = pivot_df_local.with_columns(pl.col(c_lap_cnt).fill_null(0).alias(c_lap_cnt))
            if c_test_cnt in pivot_df_local.columns:  pivot_df_local = pivot_df_local.with_columns(pl.col(c_test_cnt).fill_null(0).alias(c_test_cnt))
            if c_hw in pivot_df_local.columns and c_class in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_hw)).alias(c_hw))
            if c_lap_rate in pivot_df_local.columns and c_class in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_lap_rate)).alias(c_lap_rate))
            if c_test_rate in pivot_df_local.columns and c_class in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_test_rate)).alias(c_test_rate))

        desired_column_order_local = ["担当講師N0", "担当講師名", "教室", "総授業数"]
        rename_map_local = {"担当講師N0": "No", "担当講師名": "氏名", "教室": "教室", "総授業数": "総授業数"}

        week_period_labels_local = []
        for i, w in enumerate(week_numbers_local, start=1):
            c_class, c_hw, c_lap_cnt, c_lap_rate, c_test_cnt, c_test_rate = f"授業数_{w}", f"宿題平均_{w}", f"ラップ回数_{w}", f"ラップ率_{w}", f"テスト回数_{w}", f"テスト率_{w}"
            desired_column_order_local.extend([c_class, c_hw, c_lap_cnt, c_lap_rate, c_test_cnt, c_test_rate])
            rename_map_local[c_class], rename_map_local[c_hw], rename_map_local[c_lap_cnt] = f"授業{i}", f"平均HW{i}", f"Lap数{i}"
            rename_map_local[c_lap_rate], rename_map_local[c_test_cnt], rename_map_local[c_test_rate] = f"Lap％{i}", f"テスト数{i}", f"テスト％{i}"
            week_period_labels_local.append(week_period_by_week_number.get(str(w), f"Week {i}"))

        existing_desired_columns_local = [c for c in desired_column_order_local if c in pivot_df_local.columns]
        return pivot_df_local.select(existing_desired_columns_local).rename({k: v for k, v in rename_map_local.items() if k in existing_desired_columns_local}).sort(["教室", "No"]), week_period_labels_local

    progress("「全集計」用データのマトリクス変換を実行中")
    pivot_df_all, week_period_labels_all = aggregate_and_pivot(attendance_metrics_df)

    grade_exclusion_condition = pl.col("学年").cast(pl.Utf8).str.contains(r"(高3|高３|高卒)").fill_null(False) if "学年" in cols else pl.lit(False)
    subject_exclusion_condition = pl.col("科目").cast(pl.Utf8).str.contains(r"(国語|小論文)").fill_null(False) if "科目" in cols else pl.lit(False)

    progress("「除外集計」用データの抽出と集計を実行中")
    excluded_metrics_df = attendance_metrics_df.filter(~grade_exclusion_condition & ~subject_exclusion_condition)
    excluded_pivot_df, week_period_labels_excl = aggregate_and_pivot(excluded_metrics_df)

    classroom_summary_df = (
        attendance_metrics_df.group_by("教室")
        .agg(pl.len().alias("授業数"), pl.col("担当講師N0").n_unique().alias("講師数"), pl.col("hw_pages").sum().alias("宿計"), pl.col("has_lap").cast(pl.Int32).sum().alias("ラ回"), pl.col("has_test").cast(pl.Int32).sum().alias("テ回"))
        .with_columns((pl.col("宿計") / pl.col("授業数")).round(1).alias("宿題平均"), (pl.col("ラ回") / pl.col("授業数")).alias("ラップ率"), (pl.col("テ回") / pl.col("授業数")).alias("テスト率")).sort("教室")
    )

    excluded_classroom_summary_df = (
        excluded_metrics_df.group_by("教室")
        .agg(pl.len().alias("授業数"), pl.col("担当講師N0").n_unique().alias("講師数"), pl.col("hw_pages").sum().alias("宿計"), pl.col("has_lap").cast(pl.Int32).sum().alias("ラ回"), pl.col("has_test").cast(pl.Int32).sum().alias("テ回"))
        .with_columns((pl.col("宿計") / pl.col("授業数")).round(1).alias("宿題平均"), (pl.col("ラ回") / pl.col("授業数")).alias("ラップ率"), (pl.col("テ回") / pl.col("授業数")).alias("テスト率")).sort("教室")
    )

    del df, attendance_metrics_df, excluded_metrics_df
    gc.collect()

    return pivot_df_all, excluded_pivot_df, classroom_summary_df, excluded_classroom_summary_df, output_file, date_range_str, sheet_name, week_period_labels_all, week_period_labels_excl

def format_excel_fast(
    all_pivot_df: pl.DataFrame,
    excluded_pivot_df: pl.DataFrame,
    classroom_summary_df: pl.DataFrame,
    excluded_classroom_summary_df: pl.DataFrame,
    output_file: str,
    date_range_str: str,
    sheet_name: str,
    week_period_labels_all: list[str],
    week_period_labels_excl: list[str],
):
    progress("xlsxwriterによるデータ書き出し・ベース書式設定中...")

    with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
        workbook = writer.book
        ud_font = select_font()

        fmt_title = workbook.add_format({"font_name": ud_font, "bold": True, "font_size": 11, "align": "left", "valign": "vcenter"})
        fmt_title_center = workbook.add_format({"font_name": ud_font, "bold": True, "font_size": 11, "align": "center", "valign": "vcenter", "bg_color": "#E8F6F3", "border": 1})
        fmt_header = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#D3D3D3", "bold": True})
        fmt_white = workbook.add_format({"font_name": ud_font, "border": 1})
        fmt_khaki = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#F0E68C"})
        fmt_lavender = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#E6E6FA"})
        fmt_white_pct = workbook.add_format({"font_name": ud_font, "border": 1, "num_format": "0.0%"})
        fmt_cond_orange = workbook.add_format({"font_name": ud_font, "bg_color": "#FFA500", "font_color": "#000000"})
        fmt_cond_firebrick = workbook.add_format({"font_name": ud_font, "bg_color": "#B22222", "font_color": "#FFFFFF"})

        def write_sheet_v7(writer, sheet_df: pl.DataFrame, summary_df: pl.DataFrame, week_period_labels: list[str], sheet_title: str):
            progress(f"シート [{sheet_title}] へデータをロード中...")
            headers = list(sheet_df.columns)
            max_col = len(headers) - 1
            header_row, data_start, nrows = 1, 2, sheet_df.height

            col_values: dict[str, list] = {}
            for col in headers:
                try:
                    s = sheet_df.get_column(col)
                except Exception:
                    s = pl.Series(col, [None] * nrows)

                if col == "No":
                    vals = [None if v is None else str(v).strip() for v in s.to_list()]
                    col_values[col] = [None if v is None else (v.lstrip("0") if v.lstrip("0") != "" else v) for v in vals]
                    continue
                if any(k in col for k in ("授業", "Lap数", "テスト数", "総授業数")):
                    try:
                        col_values[col] = [0.0 if v is None else float(v) for v in s.cast(pl.Float64, strict=False).fill_null(0).to_list()]
                    except Exception:
                        col_values[col] = [0.0] * nrows
                    continue
                if ("平均" in col) or ("％" in col) or ("率" in col):
                    try:
                        col_values[col] = [None if v is None else float(v) for v in s.cast(pl.Float64, strict=False).to_list()]
                    except Exception:
                        col_values[col] = [None] * nrows
                    continue
                col_values[col] = [None if v is None else str(v) for v in s.to_list()]

            lesson_nums = []
            for h in headers:
                m = re.search(r"授業(\d+)", str(h))
                if m: lesson_nums.append(int(m.group(1)))
            lesson_nums = sorted(list(set(lesson_nums)))

            for i in lesson_nums:
                lesson_col = f"授業{i}"
                if lesson_col not in col_values: continue
                lesson_vals = col_values[lesson_col]
                related = [f"平均HW{i}", f"Lap数{i}", f"Lap％{i}", f"テスト数{i}", f"テスト％{i}"]
                for idx, lv in enumerate(lesson_vals):
                    if lv == 0 or lv == 0.0:
                        for c in related:
                            if c in col_values: col_values[c][idx] = None

            if nrows > 0:
                progress(f"シート [{sheet_title}] の『全体AVG行』および『各週ごとの教室平均ブロック』の集計・合成処理を実行中...")
                row_classrooms = col_values.get("教室", [""] * nrows)
                unique_classrooms = sorted(list(set([v for v in row_classrooms if v])))
                final_col_values: dict[str, list] = {c: [] for c in headers}

                # --- 1. 全体平均行（No: 0, 氏名: AVG, 教室: FS）の作成 ---
                total_lessons_all = sum(v for v in col_values["総授業数"] if isinstance(v, (int, float)))
                final_col_values["No"].append("0")
                final_col_values["氏名"].append("AVG")
                final_col_values["教室"].append("FS")
                final_col_values["総授業数"].append(total_lessons_all)

                for i in lesson_nums:
                    w_lessons = sum(v for v in col_values[f"授業{i}"] if isinstance(v, (int, float)))
                    final_col_values[f"授業{i}"].append(w_lessons)
                    if w_lessons == 0:
                        final_col_values[f"平均HW{i}"].append(None)
                        final_col_values[f"Lap数{i}"].append(0.0)
                        final_col_values[f"Lap％{i}"].append(None)
                        final_col_values[f"テスト数{i}"].append(0.0)
                        final_col_values[f"テスト％{i}"].append(None)
                    else:
                        w_laps = sum(v for v in col_values[f"Lap数{i}"] if isinstance(v, (int, float)))
                        w_tests = sum(v for v in col_values[f"テスト数{i}"] if isinstance(v, (int, float)))
                        final_col_values[f"Lap数{i}"].append(w_laps)
                        final_col_values[f"テスト数{i}"].append(w_tests)
                        final_col_values[f"Lap％{i}"].append(w_laps / w_lessons)
                        final_col_values[f"テスト％{i}"].append(w_tests / w_lessons)

                        total_hw_pages = 0.0
                        for idx in range(nrows):
                            l_val, h_val = col_values[f"授業{i}"][idx], col_values[f"平均HW{i}"][idx]
                            if isinstance(l_val, (int, float)) and isinstance(h_val, (int, float)):
                                total_hw_pages += h_val * l_val
                        final_col_values[f"平均HW{i}"].append(round(total_hw_pages / w_lessons, 1))

                # --- 2. 各教室ごとに「教室平均行」＋「講師データ」をブロック化 ---
                for target_cls in unique_classrooms:
                    cls_indices = [idx for idx, cls in enumerate(row_classrooms) if cls == target_cls]
                    if not cls_indices: continue

                    try:
                        cls_sum_row = summary_df.filter(pl.col("教室").cast(pl.Utf8).str.strip_chars() == target_cls)
                        cls_total_lessons = float(cls_sum_row.row(0)[1]) if cls_sum_row.height == 1 else 0.0
                    except Exception:
                        cls_total_lessons = sum(col_values["総授業数"][idx] for idx in cls_indices)

                    final_col_values["No"].append("0")
                    final_col_values["氏名"].append("AVG")
                    final_col_values["教室"].append(target_cls)
                    final_col_values["総授業数"].append(cls_total_lessons)

                    for i in lesson_nums:
                        w_lessons_cls = sum(col_values[f"授業{i}"][idx] for idx in cls_indices if isinstance(col_values[f"授業{i}"][idx], (int, float)))
                        final_col_values[f"授業{i}"].append(w_lessons_cls)
                        if w_lessons_cls == 0:
                            final_col_values[f"平均HW{i}"].append(None)
                            final_col_values[f"Lap数{i}"].append(0.0)
                            final_col_values[f"Lap％{i}"].append(None)
                            final_col_values[f"テスト数{i}"].append(0.0)
                            final_col_values[f"テスト％{i}"].append(None)
                        else:
                            w_laps_cls = sum(col_values[f"Lap数{i}"][idx] for idx in cls_indices if isinstance(col_values[f"Lap数{i}"][idx], (int, float)))
                            w_tests_cls = sum(col_values[f"テスト数{i}"][idx] for idx in cls_indices if isinstance(col_values[f"テスト数{i}"][idx], (int, float)))
                            final_col_values[f"Lap数{i}"].append(w_laps_cls)
                            final_col_values[f"テスト数{i}"].append(w_tests_cls)
                            final_col_values[f"Lap％{i}"].append(w_laps_cls / w_lessons_cls)
                            final_col_values[f"テスト％{i}"].append(w_tests_cls / w_lessons_cls)
                            total_hw_pages_cls = 0.0
                            for idx in cls_indices:
                                l_val, h_val = col_values[f"授業{i}"][idx], col_values[f"平均HW{i}"][idx]
                                if isinstance(l_val, (int, float)) and isinstance(h_val, (int, float)):
                                    total_hw_pages_cls += h_val * l_val
                            final_col_values[f"平均HW{i}"].append(round(total_hw_pages_cls / w_lessons_cls, 1))

                    for idx in cls_indices:
                        for col in headers:
                            final_col_values[col].append(col_values[col][idx])

                col_values, nrows = final_col_values, len(final_col_values["No"])

            for col, vals in col_values.items():
                for j, v in enumerate(vals):
                    if v is None: vals[j] = "-"

            if sheet_title in writer.sheets:
                worksheet = writer.sheets[sheet_title]
            else:
                worksheet = workbook.add_worksheet(sheet_title)
                writer.sheets[sheet_title] = worksheet

            for col_idx, col_name in enumerate(headers):
                worksheet.write_column(data_start, col_idx, col_values.get(col_name, ["-"] * nrows))

            # xlsxwriter側でのベースカラーパレット設定
            if nrows > 0:
                for col_idx, col_name in enumerate(headers):
                    if col_idx in (0, 1, 2):    worksheet.set_column(col_idx, col_idx, 10, fmt_white)
                    elif col_idx == 3:          worksheet.set_column(col_idx, col_idx, 10, fmt_lavender)
                    else:
                        offset = col_idx - (4 + ((col_idx - 4) // 6) * 6)
                        if offset == 0:         worksheet.set_column(col_idx, col_idx, 9, fmt_khaki)
                        elif offset in (3, 5):  worksheet.set_column(col_idx, col_idx, 9, fmt_white_pct)
                        else:                   worksheet.set_column(col_idx, col_idx, 9, fmt_white)

            try:
                worksheet.merge_range(0, 0, 0, min(3, max_col), date_range_str, fmt_title)
            except Exception:
                worksheet.write(0, 0, date_range_str, fmt_title)
            worksheet.set_row(0, 20)
            worksheet.write_row(header_row, 0, headers, fmt_header)

            for i, period_str in enumerate(week_period_labels):
                start_col = 4 + i * 6
                if start_col + 5 <= max_col:
                    worksheet.merge_range(0, start_col, 0, start_col + 5, period_str, fmt_title_center)

            last_row = data_start + nrows - 1
            worksheet.autofilter(header_row, 0, last_row, max_col)

            for c_base in range(4, len(headers), 6):
                if c_base + 5 > max_col: break
                worksheet.conditional_format(data_start, c_base + 1, last_row, c_base + 1, {"type": "cell", "criteria": "<", "value": 6, "format": fmt_cond_orange})
                worksheet.conditional_format(data_start, c_base + 3, last_row, c_base + 3, {"type": "cell", "criteria": "<", "value": 0.7, "format": fmt_cond_firebrick})
                worksheet.conditional_format(data_start, c_base + 5, last_row, c_base + 5, {"type": "cell", "criteria": "<", "value": 0.7, "format": fmt_cond_firebrick})

            worksheet.freeze_panes(data_start, 4)
            progress(f"シート [{sheet_title}] のデータ一括配置および条件付き書式の設定が完了しました。")

        # 2シート体制での出力
        write_sheet_v7(writer, all_pivot_df, classroom_summary_df, week_period_labels_all, "全集計")
        write_sheet_v7(writer, excluded_pivot_df, excluded_classroom_summary_df, week_period_labels_excl, "除外集計")

    # =========================================================================
    # 【openpyxl】による後処理（Autofit、改ページ・印刷設定の強制適用）
    # =========================================================================
    progress("openpyxlエンジンを起動。余白の調整、A:D列のAutofit、垂直改ページの強制固定処理を開始します...")
    wb = openpyxl.load_workbook(output_file)

    for sheet in wb.worksheets:
        progress(f"シート [{sheet.title}] の最終ページレイアウト調整中...")
        # A:D列に対して明示的に文字数測定のAutofitを完全適用（上書きバグ防止）
        for col_letter in ['A', 'B', 'C', 'D']:
            max_len = 0
            for row_idx in range(3, sheet.max_row + 1):
                val = sheet[f"{col_letter}{row_idx}"].value
                if val is not None: max_len = max(max_len, len(str(val)))

            if col_letter == 'A':    sheet.column_dimensions[col_letter].width = max(max_len * 2.0, 6)
            elif col_letter == 'B':  sheet.column_dimensions[col_letter].width = max(max_len * 1.8, 14)
            elif col_letter == 'C':  sheet.column_dimensions[col_letter].width = max(max_len * 1.8, 12)
            elif col_letter == 'D':  sheet.column_dimensions[col_letter].width = max(max_len * 1.5, 11)

        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

        # 【修正箇所】Pylanceの型エラーの原因だった古い引数呼び出し（margin_to_indexなど）を完全に排除
        # 直接インチ数値を指定する高精度かつ安全な記述のみに一元化しました。
        margin_inch = 1.0 / 2.54
        sheet.page_margins.left = margin_inch
        sheet.page_margins.right = margin_inch
        sheet.page_margins.top = margin_inch
        sheet.page_margins.bottom = margin_inch
        sheet.page_margins.header = 0.5 / 2.54
        sheet.page_margins.footer = 0.5 / 2.54

        if sheet.sheet_properties.pageSetUpPr is None:
            from openpyxl.worksheet.properties import PageSetupProperties
            sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        else:
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth, sheet.page_setup.fitToHeight = 1, 0

        # タイトル行・列の設定
        sheet.print_title_cols = "A:D"
        sheet.print_title_rows = "1:3"

        # 垂直方向の『1期間（6列）ごと』の強制改ページ制御
        sheet.col_breaks = ColBreak()
        col_count = sheet.max_column
        current_break_col = 4 + 6  # 10列目 (第1期間の右端)
        while current_break_col < col_count:
            from openpyxl.worksheet.pagebreak import Break
            sheet.col_breaks.append(Break(id=current_break_col))
            current_break_col += 6

        from openpyxl.utils import get_column_letter
        sheet.print_area = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    progress("調整されたページ設定情報をブックへ書き込み、ファイルを確定保存しています...")
    wb.save(output_file)
    wb.close()

    # 【メモリ即時破棄】データフレームを完全消去してメモリ返却
    progress("処理完了に伴うメモリの即時解放（ガベージコレクション）を実行中...")
    del all_pivot_df, excluded_pivot_df, classroom_summary_df, excluded_classroom_summary_df
    gc.collect()

def show_error_dialog(title: str, message: str, parent: tk.Tk | None = None) -> None:
    dlg = tk.Toplevel(parent) if parent else tk.Toplevel()
    dlg.title(title)
    try:
        scale = float(parent.tk.call("tk", "scaling")) if parent else float(dlg.tk.call("tk", "scaling"))
    except Exception:
        scale = 1.0

    base_font = tkfont.nametofont("TkDefaultFont")
    fsize = max(int(base_font.cget("size") * scale), 10) if Exception else 11
    dlg_font = tkfont.Font(family=base_font.cget("family"), size=fsize)

    dlg.geometry(f"{int(700*scale)}x{int(360*scale)}")
    txt = tk.Text(dlg, wrap="word", font=dlg_font)
    txt.insert("1.0", message)
    txt.configure(state="disabled")
    txt.pack(expand=True, fill="both", padx=int(6 * scale), pady=int(6 * scale))

    frm = tk.Frame(dlg)
    frm.pack(fill="x", padx=int(6 * scale), pady=int(6 * scale))

    def _copy():
        try:
            dlg.clipboard_clear()
            dlg.clipboard_append(message)
        except Exception:
            pass

    tk.Button(frm, text="エラーメッセージをコピー", command=_copy, font=dlg_font).pack(side="left")
    tk.Button(frm, text="閉じる", command=dlg.destroy, font=dlg_font).pack(side="right")
    try:
        dlg.transient(parent)
        dlg.grab_set()
        dlg.wait_window()
    except Exception:
        pass


def main():
    enable_windows_dpi_awareness()
    root = tk.Tk()
    configure_tk_scaling(root)
    root.withdraw()
    try:
        root.attributes("-topmost", True)
        root.lift()
        root.focus_force()

        print("集計元のExcelファイルを選択してください...")
        input_file = filedialog.askopenfilename(
            title="集計元のExcelファイルを選択してください",
            filetypes=[("MSL出力ファイル", "scube2*.xlsx"), ("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")],
            parent=root,
        )
        root.attributes("-topmost", False)

        if not input_file:
            print("ファイルの選択がキャンセルされました。")
            return

        total_start_time = time.perf_counter()

        (
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
        ) = process_attendance_data(input_file)

        format_excel_fast(
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
        )

        total_elapsed = time.perf_counter() - total_start_time
        processed_classroom_count = classroom_summary_df.height
        try:
            processed_lesson_count = int(classroom_summary_df.get_column("授業数").sum())
        except Exception:
            processed_lesson_count = 0

        result_msg = (
            f"処理が完了しました！\n"
            f"実行時間: {total_elapsed:.2f}秒\n\n"
            f"処理教室数: {processed_classroom_count}教室\n"
            f"授業数: {processed_lesson_count}件\n\n"
            f"出力先:\n{output_file}"
        )

        root.attributes("-topmost", True)
        messagebox.showinfo("集計完了", result_msg, parent=root)
        root.attributes("-topmost", False)
        progress(f"処理完了(実行時間: {total_elapsed:.2f}秒, 処理教室数: {processed_classroom_count}教室, 授業数: {processed_lesson_count}件)")

    except KeyboardInterrupt:
        print("処理を中断しました。")
    except Exception as e:
        tb = traceback.format_exc()
        print("\n=== エラー詳細 ===")
        print(tb)
        print("==================\n")
        try:
            root.attributes("-topmost", True)
            show_error_dialog("処理中にエラーが発生しました", tb, parent=root)
        except Exception:
            messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}", parent=root)
    finally:
        root.destroy()


if __name__ == "__main__":
    main()