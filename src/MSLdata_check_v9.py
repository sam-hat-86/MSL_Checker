import polars as pl
import tkinter as tk
from tkinter import filedialog, messagebox
import xlsxwriter
import os
from datetime import datetime
from typing import Any, cast
import re
import unicodedata
import pandas as pd
import sys
import ctypes
import tkinter.font as tkfont
import traceback
import time
import gc

def progress(msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    print(f"[{ts}] {msg}")

EXCLUDED_HW_TERMS = [
    "タンゴ", "単語", "シスタン", "シス単", "ターゲット", "リープ",
    "leap", "ゲットスルー", "パスタン", "パス単", "マドンナ"
]

FONT_PREFERRED = ["Noto Sans JP", "Arial Unicode MS", "Arial"]

def select_font(preferred: list[str] | None = None) -> str:
    prefs = preferred or FONT_PREFERRED
    try:
        root_tmp = tk.Tk()
        root_tmp.withdraw()
        fams = set(tkfont.families())
        root_tmp.destroy()
        for f in prefs:
            if f in fams: return f
    except Exception:
        pass
    for f in prefs: return f
    return "Calibri"

def normalize_text_for_matching(s: object) -> str:
    if s is None: return ""
    t = str(s)
    t = unicodedata.normalize("NFKC", t).lower()
    t = re.sub(r"\s+", "", t)
    t = "".join(chr(ord(c) + 96) if "ぁ" <= c <= "ん" else c for c in t)
    return t

def enable_windows_dpi_awareness() -> None:
    if sys.platform != "win32": return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

def build_hw_pages_expr(hw_name_column: str, hw_start_column: str, hw_end_column: str, excluded_terms: list[str], page_limit: int = 30) -> pl.Expr:
    hw_name_expr = pl.col(hw_name_column).cast(pl.Utf8).fill_null("").str.strip_chars()
    active_excluded_hw_terms = [term for term in excluded_terms if str(term).strip()]
    is_excluded_hw_expr = pl.any_horizontal([hw_name_expr.str.contains(term, literal=True) for term in active_excluded_hw_terms]) if active_excluded_hw_terms else pl.lit(False)
    hw_start_page_expr = pl.col(hw_start_column).cast(pl.Utf8).str.extract(r"(\d+)").cast(pl.Int64, strict=False)
    hw_end_page_expr = pl.col(hw_end_column).cast(pl.Utf8).str.extract(r"(\d+)").cast(pl.Int64, strict=False)
    hw_page_count_expr = hw_end_page_expr - hw_start_page_expr + 1
    capped_hw_page_count_expr = pl.when(hw_page_count_expr.is_null()).then(0).when(hw_page_count_expr < 0).then(0).when(hw_page_count_expr > page_limit).then(page_limit).otherwise(hw_page_count_expr)
    return pl.when(hw_name_expr != "").then(pl.when(~is_excluded_hw_expr).then(capped_hw_page_count_expr).otherwise(0)).otherwise(0)

def build_test_presence_exprs(lap_column_names: list[str], kaku_column_names: list[str]) -> tuple[pl.Expr, pl.Expr, pl.Expr]:
    def _presence_expr(column_names: list[str]) -> pl.Expr:
        presence_checks = [pl.col(column_name).is_not_null() & (pl.col(column_name).cast(pl.Utf8).str.strip_chars() != "") for column_name in column_names]
        return pl.any_horizontal(presence_checks).fill_null(False) if presence_checks else pl.lit(False)
    lap_expr = _presence_expr(lap_column_names)
    kaku_expr = _presence_expr(kaku_column_names)
    return lap_expr.alias("has_lap"), kaku_expr.alias("has_kaku"), (lap_expr | kaku_expr).fill_null(False).alias("has_test")

def configure_tk_scaling(root: tk.Tk) -> None:
    try:
        scale = float(root.tk.call("tk", "scaling"))
        root.tk.call("tk", "scaling", scale)
    except Exception:
        pass

def load_sheet_preserve_extra_columns(path: str, sheet_index: int = 1) -> pl.DataFrame:
    try:
        import fastexcel
        excel_reader = fastexcel.read_excel(path)
        sheet = excel_reader.load_sheet(sheet_index - 1, header_row=None)
        df_raw = sheet.to_polars()
        if df_raw.height == 0: return pl.DataFrame()
        progress("fastexcel読み込みモード")
        header = [str(x) if x is not None else "" for x in df_raw.row(0)]
        hw_base_idx = -1
        for i, h in enumerate(header):
            if "宿題名" in h:
                hw_base_idx = i
                break
        if hw_base_idx != -1:
            pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
            k, pattern_idx = 2, 0
            for i in range(hw_base_idx + 1, len(header)):
                if header[i].strip() == "":
                    header[i] = f"{pattern[pattern_idx]}_{k}"
                    pattern_idx += 1
                    if pattern_idx >= 3:
                        pattern_idx, k = 0, k + 1
        else:
            idx = 1
            for i in range(len(header)):
                if header[i].strip() == "":
                    header[i] = f"追加列_{idx}"
                    idx += 1
        unique_headers = []
        seen = set()
        for h in header:
            h = h.strip()
            if h == "": h = "無名列"
            original_h, counter = h, 1
            while h in seen:
                h = f"{original_h}_{counter}"
                counter += 1
            seen.add(h)
            unique_headers.append(h)
        df_data = df_raw.slice(1)
        return df_data.rename(dict(zip(df_data.columns, unique_headers)))
    except Exception as e:
        progress(f"本番Excelフォールバック読み込み: {e}")
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[sheet_index - 1]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
        if not rows: return pl.DataFrame()
        header = [str(c) if c is not None else "" for c in rows[0]]
        max_cols = max(len(r) for r in rows)
        if max_cols > len(header):
            try:
                base = header.index("宿題名")
                pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
                k = 2
                while len(header) < max_cols:
                    for p in pattern:
                        header.append(f"{p}_{k}")
                        if len(header) >= max_cols: break
                    k += 1
            except ValueError:
                idx = 1
                while len(header) < max_cols:
                    header.append(f"追加列_{idx}")
                    idx += 1
        header_len = len(header)
        padded_rows = [(tuple(row) + (None,) * (header_len - len(row)) if len(row) < header_len else tuple(row)) for row in rows[1:]]
        return pl.DataFrame(padded_rows, schema=header, orient="row")

def get_jp_weekday(dt_obj):
    return ["月", "火", "水", "木", "金", "土", "日"][dt_obj.weekday()]

def process_attendance_data( input_file: str,) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame, pl.DataFrame, str, str, str, list[str], list[str], dict[str, int], dict[str, int]]:
    progress("出欠データの解析開始")
    df = load_sheet_preserve_extra_columns(input_file, sheet_index=1)
    cols = df.columns

    if "教室" in cols:
        progress("教室名から'FS'の除去")
        df = df.with_columns(pl.col("教室").cast(pl.Utf8).str.replace_all(r"(ＦＳ|FS|ｆｓ|fs)", "").str.strip_chars().alias("教室"))

    hw_name_cols = [c for c in cols if c and "宿題名" in c]
    if hw_name_cols:
        progress(f"宿題名列(計{len(hw_name_cols)}列)の正規化を実行")
        norm_series_list = []
        for c in hw_name_cols:
            try: orig_vals = df.get_column(c).to_list()
            except Exception: orig_vals = []
            norm_series_list.append(pl.Series(f"{c}__norm", [normalize_text_for_matching(v) for v in orig_vals]))
        if norm_series_list:
            df = df.with_columns(norm_series_list)
            cols = df.columns

    possible_date_cols = ["日付", "授業日"]
    date_col = next((dc for dc in possible_date_cols if dc in cols), None)
    valid_dates = df.select(pl.col(date_col).cast(pl.Date, strict=False)).drop_nulls() if date_col else pl.DataFrame()

    if valid_dates.height > 0:
        min_dt, max_dt = valid_dates.min().item(), valid_dates.max().item()
        min_date_file, max_date_file = min_dt.strftime("%Y%m%d"), max_dt.strftime("%Y%m%d")
        date_range_str = f"集計期間: {min_dt.strftime('%Y/%m/%d')}({get_jp_weekday(min_dt)}) ～ {max_dt.strftime('%Y/%m/%d')}({get_jp_weekday(max_dt)})"
    else:
        min_date_file, max_date_file = "不明", "不明"
        date_range_str = "集計期間: 不明"

    sheet_name = "集計結果"
    output_filename = f"集計結果_[{min_date_file}-{max_date_file}]_{datetime.now().strftime('%y%m%d-%H%M%S')}.xlsx"
    output_file = os.path.join(os.path.dirname(input_file), output_filename)

    dt_iso = pl.col(date_col).cast(pl.Date, strict=False).dt.week() if date_col else pl.lit(None)
    lap_section_start_index = next((i for i, c in enumerate(cols) if c and ("ラップテスト" in c or "ラップ" in c)), 0)
    hw_section_start_index = next((i for i, c in enumerate(cols) if c and "宿題名" in c), len(cols))

    test_section_column_names = cols[lap_section_start_index:hw_section_start_index]
    lap_count_column_names = [c for c in test_section_column_names if "ラップ" in c and "（分子）" in c]
    confirmation_count_column_names = [c for c in test_section_column_names if "確認" in c and "（分子）" in c]

    hw_page_exprs = []
    for i in range(hw_section_start_index, len(cols), 3):
        if i + 2 < len(cols):
            hw_name_for_expr = f"{cols[i]}__norm" if f"{cols[i]}__norm" in cols else cols[i]
            hw_page_exprs.append(build_hw_pages_expr(hw_name_for_expr, cols[i+1], cols[i+2], EXCLUDED_HW_TERMS))

    if any(req not in cols for req in ["出欠", "担当講師名", "担当講師N0", "教室"]):
        raise ValueError("必須の列が見つかりません")

    progress("出席データの抽出および宿題・テスト集計ロジックを構築中")
    attendance_metrics_df = df.filter(pl.col("出欠") == "出席").with_columns(
        pl.col("担当講師名").cast(pl.Utf8).str.replace_all(r"[Ａ-Ｚ_]", "").alias("担当講師名"),
        dt_iso.cast(pl.Utf8).alias("week_num"),
        (pl.col(date_col).cast(pl.Date, strict=False) if date_col else pl.lit(None)).alias("date_val"),
        (pl.sum_horizontal(hw_page_exprs) if hw_page_exprs else pl.lit(0)).alias("hw_pages"),
    ).with_columns(*build_test_presence_exprs(lap_count_column_names, confirmation_count_column_names))

    week_period_by_week_number = {}
    if "date_val" in attendance_metrics_df.columns:
        week_ranges_df = attendance_metrics_df.filter(pl.col("date_val").is_not_null()).group_by("week_num").agg(pl.col("date_val").min().alias("min_dt"), pl.col("date_val").max().alias("max_dt"))
        for row in week_ranges_df.iter_rows(named=True):
            md, xd = row["min_dt"], row["max_dt"]
            if md and xd: week_period_by_week_number[str(row["week_num"])] = f"{md.strftime('%Y/%m/%d')}({get_jp_weekday(md)}) ～ {xd.strftime('%Y/%m/%d')}({get_jp_weekday(xd)})"

    def aggregate_and_pivot(metrics_df: pl.DataFrame):
        aggregated = metrics_df.group_by(["担当講師N0", "担当講師名", "教室", "week_num"]).agg(
            pl.len().alias("授業数"), pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"), pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数")
        ).with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率")
        ).select(["担当講師N0", "担当講師名", "教室", "week_num", "授業数", "宿題平均", "ラップ回数", "ラップ率", "テスト回数", "テスト率"])

        pivot_df = aggregated.pivot(on="week_num", index=["担当講師N0", "担当講師名", "教室"], values=["授業数", "宿題平均", "ラップ回数", "ラップ率", "テスト回数", "テスト率"])
        week_nums = sorted(list(set([c.split("_")[-1] for c in pivot_df.columns if "_" in c])))
        pivot_df = pivot_df.with_columns(pl.sum_horizontal([pl.col(f"授業数_{w}") for w in week_nums if f"授業数_{w}" in pivot_df.columns]).fill_null(0).alias("総授業数"))

        for w in week_nums:
            c_class, c_hw, c_lap_cnt, c_lap_rate, c_test_cnt, c_test_rate = f"授業数_{w}", f"宿題平均_{w}", f"ラップ回数_{w}", f"ラップ率_{w}", f"テスト回数_{w}", f"テスト率_{w}"
            if c_class in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.col(c_class).fill_null(0).alias(c_class))
            if c_lap_cnt in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.col(c_lap_cnt).fill_null(0).alias(c_lap_cnt))
            if c_test_cnt in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.col(c_test_cnt).fill_null(0).alias(c_test_cnt))
            if c_hw in pivot_df.columns and c_class in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_hw)).alias(c_hw))
            if c_lap_rate in pivot_df.columns and c_class in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_lap_rate)).alias(c_lap_rate))
            if c_test_rate in pivot_df.columns and c_class in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_test_rate)).alias(c_test_rate))

        desired_order = ["担当講師N0", "担当講師名", "教室", "総授業数"]
        rename_map = {"担当講師N0": "No", "担当講師名": "氏名", "教室": "教室", "総授業数": "総授業数"}
        week_labels = []
        for i, w in enumerate(week_nums, start=1):
            desired_order.extend([f"授業数_{w}", f"宿題平均_{w}", f"ラップ回数_{w}", f"ラップ率_{w}", f"テスト回数_{w}", f"テスト率_{w}"])
            rename_map[f"授業数_{w}"], rename_map[f"宿題平均_{w}"], rename_map[f"ラップ回数_{w}"] = f"授業{i}", f"平均HW{i}", f"Lap数{i}"
            rename_map[f"ラップ率_{w}"], rename_map[f"テスト回数_{w}"], rename_map[f"テスト率_{w}"] = f"Lap％{i}", f"テスト数{i}", f"テスト％{i}"
            week_labels.append(week_period_by_week_number.get(str(w), f"Week {i}"))

        exist_cols = [c for c in desired_order if c in pivot_df.columns]
        return pivot_df.select(exist_cols).rename({k: v for k, v in rename_map.items() if k in exist_cols}).sort(["教室", "No"]), week_labels

    progress("「全集計」用データのマトリクス変換を実行中")
    pivot_df_all, week_period_labels_all = aggregate_and_pivot(attendance_metrics_df)

    grade_excl = pl.col("学年").cast(pl.Utf8).str.contains(r"(高3|高３|高卒)").fill_null(False) if "学年" in cols else pl.lit(False)
    subj_excl = pl.col("科目").cast(pl.Utf8).str.contains(r"(国語|小論文)").fill_null(False) if "科目" in cols else pl.lit(False)

    progress("「除外集計」用データの抽出と集計を実行中")
    excluded_df = attendance_metrics_df.filter(~grade_excl & ~subj_excl)
    excluded_pivot_df, week_period_labels_excl = aggregate_and_pivot(excluded_df)

    def get_summary(m_df):
        return m_df.group_by("教室").agg(
            pl.len().alias("授業数"), pl.col("担当講師N0").n_unique().alias("講師数"), pl.col("hw_pages").sum().alias("宿計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラ回"), pl.col("has_test").cast(pl.Int32).sum().alias("テ回")
        ).with_columns((pl.col("宿計")/pl.col("授業数")).round(1).alias("宿題平均"), (pl.col("ラ回")/pl.col("授業数")).alias("ラップ率"), (pl.col("テ回")/pl.col("授業数")).alias("テスト率")).sort("教室")

    classroom_summary_df = get_summary(attendance_metrics_df)
    excluded_classroom_summary_df = get_summary(excluded_df)

    # --- 高速Autofit用の文字数事前集計処理 (全集計と除外集計の両方を計算) ---
    progress("Polarsネイティブで全列の最大文字長(Autofit用)を一斉並列計算中...")

    # 全集計シート用の列幅マッピング
    max_lens_all = pivot_df_all.select([
        pl.col(c).cast(pl.Utf8).str.len_chars().max().fill_null(0).alias(c)
        for c in pivot_df_all.columns
    ]).row(0, named=True)

    # 除外集計シート用の列幅マッピング
    max_lens_excl = excluded_pivot_df.select([
        pl.col(c).cast(pl.Utf8).str.len_chars().max().fill_null(0).alias(c)
        for c in excluded_pivot_df.columns
    ]).row(0, named=True)

    del df, attendance_metrics_df, excluded_df
    gc.collect()

    # 戻り値の末尾に2つの文字数マッピング辞書を追加
    return (
        pivot_df_all, excluded_pivot_df, classroom_summary_df, excluded_classroom_summary_df,
        output_file, date_range_str, sheet_name, week_period_labels_all, week_period_labels_excl,
        max_lens_all, max_lens_excl
    )

def show_error_dialog(title: str, message: str, parent: tk.Tk | None = None) -> None:
    dlg = tk.Toplevel(parent) if parent else tk.Toplevel()
    dlg.title(title)
    try: scale = float(parent.tk.call("tk", "scaling")) if parent else float(dlg.tk.call("tk", "scaling"))
    except Exception: scale = 1.0
    base_font = tkfont.nametofont("TkDefaultFont")
    dlg_font = tkfont.Font(family=base_font.cget("family"), size=max(int(base_font.cget("size") * scale), 10))
    dlg.geometry(f"{int(700*scale)}x{int(360*scale)}")
    txt = tk.Text(dlg, wrap="word", font=dlg_font)
    txt.insert("1.0", message)
    txt.configure(state="disabled")
    txt.pack(expand=True, fill="both", padx=int(6 * scale), pady=int(6 * scale))
    frm = tk.Frame(dlg); frm.pack(fill="x", padx=int(6 * scale), pady=int(6 * scale))
    def _copy():
        try: dlg.clipboard_clear(); dlg.clipboard_append(message)
        except Exception: pass
    tk.Button(frm, text="エラーメッセージをコピー", command=_copy, font=dlg_font).pack(side="left")
    tk.Button(frm, text="閉じる", command=dlg.destroy, font=dlg_font).pack(side="right")
    try: dlg.transient(parent); dlg.grab_set(); dlg.wait_window()
    except Exception: pass
def format_excel_fast(
    all_pivot_df: pl.DataFrame,
    excluded_pivot_df: pl.DataFrame,
    classroom_summary_df: pl.DataFrame,
    excluded_classroom_summary_df: pl.DataFrame,
    output_file: str,
    date_range_str: str,
    sheet_name: str,
    week_period_labels_all: list[str],
    week_period_labels_excl: list[str],
    max_lens_all: dict[str, int],   # 引数を追加
    max_lens_excl: dict[str, int]   # 引数を追加
):
    progress("xlsxwriter単一ストリームによる高速レイアウト処理を開始")
    # 常時メモリに溜め込まず、順次ディスクへフラッシュする定数メモリモードを有効化
    workbook = xlsxwriter.Workbook(output_file, {'constant_memory': True})
    ud_font = select_font()

    fmt_title = workbook.add_format({"font_name": ud_font, "bold": True, "font_size": 11, "align": "left", "valign": "vcenter"})
    fmt_title_center = workbook.add_format({"font_name": ud_font, "bold": True, "font_size": 11, "align": "center", "valign": "vcenter", "bg_color": "#E8F6F3", "border": 1})
    fmt_header = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#D3D3D3", "bold": True, "align": "center", "valign": "vcenter"})
    fmt_white = workbook.add_format({"font_name": ud_font, "border": 1, "valign": "vcenter"})
    fmt_khaki = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#F0E68C", "valign": "vcenter"})
    fmt_lavender = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#E6E6FA", "valign": "vcenter"})
    fmt_white_pct = workbook.add_format({"font_name": ud_font, "border": 1, "num_format": "0.0%", "valign": "vcenter"})
    fmt_cond_orange = workbook.add_format({"font_name": ud_font, "bg_color": "#FFA500", "font_color": "#000000"})
    fmt_cond_firebrick = workbook.add_format({"font_name": ud_font, "bg_color": "#B22222", "font_color": "#FFFFFF"})

    # 内包関数の引数に max_lens_sheet を追加し、シートごとに切り替えられるように設計
    def write_sheet_v7(sheet_df: pl.DataFrame, summary_df: pl.DataFrame, week_period_labels: list[str], sheet_title: str, max_lens_sheet: dict[str, int]):
        progress(f"シート [{sheet_title}] のデータ一括合成中...")
        headers = list(sheet_df.columns)
        max_col, nrows = len(headers) - 1, sheet_df.height
        header_row, data_start = 1, 2

        col_values: dict[str, list] = {}
        for col in headers:
            try: s = sheet_df.get_column(col)
            except Exception: s = pl.Series(col, [None] * nrows)
            if col == "No":
                vals = [None if v is None else str(v).strip() for v in s.to_list()]
                col_values[col] = [None if v is None else (v.lstrip("0") if v.lstrip("0") != "" else v) for v in vals]
            elif any(k in col for k in ("授業", "Lap数", "テスト数", "総授業数")):
                col_values[col] = [0.0 if v is None else float(v) for v in s.cast(pl.Float64, strict=False).fill_null(0).to_list()]
            elif ("平均" in col) or ("％" in col) or ("率" in col):
                col_values[col] = [None if v is None else float(v) for v in s.cast(pl.Float64, strict=False).to_list()]
            else:
                col_values[col] = [None if v is None else str(v) for v in s.to_list()]

        lesson_nums = []
        for h in headers:
            m = re.search(r"授業(\d+)", str(h))
            if m:
                lesson_nums.append(int(m.group(1)))
        lesson_nums = sorted(list(set(lesson_nums)))
        for i in lesson_nums:
            if f"授業{i}" not in col_values: continue
            for idx, lv in enumerate(col_values[f"授業{i}"]):
                if lv == 0 or lv == 0.0:
                    for c in [f"平均HW{i}", f"Lap数{i}", f"Lap％{i}", f"テスト数{i}", f"テスト％{i}"]:
                        if c in col_values: col_values[c][idx] = None

        if nrows > 0:
            row_classrooms = col_values.get("教室", [""] * nrows)
            unique_classrooms = sorted(list(set([v for v in row_classrooms if v])))
            final_col_values: dict[str, list] = {c: [] for c in headers}

            # 全体平均行
            final_col_values["No"].append("0"); final_col_values["氏名"].append("AVG"); final_col_values["教室"].append("FS")
            final_col_values["総授業数"].append(sum(v for v in col_values["総授業数"] if isinstance(v, (int, float))))
            for i in lesson_nums:
                w_lessons = sum(v for v in col_values[f"授業{i}"] if isinstance(v, (int, float)))
                final_col_values[f"授業{i}"].append(w_lessons)
                if w_lessons == 0:
                    for c in [f"Lap％{i}", f"テスト％{i}", f"平均HW{i}"]: final_col_values[c].append(None)
                    for c in [f"Lap数{i}", f"テスト数{i}"]: final_col_values[c].append(0.0)
                else:
                    wl, wt = sum(v for v in col_values[f"Lap数{i}"] if isinstance(v, (int, float))), sum(v for v in col_values[f"テスト数{i}"] if isinstance(v, (int, float)))
                    final_col_values[f"Lap数{i}"].append(wl); final_col_values[f"テスト数{i}"].append(wt)
                    final_col_values[f"Lap％{i}"].append(wl / w_lessons); final_col_values[f"テスト％{i}"].append(wt / w_lessons)
                    thw = sum(col_values[f"平均HW{i}"][idx]*col_values[f"授業{i}"][idx] for idx in range(nrows) if isinstance(col_values[f"授業{i}"][idx], (int, float)) and isinstance(col_values[f"平均HW{i}"][idx], (int, float)))
                    final_col_values[f"平均HW{i}"].append(round(thw / w_lessons, 1))

            # 教室別ループ
            for target_cls in unique_classrooms:
                cls_indices = [idx for idx, cls in enumerate(row_classrooms) if cls == target_cls]
                try:
                    cls_sum_row = summary_df.filter(pl.col("教室").cast(pl.Utf8).str.strip_chars() == target_cls)
                    cls_total_lessons = float(cls_sum_row.row(0)[1]) if cls_sum_row.height == 1 else 0.0
                except Exception:
                    cls_total_lessons = sum(col_values["総授業数"][idx] for idx in cls_indices)

                final_col_values["No"].append("0"); final_col_values["氏名"].append("AVG"); final_col_values["教室"].append(target_cls)
                final_col_values["総授業数"].append(cls_total_lessons)

                for i in lesson_nums:
                    w_lessons_cls = sum(col_values[f"授業{i}"][idx] for idx in cls_indices if isinstance(col_values[f"授業{i}"][idx], (int, float)))
                    final_col_values[f"授業{i}"].append(w_lessons_cls)
                    if w_lessons_cls == 0:
                        for c in [f"Lap％{i}", f"テスト％{i}", f"平均HW{i}"]: final_col_values[c].append(None)
                        for c in [f"Lap数{i}", f"テスト数{i}"]: final_col_values[c].append(0.0)
                    else:
                        wl_c = sum(col_values[f"Lap数{i}"][idx] for idx in cls_indices if isinstance(col_values[f"Lap数{i}"][idx], (int, float)))
                        wt_c = sum(col_values[f"テスト数{i}"][idx] for idx in cls_indices if isinstance(col_values[f"テスト数{i}"][idx], (int, float)))
                        final_col_values[f"Lap数{i}"].append(wl_c); final_col_values[f"テスト数{i}"].append(wt_c)
                        final_col_values[f"Lap％{i}"].append(wl_c / w_lessons_cls); final_col_values[f"テスト％{i}"].append(wt_c / w_lessons_cls)
                        thw_c = sum(col_values[f"平均HW{i}"][idx]*col_values[f"授業{i}"][idx] for idx in cls_indices if isinstance(col_values[f"授業{i}"][idx], (int, float)) and isinstance(col_values[f"平均HW{i}"][idx], (int, float)))
                        final_col_values[f"平均HW{i}"].append(round(thw_c / w_lessons_cls, 1))

                for idx in cls_indices:
                    for col in headers: final_col_values[col].append(col_values[col][idx])
            col_values, nrows = final_col_values, len(final_col_values["No"])

        for col, vals in col_values.items():
            for j, v in enumerate(vals):
                if v is None: vals[j] = "-"

        worksheet = workbook.add_worksheet(sheet_title)

        # 【超高速Autofit + レイアウト一括制御】
        for col_idx, col_name in enumerate(headers):
            # 修正：Pylanceエラーの原因だった未定義変数を、引数のマッピング参照（O(1)）に修正
            max_val_len = max_lens_sheet.get(col_name, 0)
            max_len = max(max_val_len, len(col_name))
            if col_idx == 0: col_width = max(max_len * 2.0, 6)
            elif col_idx == 1: col_width = max(max_len * 1.8, 14)
            elif col_idx == 2: col_width = max(max_len * 1.8, 12)
            elif col_idx == 3: col_width = max(max_len * 1.5, 11)
            else: col_width = 9

            if col_idx in (0, 1, 2): worksheet.set_column(col_idx, col_idx, col_width, fmt_white)
            elif col_idx == 3: worksheet.set_column(col_idx, col_idx, col_width, fmt_lavender)
            else:
                offset = col_idx - (4 + ((col_idx - 4) // 6) * 6)
                if offset == 0: worksheet.set_column(col_idx, col_idx, col_width, fmt_khaki)
                elif offset in (3, 5): worksheet.set_column(col_idx, col_idx, col_width, fmt_white_pct)
                else: worksheet.set_column(col_idx, col_idx, col_width, fmt_white)

        # 1行ずつデータを辞書から引き出して横方向に一括配置（これですべての列が正常に出現します）
        for r_idx in range(nrows):
            row_data = [col_values[col_name][r_idx] for col_name in headers]
            # data_start (2行目) から順に下へ向かって行データを書き出す
            worksheet.write_row(data_start + r_idx, 0, row_data)

        try: worksheet.merge_range(0, 0, 0, min(3, max_col), date_range_str, fmt_title)
        except Exception: worksheet.write(0, 0, date_range_str, fmt_title)
        worksheet.set_row(0, 20)
        worksheet.write_row(header_row, 0, headers, fmt_header)

        for i, period_str in enumerate(week_period_labels):
            start_col = 4 + i * 6
            if start_col + 5 <= max_col: worksheet.merge_range(0, start_col, 0, start_col + 5, period_str, fmt_title_center)

        last_row = data_start + nrows - 1
        worksheet.autofilter(header_row, 0, last_row, max_col)

        for c_base in range(4, len(headers), 6):
            if c_base + 5 > max_col: break
            worksheet.conditional_format(data_start, c_base + 1, last_row, c_base + 1, {"type": "cell", "criteria": "<", "value": 6, "format": fmt_cond_orange})
            worksheet.conditional_format(data_start, c_base + 3, last_row, c_base + 3, {"type": "cell", "criteria": "<", "value": 0.7, "format": fmt_cond_firebrick})
            worksheet.conditional_format(data_start, c_base + 5, last_row, c_base + 5, {"type": "cell", "criteria": "<", "value": 0.7, "format": fmt_cond_firebrick})

        worksheet.freeze_panes(data_start, 4)

        worksheet.set_portrait()
        worksheet.set_paper(9) # A4
        margin_inch = 1.0 / 2.54
        worksheet.set_margins(left=margin_inch, right=margin_inch, top=margin_inch, bottom=margin_inch)
        worksheet.set_header('', {'margin': 0.5 / 2.54})
        worksheet.set_footer('', {'margin': 0.5 / 2.54})
        worksheet.repeat_columns(0, 3)
        worksheet.repeat_rows(0, 2)

        v_breaks = []
        current_break_col = 4 + 6
        while current_break_col < len(headers):
            v_breaks.append(current_break_col)
            current_break_col += 6
        worksheet.set_v_pagebreaks(v_breaks)
        worksheet.print_area(0, 0, last_row, max_col)
        worksheet.fit_to_pages(1, 0)
        progress(f"シート [{sheet_title}] の全レイアウト・印刷設定が一発で完了しました。")

    # 修正：それぞれのシートに対応する辞書（max_lens_all, max_lens_excl）を渡す
    write_sheet_v7(all_pivot_df, classroom_summary_df, week_period_labels_all, "全集計", max_lens_all)
    write_sheet_v7(excluded_pivot_df, excluded_classroom_summary_df, week_period_labels_excl, "除外集計", max_lens_excl)

    progress("Excelブックのクローズ処理中...")
    workbook.close()
    progress("GCを実行中")
    del all_pivot_df, excluded_pivot_df, classroom_summary_df, excluded_classroom_summary_df
    gc.collect()


def main():
    enable_windows_dpi_awareness()
    root = tk.Tk()
    configure_tk_scaling(root)
    root.withdraw()
    try:
        root.attributes("-topmost", True); root.lift(); root.focus_force()
        input_file = filedialog.askopenfilename(title="集計元のExcelファイルを選択してください", filetypes=[("Excel", "*.xlsx"), ("すべてのファイル", "*.*")], parent=root)
        root.attributes("-topmost", False)
        if not input_file: return

        total_start_time = time.perf_counter()

        # 修正：戻り値に増えた辞書（max_lens_all, max_lens_excl）を正しくアンパックして受け取る
        (
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
            max_lens_all,
            max_lens_excl
        ) = process_attendance_data(input_file)

        # 修正：引数のミスマッチ（9個必要なところに足りていなかった問題）を11個に増やして完全適合
        format_excel_fast(
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
            max_lens_all,
            max_lens_excl
        )

        total_elapsed = time.perf_counter() - total_start_time
        processed_classroom_count = classroom_summary_df.height
        try: processed_lesson_count = int(classroom_summary_df.get_column("授業数").sum())
        except Exception: processed_lesson_count = 0

        result_msg = f"処理が完了しました！\n実行時間: {total_elapsed:.2f}秒\n\n処理教室数: {processed_classroom_count}教室\n授業数: {processed_lesson_count}件\n\n出力先:\n{output_file}"
        root.attributes("-topmost", True); messagebox.showinfo("集計完了", result_msg, parent=root); root.attributes("-topmost", False)
        progress(f"処理完了(実行時間: {total_elapsed:.2f}秒)")
    except KeyboardInterrupt: pass
    except Exception as e:
        tb = traceback.format_exc()
        try: root.attributes("-topmost", True); show_error_dialog("処理中にエラーが発生しました", tb, parent=root)
        except Exception: messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}", parent=root)
    finally: root.destroy()
if __name__ == "__main__": main()