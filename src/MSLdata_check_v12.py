import polars as pl
import tkinter as tk
from tkinter import filedialog, messagebox
import xlsxwriter
import os
from datetime import datetime
import re
import unicodedata
import sys
import ctypes
import tkinter.font as tkfont
import traceback
import time
import gc

def progress(msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    print(f"[{ts}] {msg}")

def dismiss_splash() -> None:
    try:
        if sys.platform == "win32":
            import ctypes
            if hasattr(ctypes, "windll") and hasattr(ctypes.windll, "user32"):
                ctypes.windll.user32.PostMessageW(0xFFFF, 0x0400 + 5, 0, 0)
    except Exception:
        pass

EXCLUDED_HW_TERMS = [
    "タンゴ", "単語", "シスタン", "シス単", "ターゲット", "リープ",
    "leap", "ゲットスルー", "パスタン", "パス単", "マドンナ"
]

FONT_PREFERRED = ["Noto Sans JP", "Arial Unicode MS", "Arial"]

def select_font(preferred: list[str] | None = None) -> str:
    prefs = preferred or FONT_PREFERRED
    try:
        root_tmp = tk.Tk()
        root_tmp.withdraw()
        fams = set(tkfont.families())
        root_tmp.destroy()
        for f in prefs:
            if f in fams: return f
    except Exception:
        pass
    for f in prefs: return f
    return "Calibri"

def normalize_text_for_matching(s: object) -> str:
    if s is None: return ""
    t = str(s)
    t = unicodedata.normalize("NFKC", t).lower()
    t = re.sub(r"\s+", "", t)
    t = "".join(chr(ord(c) + 96) if "ぁ" <= c <= "ん" else c for c in t)
    return t

def enable_windows_dpi_awareness() -> None:
    if sys.platform != "win32": return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

def build_hw_pages_expr(hw_name_column: str, hw_start_column: str, hw_end_column: str, excluded_terms: list[str], page_limit: int = 30) -> pl.Expr:
    hw_name_expr = pl.col(hw_name_column).cast(pl.Utf8).fill_null("").str.strip_chars()
    active_excluded_hw_terms = [term for term in excluded_terms if str(term).strip()]
    is_excluded_hw_expr = pl.any_horizontal([hw_name_expr.str.contains(term, literal=True) for term in active_excluded_hw_terms]) if active_excluded_hw_terms else pl.lit(False)
    hw_start_page_expr = pl.col(hw_start_column).cast(pl.Utf8).str.extract(r"(\d+)").cast(pl.Int64, strict=False)
    hw_end_page_expr = pl.col(hw_end_column).cast(pl.Utf8).str.extract(r"(\d+)").cast(pl.Int64, strict=False)
    hw_page_count_expr = hw_end_page_expr - hw_start_page_expr + 1
    capped_hw_page_count_expr = pl.when(hw_page_count_expr.is_null()).then(0).when(hw_page_count_expr < 0).then(0).when(hw_page_count_expr > page_limit).then(page_limit).otherwise(hw_page_count_expr)
    return pl.when(hw_name_expr != "").then(pl.when(~is_excluded_hw_expr).then(capped_hw_page_count_expr).otherwise(0)).otherwise(0)

def build_test_presence_exprs(lap_column_names: list[str], kaku_column_names: list[str]) -> tuple[pl.Expr, pl.Expr, pl.Expr]:
    def _presence_expr(column_names: list[str]) -> pl.Expr:
        presence_checks = [pl.col(column_name).is_not_null() & (pl.col(column_name).cast(pl.Utf8).str.strip_chars() != "") for column_name in column_names]
        return pl.any_horizontal(presence_checks).fill_null(False) if presence_checks else pl.lit(False)
    lap_expr = _presence_expr(lap_column_names)
    kaku_expr = _presence_expr(kaku_column_names)
    return lap_expr.alias("has_lap"), kaku_expr.alias("has_kaku"), (lap_expr | kaku_expr).fill_null(False).alias("has_test")

def configure_tk_scaling(root: tk.Tk) -> None:
    try:
        scale = float(root.tk.call("tk", "scaling"))
        root.tk.call("tk", "scaling", scale)
    except Exception:
        pass

def load_sheet_preserve_extra_columns(path: str, sheet_index: int = 1) -> pl.DataFrame:
    try:
        import fastexcel
        excel_reader = fastexcel.read_excel(path)
        sheet = excel_reader.load_sheet(sheet_index - 1, header_row=None)
        df_raw = sheet.to_polars()
        if df_raw.height == 0: return pl.DataFrame()
        progress("fastexcel読み込みモード")
        header = [str(x) if x is not None else "" for x in df_raw.row(0)]
        hw_base_idx = -1
        for i, h in enumerate(header):
            if "宿題名" in h:
                hw_base_idx = i
                break
        if hw_base_idx != -1:
            pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
            k, pattern_idx = 2, 0
            for i in range(hw_base_idx + 1, len(header)):
                if header[i].strip() == "":
                    header[i] = f"{pattern[pattern_idx]}_{k}"
                    pattern_idx += 1
                    if pattern_idx >= 3:
                        pattern_idx, k = 0, k + 1
        else:
            idx = 1
            for i in range(len(header)):
                if header[i].strip() == "":
                    header[i] = f"追加列_{idx}"
                    idx += 1
        unique_headers = []
        seen = set()
        for h in header:
            h = h.strip()
            if h == "": h = "無名列"
            original_h, counter = h, 1
            while h in seen:
                h = f"{original_h}_{counter}"
                counter += 1
            seen.add(h)
            unique_headers.append(h)
        df_data = df_raw.slice(1)
        return df_data.rename(dict(zip(df_data.columns, unique_headers)))
    except Exception as e:
        progress(f"本番Excelフォールバック読み込み: {e}")
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[sheet_index - 1]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()
        if not rows: return pl.DataFrame()
        header = [str(c) if c is not None else "" for c in rows[0]]
        max_cols = max(len(r) for r in rows)
        if max_cols > len(header):
            try:
                base = header.index("宿題名")
                pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
                k = 2
                while len(header) < max_cols:
                    for p in pattern:
                        header.append(f"{p}_{k}")
                        if len(header) >= max_cols: break
                    k += 1
            except ValueError:
                idx = 1
                while len(header) < max_cols:
                    header.append(f"追加列_{idx}")
                    idx += 1
        header_len = len(header)
        padded_rows = [(tuple(row) + (None,) * (header_len - len(row)) if len(row) < header_len else tuple(row)) for row in rows[1:]]
        return pl.DataFrame(padded_rows, schema=header, orient="row")

def get_jp_weekday(dt_obj):
    return ["月", "火", "水", "木", "金", "土", "日"][dt_obj.weekday()]

def calculate_visual_width(text: str) -> float:
    if not text: return 0.0
    width = 0.0
    for char in text:
        status = unicodedata.east_asian_width(char)
        if status in ("W", "F", "A"):
            width += 2.0
        else:
            width += 1.0
    return width

def process_attendance_data(input_file: str, start_from_saturday: bool) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame, pl.DataFrame, str, str, str, list[str], list[str], dict[str, int], dict[str, int], dict[str, str]]:
    progress("出欠データの解析開始")
    df = load_sheet_preserve_extra_columns(input_file, sheet_index=1)
    cols = df.columns

    if "教室" in cols:
        progress("教室名から'FS'の除去")
        df = df.with_columns(pl.col("教室").cast(pl.Utf8).str.replace_all(r"(ＦＳ|FS|ｆｓ|fs)", "").str.strip_chars().alias("教室"))

    hw_name_cols = [c for c in cols if c and "宿題名" in c]
    if hw_name_cols:
        progress(f"宿題名列(計{len(hw_name_cols)}列)の正規化を実行")
        norm_series_list = []
        for c in hw_name_cols:
            try: orig_vals = df.get_column(c).to_list()
            except Exception: orig_vals = []
            norm_series_list.append(pl.Series(f"{c}__norm", [normalize_text_for_matching(v) for v in orig_vals]))
        if norm_series_list:
            df = df.with_columns(norm_series_list)
            cols = df.columns

    possible_date_cols = ["日付", "授業日"]
    date_col = next((dc for dc in possible_date_cols if dc in cols), None)
    valid_dates = df.select(pl.col(date_col).cast(pl.Date, strict=False)).drop_nulls() if date_col else pl.DataFrame()

    if valid_dates.height > 0:
        min_dt, max_dt = valid_dates.min().item(), valid_dates.max().item()
        min_date_file, max_date_file = min_dt.strftime("%Y%m%d"), max_dt.strftime("%Y%m%d")

        min_yy = min_dt.strftime("%y")
        max_yy = max_dt.strftime("%y")
        date_range_str = (
            f"期間:{min_yy}/{min_dt.month}/{min_dt.day}({get_jp_weekday(min_dt)}) ～ "
            f"{max_yy}/{max_dt.month}/{max_dt.day}({get_jp_weekday(max_dt)})"
        )
    else:
        min_date_file, max_date_file = "不明", "不明"
        date_range_str = "期間:不明"

    sheet_name = "集計結果"
    output_filename = f"集計結果_[{min_date_file}-{max_date_file}]_{datetime.now().strftime('%y%m%d-%H%M%S')}.xlsx"
    output_file = os.path.join(os.path.dirname(input_file), output_filename)

    if date_col:
        parsed_date_expr = pl.col(date_col).cast(pl.Date, strict=False)
        if start_from_saturday:
            dt_iso = parsed_date_expr.dt.offset_by("2d").dt.week()
        else:
            dt_iso = parsed_date_expr.dt.week()
    else:
        dt_iso = pl.lit(None)

    lap_section_start_index = next((i for i, c in enumerate(cols) if c and ("ラップテスト" in c or "ラップ" in c)), 0)
    hw_section_start_index = next((i for i, c in enumerate(cols) if c and "宿題名" in c), len(cols))

    test_section_column_names = cols[lap_section_start_index:hw_section_start_index]
    lap_count_column_names = [c for c in test_section_column_names if "ラップ" in c and "（分子）" in c]
    confirmation_count_column_names = [c for c in test_section_column_names if "確認" in c and "（分子）" in c]

    hw_page_exprs = []
    for i in range(hw_section_start_index, len(cols), 3):
        if i + 2 < len(cols):
            hw_name_for_expr = f"{cols[i]}__norm" if f"{cols[i]}__norm" in cols else cols[i]
            hw_page_exprs.append(build_hw_pages_expr(hw_name_for_expr, cols[i+1], cols[i+2], EXCLUDED_HW_TERMS))

    if any(req not in cols for req in ["出欠", "担当講師名", "担当講師N0", "教室"]):
        raise ValueError("必須の列が見つかりません")

    progress("出席データの抽出および宿題・テスト集計ロジックを構築中")
    attendance_metrics_df = df.filter(pl.col("出欠") == "出席").with_columns(
        pl.col("担当講師名").cast(pl.Utf8).str.replace_all(r"[Ａ-Ｚ_]", "").alias("担当講師名"),
        dt_iso.cast(pl.Utf8).alias("week_num"),
        (pl.col(date_col).cast(pl.Date, strict=False) if date_col else pl.lit(None)).alias("date_val"),
        (pl.sum_horizontal(hw_page_exprs) if hw_page_exprs else pl.lit(0)).alias("hw_pages"),
    ).with_columns(*build_test_presence_exprs(lap_count_column_names, confirmation_count_column_names))

    week_period_by_week_number = {}
    if "date_val" in attendance_metrics_df.columns:
        week_ranges_df = attendance_metrics_df.filter(pl.col("date_val").is_not_null()).group_by("week_num").agg(pl.col("date_val").min().alias("min_dt"), pl.col("date_val").max().alias("max_dt"))
        for row in week_ranges_df.iter_rows(named=True):
            md, xd = row["min_dt"], row["max_dt"]
            if md and xd: week_period_by_week_number[str(row["week_num"])] = f"{md.strftime('%Y/%m/%d')}({get_jp_weekday(md)}) ～ {xd.strftime('%Y/%m/%d')}({get_jp_weekday(xd)})"

    global_rename_map = {}

    def aggregate_and_pivot(metrics_df: pl.DataFrame):
        aggregated = metrics_df.group_by(["担当講師N0", "担当講師名", "教室", "week_num"]).agg(
            pl.len().alias("授業数"), pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"), pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数")
        ).with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率")
        ).select(["担当講師N0", "担当講師名", "教室", "week_num", "授業数", "宿題平均", "ラップ回数", "ラップ率", "テスト回数", "テスト率"])

        pivot_df = aggregated.pivot(on="week_num", index=["担当講師N0", "担当講師名", "教室"], values=["授業数", "宿題平均", "ラップ回数", "ラップ率", "テスト回数", "テスト率"])
        week_nums = sorted(list(set([c.split("_")[-1] for c in pivot_df.columns if "_" in c])))
        pivot_df = pivot_df.with_columns(pl.sum_horizontal([pl.col(f"授業数_{w}") for w in week_nums if f"授業数_{w}" in pivot_df.columns]).fill_null(0).alias("総授業数"))

        for w in week_nums:
            c_class, c_hw, c_lap_cnt, c_lap_rate, c_test_cnt, c_test_rate = f"授業数_{w}", f"宿題平均_{w}", f"ラップ回数_{w}", f"ラップ率_{w}", f"テスト回数_{w}", f"テスト率_{w}"
            if c_class in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.col(c_class).fill_null(0).alias(c_class))
            if c_lap_cnt in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.col(c_lap_cnt).fill_null(0).alias(c_lap_cnt))
            if c_test_cnt in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.col(c_test_cnt).fill_null(0).alias(c_test_cnt))
            if c_hw in pivot_df.columns and c_class in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_hw)).alias(c_hw))
            if c_lap_rate in pivot_df.columns and c_class in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_lap_rate)).alias(c_lap_rate))
            if c_test_rate in pivot_df.columns and c_class in pivot_df.columns: pivot_df = pivot_df.with_columns(pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_test_rate)).alias(c_test_rate))

        desired_order = ["担当講師N0", "担当講師名", "教室", "総授業数"]
        rename_map = {"担当講師N0": "No", "担当講師名": "氏名", "教室": "教室", "総授業数": "総授業"}
        week_labels = []
        for i, w in enumerate(week_nums, start=1):
            desired_order.extend([f"授業数_{w}", f"宿題平均_{w}", f"ラップ回数_{w}", f"ラップ率_{w}", f"テスト回数_{w}", f"テスト率_{w}"])
            rename_map[f"授業数_{w}"], rename_map[f"宿題平均_{w}"], rename_map[f"ラップ回数_{w}"] = f"授業_{i}", f"平均HW_{i}", f"Lap回数_{i}"
            rename_map[f"ラップ率_{w}"], rename_map[f"テスト回数_{w}"], rename_map[f"テスト率_{w}"] = f"Lap％_{i}", f"テスト回数_{i}", f"テスト％_{i}"
            week_labels.append(week_period_by_week_number.get(str(w), f"Week {i}"))

        global_rename_map.update(rename_map)
        exist_cols = [c for c in desired_order if c in pivot_df.columns]
        return pivot_df.select(exist_cols).rename({k: v for k, v in rename_map.items() if k in exist_cols}).sort(["教室", "No"]), week_labels

    progress("「全集計」用データのマトリクス変換を実行中")
    pivot_df_all, week_period_labels_all = aggregate_and_pivot(attendance_metrics_df)

    grade_excl = pl.col("学年").cast(pl.Utf8).str.contains(r"(高3|高３|高卒)").fill_null(False) if "学年" in cols else pl.lit(False)
    subj_excl = pl.col("科目").cast(pl.Utf8).str.contains(r"(国語|小論文)").fill_null(False) if "科目" in cols else pl.lit(False)

    progress("「除外済み集計」用データの抽出と集計を実行中")
    excluded_df = attendance_metrics_df.filter(~grade_excl & ~subj_excl)
    excluded_pivot_df, week_period_labels_excl = aggregate_and_pivot(excluded_df)

    def get_summary(m_df):
        return m_df.group_by("教室").agg(
            pl.len().alias("授業数"), pl.col("担当講師N0").n_unique().alias("講師数"), pl.col("hw_pages").sum().alias("宿計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラ回"), pl.col("has_test").cast(pl.Int32).sum().alias("テ回")
        ).with_columns((pl.col("宿計")/pl.col("授業数")).round(1).alias("宿題平均"), (pl.col("ラ回")/pl.col("授業数")).alias("ラップ率"), (pl.col("テ回")/pl.col("授業数")).alias("テスト率")).sort("教室")

    classroom_summary_df = get_summary(attendance_metrics_df)
    excluded_classroom_summary_df = get_summary(excluded_df)

    progress("全列の視覚的文字幅によるAutofit計算を実行中...")

    def get_visual_max_lens(pdf: pl.DataFrame) -> dict[str, int]:
        max_lens = {}
        for c in pdf.columns:
            vals = pdf.get_column(c).cast(pl.Utf8).fill_null("").to_list()
            max_w = 0.0
            for v in vals:
                w = calculate_visual_width(v)
                if w > max_w: max_w = w
            max_lens[c] = int(max_w)
        return max_lens

    max_lens_all = get_visual_max_lens(pivot_df_all)
    max_lens_excl = get_visual_max_lens(excluded_pivot_df)

    del df, attendance_metrics_df, excluded_df
    gc.collect()

    return (
        pivot_df_all, excluded_pivot_df, classroom_summary_df, excluded_classroom_summary_df,
        output_file, date_range_str, sheet_name, week_period_labels_all, week_period_labels_excl,
        max_lens_all, max_lens_excl, global_rename_map
    )

def show_error_dialog(title: str, message: str, parent: tk.Tk | None = None) -> None:
    dlg = tk.Toplevel(parent) if parent else tk.Toplevel()
    dlg.title(title)
    try: scale = float(parent.tk.call("tk", "scaling")) if parent else float(dlg.tk.call("tk", "scaling"))
    except Exception: scale = 1.0
    base_font = tkfont.nametofont("TkDefaultFont")
    dlg_font = tkfont.Font(family=base_font.cget("family"), size=max(int(base_font.cget("size") * scale), 10))
    dlg.geometry(f"{int(700*scale)}x{int(360*scale)}")
    txt = tk.Text(dlg, wrap="word", font=dlg_font)
    txt.insert("1.0", message)
    txt.configure(state="disabled")
    txt.pack(expand=True, fill="both", padx=int(6 * scale), pady=int(6 * scale))
    frm = tk.Frame(dlg); frm.pack(fill="x", padx=int(6 * scale), pady=int(6 * scale))
    def _copy():
        try: dlg.clipboard_clear(); dlg.clipboard_append(message)
        except Exception: pass
    tk.Button(frm, text="エラーメッセージをコピー", command=_copy, font=dlg_font).pack(side="left")
    tk.Button(frm, text="閉じる", command=dlg.destroy, font=dlg_font).pack(side="right")
    try: dlg.transient(parent); dlg.grab_set(); dlg.wait_window()
    except Exception: pass

def get_column_style_config(orig_key: str, clean_name: str, max_val_len: int) -> tuple[float, str, str | None]:
    if clean_name in ("No", "氏名", "教室"):
        max_len = float(max_val_len)
        if clean_name == "No": return max(max_len + 1.5, 5.0), "white", None
        if clean_name == "氏名": return max(max_len + 1.5, 11.0), "white", None
        if clean_name == "教室": return max(max_len + 1.5, 9.0), "white", None

    header_visual_width = calculate_visual_width(clean_name)
    max_len = header_visual_width

    if clean_name == "総授業": return max(max_len + 2.0, 9.0), "lavender", None

    col_width = max(max_len + 2.0, 8.0)
    if "授業数" in orig_key: return col_width, "khaki", None
    if any(k in orig_key for k in ("率", "％")): return col_width, "white_pct", "pct"
    return col_width, "white", None

def format_excel_fast(
    all_pivot_df: pl.DataFrame,
    excluded_pivot_df: pl.DataFrame,
    classroom_summary_df: pl.DataFrame,
    excluded_classroom_summary_df: pl.DataFrame,
    output_file: str,
    date_range_str: str,
    sheet_name: str,
    week_period_labels_all: list[str],
    week_period_labels_excl: list[str],
    max_lens_all: dict[str, int],
    max_lens_excl: dict[str, int],
    global_rename_map: dict[str, str]
):
    progress("xlsxwriter単一ストリームによる高速レイアウト処理を開始")
    workbook = xlsxwriter.Workbook(output_file, {'constant_memory': True})
    ud_font = select_font()

    fmt_title = workbook.add_format({"font_name": ud_font, "bold": True, "font_size": 11, "align": "left", "valign": "vcenter"})
    fmt_title_center = workbook.add_format({"font_name": ud_font, "bold": True, "font_size": 11, "align": "center", "valign": "vcenter", "bg_color": "#E8F6F3", "border": 1})
    fmt_header = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#D3D3D3", "bold": True, "align": "center", "valign": "vcenter"})
    fmt_system_sub = workbook.add_format({"font_name": ud_font, "border": 1, "font_size": 8, "font_color": "#A0A0A0", "align": "center", "valign": "vcenter"})
    fmt_cond_orange = workbook.add_format({"font_name": ud_font, "bg_color": "#FFA500", "font_color": "#000000"})
    fmt_cond_firebrick = workbook.add_format({"font_name": ud_font, "bg_color": "#B22222", "font_color": "#FFFFFF"})

    formats_matrix = {
        "base": {
            "white": workbook.add_format({"font_name": ud_font, "border": 1, "valign": "vcenter"}),
            "khaki": workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#F0E68C", "valign": "vcenter"}),
            "lavender": workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#E6E6FA", "valign": "vcenter"}),
            "white_pct": workbook.add_format({"font_name": ud_font, "border": 1, "num_format": "0.0%", "valign": "vcenter"})
        },
        "avg": {
            "white": workbook.add_format({"font_name": ud_font, "border": 1, "bold": True, "valign": "vcenter"}),
            "khaki": workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#F0E68C", "bold": True, "valign": "vcenter"}),
            "lavender": workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#E6E6FA", "bold": True, "valign": "vcenter"}),
            "white_pct": workbook.add_format({"font_name": ud_font, "border": 1, "bold": True, "num_format": "0.0%", "valign": "vcenter"})
        }
    }

    reverse_map = {v: k for k, v in global_rename_map.items()}

    def write_sheet_v7(sheet_df: pl.DataFrame, summary_df: pl.DataFrame, week_period_labels: list[str], sheet_title: str, max_lens_sheet: dict[str, int], tab_color: str):
        progress(f"シート [{sheet_title}] のデータ合成中...")
        headers = list(sheet_df.columns)

        lesson_nums = []
        for h in headers:
            orig_h = reverse_map.get(h, h)
            m = re.search(r"授業数_(\d+)", str(orig_h))
            if m:
                lesson_nums.append(int(m.group(1)))
        lesson_nums = sorted(list(set(lesson_nums)))

        progress(f"[{sheet_title}] Polarsネイティブ層による超高速積和演算の適用中...")
        prod_exprs = []
        for i in lesson_nums:
            c_class = global_rename_map.get(f"授業数_{i}", "")
            c_hw = global_rename_map.get(f"宿題平均_{i}", "")
            if c_class and c_hw:
                prod_exprs.append((pl.col(c_class).fill_null(0.0) * pl.col(c_hw).fill_null(0.0)).alias(f"__prod_hw_{i}"))

        if prod_exprs:
            sheet_df = sheet_df.with_columns(prod_exprs)

        nrows = sheet_df.height
        header_row_visual = 1
        header_row_system = 2
        data_start = 3

        progress(f"[{sheet_title}] ステップ1: 各列データのメモリ展開開始")
        col_values: dict[str, list] = {}
        for col in headers:
            try: s = sheet_df.get_column(col)
            except Exception: s = pl.Series(col, [None] * nrows)

            orig_key = reverse_map.get(col, col)
            if col.startswith("No"):
                vals = [None if v is None else str(v).strip() for v in s.to_list()]
                col_values[col] = [None if v is None else (v.lstrip("0") if v.lstrip("0") != "" else v) for v in vals]
            elif any(k in orig_key for k in ("授業数", "ラップ回数", "テスト回数", "総授業数")):
                col_values[col] = [0.0 if v is None else float(v) for v in s.cast(pl.Float64, strict=False).fill_null(0).to_list()]
            elif any(k in orig_key for k in ("平均", "％", "率")):
                col_values[col] = [None if v is None else float(v) for v in s.cast(pl.Float64, strict=False).to_list()]
            else:
                col_values[col] = [None if v is None else str(v) for v in s.to_list()]

        progress(f"[{sheet_title}] ステップ3: 授業数0の不要データのクリア処理")
        for i in lesson_nums:
            c_class = global_rename_map.get(f"授業数_{i}", "")
            if not c_class or c_class not in col_values: continue
            for idx, lv in enumerate(col_values[c_class]):
                if lv == 0 or lv == 0.0:
                    for k in (f"宿題平均_{i}", f"ラップ回数_{i}", f"ラップ率_{i}", f"テスト回数_{i}", f"テスト率_{i}"):
                        c_display = global_rename_map.get(k, "")
                        if c_display in col_values: col_values[c_display][idx] = None

        if nrows > 0:
            progress(f"[{sheet_title}] ステップ4: 教室所属データの事前マッピング生成")

            # 列名を動的に安全に取得
            c_no = global_rename_map.get("担当講師N0", "No")
            c_name = global_rename_map.get("担当講師名", "氏名")
            c_cls = global_rename_map.get("教室", "教室")
            c_total = global_rename_map.get("総授業数", "総授業")

            row_classrooms = col_values.get(c_cls, [""] * nrows)
            unique_classrooms = sorted(list(set([v for v in row_classrooms if v])))
            final_col_values: dict[str, list] = {c: [] for c in headers}

            if c_no in final_col_values: final_col_values[c_no].append("0")
            if c_name in final_col_values: final_col_values[c_name].append("AVG")
            if c_cls in final_col_values: final_col_values[c_cls].append("FS")
            if c_total in final_col_values:
                final_col_values[c_total].append(sum(v for v in col_values.get(c_total, []) if isinstance(v, (int, float))))

            cls_to_indices = {target_cls: [] for target_cls in unique_classrooms}
            for idx, cls in enumerate(row_classrooms):
                if cls in cls_to_indices:
                    cls_to_indices[cls].append(idx)

            progress(f"[{sheet_title}] ステップ5: 全体(FS)のAVG集計をPolarsデータから直接一撃で計算")
            total_w_lessons = {}
            total_wl = {}
            total_wt = {}
            total_thw = {}

            for i in lesson_nums:
                k_class, k_lap, k_test = f"授業数_{i}", f"ラップ回数_{i}", f"テスト回数_{i}"
                c_class, c_lap, c_test = [global_rename_map.get(k, "") for k in (k_class, k_lap, k_test)]

                total_w_lessons[i] = sum(v for v in col_values.get(c_class, []) if isinstance(v, (int, float)))
                total_wl[i] = sum(v for v in col_values.get(c_lap, []) if isinstance(v, (int, float)))
                total_wt[i] = sum(v for v in col_values.get(c_test, []) if isinstance(v, (int, float)))

                if f"__prod_hw_{i}" in sheet_df.columns:
                    total_thw[i] = float(sheet_df.get_column(f"__prod_hw_{i}").sum())
                else:
                    total_thw[i] = 0.0

            for i in lesson_nums:
                k_class, k_lap, k_test, k_lap_r, k_test_r, k_hw = f"授業数_{i}", f"ラップ回数_{i}", f"テスト回数_{i}", f"ラップ率_{i}", f"テスト率_{i}", f"宿題平均_{i}"
                c_class, c_lap, c_test, c_lap_r, c_test_r, c_hw = [global_rename_map.get(k, "") for k in (k_class, k_lap, k_test, k_lap_r, k_test_r, k_hw)]

                if not (c_class and c_lap and c_test and c_lap_r and c_test_r and c_hw): continue
                if c_class not in final_col_values: continue

                w_lessons = total_w_lessons[i]
                final_col_values[c_class].append(w_lessons)
                if w_lessons == 0:
                    for c in [c_lap_r, c_test_r, c_hw]:
                        if c in final_col_values: final_col_values[c].append(None)
                    for c in [c_lap, c_test]:
                        if c in final_col_values: final_col_values[c].append(0.0)
                else:
                    wl = total_wl[i]; wt = total_wt[i]
                    if c_lap in final_col_values: final_col_values[c_lap].append(wl)
                    if c_test in final_col_values: final_col_values[c_test].append(wt)
                    if c_lap_r in final_col_values: final_col_values[c_lap_r].append(wl / w_lessons)
                    if c_test_r in final_col_values: final_col_values[c_test_r].append(wt / w_lessons)
                    if c_hw in final_col_values: final_col_values[c_hw].append(round(total_thw[i] / w_lessons, 1))

            progress(f"[{sheet_title}] ステップ6: 各教室(計{len(unique_classrooms)}教室)ごとのAVG・個別データマージ中...")
            for target_cls in unique_classrooms:
                cls_indices = cls_to_indices[target_cls]
                try:
                    cls_sum_row = summary_df.filter(pl.col("教室").cast(pl.Utf8).str.strip_chars() == target_cls)
                    cls_total_lessons = float(cls_sum_row.row(0)[1]) if cls_sum_row.height == 1 else 0.0
                except Exception:
                    cls_total_lessons = sum(col_values.get(c_total, [])[idx] for idx in cls_indices if isinstance(col_values.get(c_total, [])[idx], (int, float)))

                if c_no in final_col_values: final_col_values[c_no].append("0")
                if c_name in final_col_values: final_col_values[c_name].append("AVG")
                if c_cls in final_col_values: final_col_values[c_cls].append(target_cls)
                if c_total in final_col_values: final_col_values[c_total].append(cls_total_lessons)

                for i in lesson_nums:
                    k_class, k_lap, k_test, k_lap_r, k_test_r, k_hw = f"授業数_{i}", f"ラップ回数_{i}", f"テスト回数_{i}", f"ラップ率_{i}", f"テスト率_{i}", f"宿題平均_{i}"
                    c_class, c_lap, c_test, c_lap_r, c_test_r, c_hw = [global_rename_map.get(k, "") for k in (k_class, k_lap, k_test, k_lap_r, k_test_r, k_hw)]

                    if not (c_class and c_lap and c_test and c_lap_r and c_test_r and c_hw): continue
                    if c_class not in final_col_values: continue

                    w_lessons_cls = 0.0
                    wl_c = 0.0
                    wt_c = 0.0
                    thw_c = 0.0

                    c_class_vals = col_values.get(c_class, [])
                    c_lap_vals = col_values.get(c_lap, [])
                    c_test_vals = col_values.get(c_test, [])

                    prod_vals = sheet_df.get_column(f"__prod_hw_{i}").to_list() if f"__prod_hw_{i}" in sheet_df.columns else None

                    for idx in cls_indices:
                        if idx < len(c_class_vals):
                            cv = c_class_vals[idx]
                            if isinstance(cv, (int, float)):
                                w_lessons_cls += cv
                        if idx < len(c_lap_vals):
                            lv = c_lap_vals[idx]
                            if isinstance(lv, (int, float)):
                                wl_c += lv
                        if idx < len(c_test_vals):
                            tv = c_test_vals[idx]
                            if isinstance(tv, (int, float)):
                                wt_c += tv
                        if prod_vals is not None and idx < len(prod_vals):
                            pv = prod_vals[idx]
                            if isinstance(pv, (int, float)):
                                thw_c += pv

                    final_col_values[c_class].append(w_lessons_cls)
                    if w_lessons_cls == 0:
                        for c in [c_lap_r, c_test_r, c_hw]:
                            if c in final_col_values: final_col_values[c].append(None)
                        for c in [c_lap, c_test]:
                            if c in final_col_values: final_col_values[c].append(0.0)
                    else:
                        if c_lap in final_col_values: final_col_values[c_lap].append(wl_c)
                        if c_test in final_col_values: final_col_values[c_test].append(wt_c)
                        if c_lap_r in final_col_values: final_col_values[c_lap_r].append(wl_c / w_lessons_cls)
                        if c_test_r in final_col_values: final_col_values[c_test_r].append(wt_c / w_lessons_cls)
                        if c_hw in final_col_values: final_col_values[c_hw].append(round(thw_c / w_lessons_cls, 1))

                for idx in cls_indices:
                    for col in headers: final_col_values[col].append(col_values[col][idx])
            col_values, nrows = final_col_values, len(final_col_values.get(c_no, []))

        progress(f"[{sheet_title}] ステップ7: 空白値のハイフン変換処理")
        for col, vals in col_values.items():
            for j, v in enumerate(vals):
                if v is None: vals[j] = "-"

        progress(f"[{sheet_title}] ステップ8: Excelシート作成および列幅AutoFit計算")
        worksheet = workbook.add_worksheet(sheet_title)
        worksheet.set_tab_color(tab_color)

        worksheet.center_horizontally()
        worksheet.set_paper(9)
        worksheet.set_margins(0.5/2.54, 0.5/2.54, 0.75/2.54, 0.75/2.54)

        clean_headers = [re.sub(r"_\d+$", "", h) for h in headers]
        system_col_numbers = [col_idx + 1 for col_idx in range(len(headers))]
        max_col = len(headers) - 1

        for col_idx, col_name in enumerate(headers):
            orig_key = reverse_map.get(col_name, col_name)
            max_val_len = max_lens_sheet.get(col_name, 0)

            width, format_key, _ = get_column_style_config(orig_key, clean_headers[col_idx], max_val_len)

            header_w = calculate_visual_width(clean_headers[col_idx])
            final_w = max(max_val_len + 4, header_w + 4, 12, width)

            base_format = formats_matrix["base"][format_key]
            worksheet.set_column(col_idx, col_idx, final_w, base_format)

        progress(f"[{sheet_title}] ステップ9: タイトル・期間見出し・週見出しの書き込み")
        try: worksheet.merge_range(0, 0, 0, min(3, max_col), date_range_str, fmt_title)
        except Exception: worksheet.write(0, 0, date_range_str, fmt_title)
        worksheet.set_row(0, 20)

        week_block_size = len(lesson_nums)
        first_week_col = next((idx for idx, h in enumerate(headers) if "_" in reverse_map.get(h, h)), 4)
        cols_per_week = (len(headers) - first_week_col) // week_block_size if week_block_size > 0 else 6

        for i, period_str in enumerate(week_period_labels):
            start_col = first_week_col + i * cols_per_week
            if start_col + (cols_per_week - 1) <= max_col:
                worksheet.merge_range(0, start_col, 0, start_col + (cols_per_week - 1), period_str, fmt_title_center)

        worksheet.write_row(header_row_visual, 0, clean_headers, fmt_header)

        worksheet.set_row(header_row_system, 12)
        worksheet.write_row(header_row_system, 0, system_col_numbers, fmt_system_sub)

        progress(f"[{sheet_title}] ステップ10: 実際のデータセル書き込み開始 (計 {nrows} 行)")
        for r_idx in range(nrows):
            is_avg = False
            c_name = global_rename_map.get("担当講師名", "氏名")
            if c_name in col_values and r_idx < len(col_values[c_name]):
                is_avg = (col_values[c_name][r_idx] == "AVG")

            current_row_idx = data_start + r_idx

            if is_avg:
                row_type = "avg"
                for col_idx, col_name in enumerate(headers):
                    val = col_values[col_name][r_idx] if col_name in col_values and r_idx < len(col_values[col_name]) else ""
                    orig_key = reverse_map.get(col_name, col_name)
                    _, format_key, _ = get_column_style_config(orig_key, clean_headers[col_idx], 0)
                    worksheet.write(current_row_idx, col_idx, val, formats_matrix[row_type][format_key])
            else:
                row_data = [col_values[col_name][r_idx] if col_name in col_values and r_idx < len(col_values[col_name]) else "" for col_name in headers]
                worksheet.write_row(current_row_idx, 0, row_data)

        progress(f"[{sheet_title}] ステップ11: 条件付き書式・印刷設定の適用")
        last_row = data_start + nrows - 1
        worksheet.autofilter(header_row_system, 0, last_row, max_col)

        for c_idx, col_name in enumerate(headers):
            orig_key = reverse_map.get(col_name, col_name)
            if "宿題平均" in orig_key:
                worksheet.conditional_format(data_start, c_idx, last_row, c_idx, {"type": "cell", "criteria": "<", "value": 6, "format": fmt_cond_orange})
            elif any(k in orig_key for k in ("ラップ率", "テスト率")):
                worksheet.conditional_format(data_start, c_idx, last_row, c_idx, {"type": "cell", "criteria": "<", "value": 0.7, "format": fmt_cond_firebrick})

        worksheet.freeze_panes(data_start, first_week_col)
        worksheet.set_portrait()
        worksheet.set_paper(9)
        worksheet.repeat_columns(0, first_week_col - 1)
        worksheet.repeat_rows(0, header_row_system)

        v_breaks = []
        current_break_col = first_week_col + cols_per_week
        while current_break_col < len(headers):
            v_breaks.append(current_break_col)
            current_break_col += cols_per_week
        worksheet.set_v_pagebreaks(v_breaks)
        worksheet.print_area(0, 0, last_row, max_col)
        progress(f"シート [{sheet_title}] の書き込み処理完了")

    write_sheet_v7(all_pivot_df, classroom_summary_df, week_period_labels_all, "全集計", max_lens_all, "#4F81BD")
    write_sheet_v7(excluded_pivot_df, excluded_classroom_summary_df, week_period_labels_excl, "除外済み集計", max_lens_excl, "#4B6F44")

    progress("Excelブックのクローズ処理中...")
    workbook.close()

    progress("GCを実行中")
    del all_pivot_df, excluded_pivot_df, classroom_summary_df, excluded_classroom_summary_df
    gc.collect()

def main():
    dismiss_splash()
    enable_windows_dpi_awareness()
    root = tk.Tk()
    configure_tk_scaling(root)
    root.withdraw()
    try:
        root.attributes("-topmost", True); root.lift(); root.focus_force()
        input_file = filedialog.askopenfilename(title="集計元のExcelファイルを選択してください", filetypes=[("集計データ","scube2-lesson-result_*.xlsx"), ("Excel", "*.xlsx"), ("すべてのファイル", "*.*")], parent=root)
        if not input_file:
            root.destroy()
            return

        start_from_saturday = messagebox.askyesno(
            "集計基準の選択",
            "土曜始まり（土〜金）で集計しますか？\n\n【はい】: 土曜始まり\n【いいえ】: 月曜始まり",
            parent=root
        )
        root.attributes("-topmost", False)

        total_start_time = time.perf_counter()

        (
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
            max_lens_all,
            max_lens_excl,
            global_rename_map
        ) = process_attendance_data(input_file, start_from_saturday)

        try:
            if os.path.exists(output_file):
                with open(output_file, "r+"): pass
        except IOError:
            root.attributes("-topmost", True)
            messagebox.showerror("実行エラー", f"出力先のExcelファイルが開かれたままです。\nファイルを閉じてから再度実行してください。\n\n対象ファイル:\n{os.path.basename(output_file)}", parent=root)
            root.destroy()
            return

        format_excel_fast(
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
            max_lens_all,
            max_lens_excl,
            global_rename_map
        )

        total_elapsed = time.perf_counter() - total_start_time
        processed_classroom_count = classroom_summary_df.height
        try: processed_lesson_count = int(classroom_summary_df.get_column("授業数").sum())
        except Exception: processed_lesson_count = 0

        result_msg = (
            f"処理が完了しました！\n実行時間: {total_elapsed:.2f}秒\n\n"
            f"処理教室数: {processed_classroom_count}教室\n授業数: {processed_lesson_count}件\n\n"
            f"出力先:\n{output_file}\n\n"
            f"作成されたファイルを開きますか？"
        )

        root.attributes("-topmost", True)
        open_file = messagebox.askyesno("集計完了", result_msg, parent=root)
        root.attributes("-topmost", False)

        if open_file:
            os.startfile(output_file)

        progress(f"処理完了(実行時間: {total_elapsed:.2f}秒)")
    except KeyboardInterrupt: pass
    except Exception as e:
        tb = traceback.format_exc()
        try: root.attributes("-topmost", True); show_error_dialog("処理中にエラーが発生しました", tb, parent=root)
        except Exception: messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}", parent=root)
    finally: root.destroy()

if __name__ == "__main__": main()