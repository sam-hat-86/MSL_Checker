import polars as pl
import tkinter as tk
from tkinter import filedialog, messagebox
import xlsxwriter
import openpyxl
import os
from datetime import datetime
from typing import Any, cast
import re
import unicodedata
import importlib
import pandas as pd
import sys
import ctypes
import tkinter.font as tkfont
import traceback
import time



def enable_windows_dpi_awareness() -> None:
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


# 進捗表示などの共通ヘルパーは可能なら v4 から再利用する

# Try to import helpers from v4; if missing, provide concrete fallbacks
_v4 = None
try:
    _v4 = importlib.import_module("MSLdata_check_v4")
except Exception:
    _v4 = None

if _v4 is not None and hasattr(_v4, "progress"):
    progress = _v4.progress
else:
    def progress(msg: str) -> None:
        ts = time.strftime("%H:%M:%S")
        print(f"[{ts}] {msg}")

EXCLUDED_HW_TERMS = getattr(_v4, "EXCLUDED_HW_TERMS", [
    "たんご",
    "単語",
    "しすたん",
    "シスタン",
    "シス単",
    "たーげっと",
    "ターゲット",
    "りーぷ",
    "リープ",
    "Leap",
    "LEAP",
    "まどんな",
    "マドンナ",
])

DEFAULT_NORMALIZE_WORKERS = getattr(_v4, "DEFAULT_NORMALIZE_WORKERS", 6)
DEFAULT_NORMALIZE_CHUNKSIZE = getattr(_v4, "DEFAULT_NORMALIZE_CHUNKSIZE", 100)

if _v4 is not None and hasattr(_v4, "normalize_text_for_matching"):
    normalize_text_for_matching = _v4.normalize_text_for_matching
else:
    def normalize_text_for_matching(s: object) -> str:
        if s is None:
            return ""
        t = str(s)
        t = unicodedata.normalize("NFKC", t)
        t = t.lower()
        t = re.sub(r"\s+", "", t)
        return t

if _v4 is not None and hasattr(_v4, "parallel_normalize"):
    parallel_normalize = _v4.parallel_normalize
else:
    def parallel_normalize(values, workers: int = 1, chunksize: int = 100):
        return [normalize_text_for_matching(v) for v in values]

if _v4 is not None and hasattr(_v4, "load_sheet_preserve_extra_columns"):
    load_sheet_preserve_extra_columns = _v4.load_sheet_preserve_extra_columns
else:
    def load_sheet_preserve_extra_columns(path: str, sheet_index: int = 1) -> pl.DataFrame:
        try:
            df_pd = pd.read_excel(path, sheet_name=sheet_index - 1)
            return pl.DataFrame(df_pd)
        except Exception:
            return pl.DataFrame()

if _v4 is not None and hasattr(_v4, "build_hw_pages_expr"):
    build_hw_pages_expr = _v4.build_hw_pages_expr
else:
    def build_hw_pages_expr(hw_name_column: str, hw_start_column: str, hw_end_column: str, excluded_terms: list[str], page_limit: int = 30) -> pl.Expr:
        return pl.lit(0)

if _v4 is not None and hasattr(_v4, "build_test_presence_exprs"):
    build_test_presence_exprs = _v4.build_test_presence_exprs
else:
    def build_test_presence_exprs(lap_column_names: list[str], kaku_column_names: list[str]):
        return pl.lit(False), pl.lit(False), pl.lit(False)

# will prefer v4 implementation after local definition if available

if _v4 is not None and hasattr(_v4, "configure_tk_scaling"):
    configure_tk_scaling = _v4.configure_tk_scaling
else:
    def configure_tk_scaling(root: tk.Tk) -> None:
        try:
            scale = float(root.tk.call("tk", "scaling"))
            root.tk.call("tk", "scaling", scale)
        except Exception:
            pass

def load_sheet_preserve_extra_columns(path: str, sheet_index: int = 1) -> pl.DataFrame:
    try:
        import fastexcel

        excel_reader = fastexcel.read_excel(path)
        sheet = excel_reader.load_sheet(sheet_index - 1, header_row=None)
        df_raw = sheet.to_polars()

        if df_raw.height == 0:
            return pl.DataFrame()

        progress(
            "fastexcelエンジンによる読み込み完了。Polarsネイティブでヘッダー処理を開始します。"
        )

        header = [str(x) if x is not None else "" for x in df_raw.row(0)]

        hw_base_idx = -1
        for i, h in enumerate(header):
            if "宿題名" in h:
                hw_base_idx = i
                break

        if hw_base_idx != -1:
            pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
            k = 2
            pattern_idx = 0
            for i in range(hw_base_idx + 1, len(header)):
                if header[i].strip() == "":
                    header[i] = f"{pattern[pattern_idx]}_{k}"
                    pattern_idx += 1
                    if pattern_idx >= 3:
                        pattern_idx = 0
                        k += 1
        else:
            idx = 1
            for i in range(len(header)):
                if header[i].strip() == "":
                    header[i] = f"追加列_{idx}"
                    idx += 1

        unique_headers = []
        seen = set()
        for h in header:
            h = h.strip()
            if h == "":
                h = "無名列"
            original_h = h
            counter = 1
            while h in seen:
                h = f"{original_h}_{counter}"
                counter += 1
            seen.add(h)
            unique_headers.append(h)

        df_data = df_raw.slice(1)
        df_data = df_data.rename(dict(zip(df_data.columns, unique_headers)))

        progress(f"データ構築完了: 行数={df_data.height}, 列数={df_data.width}")
        return df_data

    except Exception as e:
        progress(f"高速エンジンでの読み込み失敗。通常モードに切り替えます: {e}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[sheet_index - 1]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not rows:
            return pl.DataFrame()

        header = [str(c) if c is not None else "" for c in rows[0]]
        max_cols = max(len(r) for r in rows)

        if max_cols > len(header):
            try:
                base = header.index("宿題名")
                pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
                k = 2
                while len(header) < max_cols:
                    for p in pattern:
                        header.append(f"{p}_{k}")
                        if len(header) >= max_cols:
                            break
                    k += 1
            except ValueError:
                idx = 1
                while len(header) < max_cols:
                    header.append(f"追加列_{idx}")
                    idx += 1

        header_len = len(header)
        padded_rows = [
            (
                tuple(row) + (None,) * (header_len - len(row))
                if len(row) < header_len
                else tuple(row)
            )
            for row in rows[1:]
        ]

        df = pl.DataFrame(padded_rows, schema=header, orient="row")
        return df


def get_jp_weekday(dt_obj):
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    return weekdays[dt_obj.weekday()]

# If v4 provides a specialized implementation, prefer it (override local def)
if _v4 is not None and hasattr(_v4, "get_jp_weekday"):
    get_jp_weekday = _v4.get_jp_weekday


def process_attendance_data(
    input_file: str,
) -> tuple[
    pl.DataFrame,
    pl.DataFrame,
    pl.DataFrame,
    pl.DataFrame,
    str,
    str,
    str,
    list[str],
    list[str],
]:
    df = load_sheet_preserve_extra_columns(input_file, sheet_index=1)
    cols = df.columns

    hw_name_cols = [c for c in cols if c and "宿題名" in c]
    if hw_name_cols:
        norm_series_list = []
        for c in hw_name_cols:
            try:
                orig_vals = df.get_column(c).to_list()
            except Exception:
                orig_vals = []

            try:
                env_workers = os.environ.get("NORMALIZE_WORKERS")
                env_chunks = os.environ.get("NORMALIZE_CHUNKSIZE")

                if env_workers and env_workers.isdigit():
                    cpu = max(1, int(env_workers))
                else:
                    cpu = max(1, DEFAULT_NORMALIZE_WORKERS)

                if env_chunks and env_chunks.isdigit():
                    chunksize = max(1, int(env_chunks))
                else:
                    chunksize = max(1, DEFAULT_NORMALIZE_CHUNKSIZE)

                norm_vals = parallel_normalize(orig_vals, workers=cpu, chunksize=chunksize)
            except Exception:
                norm_vals = [normalize_text_for_matching(v) for v in orig_vals]

            norm_col_name = f"{c}__norm"
            norm_series_list.append(pl.Series(norm_col_name, norm_vals))
        if norm_series_list:
            df = df.with_columns(norm_series_list)
            cols = df.columns

    possible_date_cols = ["日付", "授業日"]
    date_col: str | None = None
    for dc in possible_date_cols:
        if dc in cols:
            date_col = dc
            break

    progress("日付情報からファイル名・シート名を生成中...")
    if date_col is not None:
        valid_dates = df.select(
            pl.col(date_col).cast(pl.Date, strict=False)
        ).drop_nulls()
    else:
        valid_dates = pl.DataFrame()

    if valid_dates.height > 0:
        min_dt = valid_dates.min().item()
        max_dt = valid_dates.max().item()

        wd_min = get_jp_weekday(min_dt)
        wd_max = get_jp_weekday(max_dt)

        min_date_file = min_dt.strftime("%Y%m%d")
        max_date_file = max_dt.strftime("%Y%m%d")

        date_range_str = f"集計期間: {min_dt.strftime('%Y/%m/%d')}({wd_min}) ～ {max_dt.strftime('%Y/%m/%d')}({wd_max})"
    else:
        min_date_file, max_date_file = "不明", "不明"
        date_range_str = "集計期間: 不明"

    sheet_name = "集計結果"
    now_str = datetime.now().strftime("%y%m%d-%H%M%S")
    input_dir = os.path.dirname(input_file)
    output_filename = f"集計結果_[{min_date_file}-{max_date_file}]_{now_str}.xlsx"
    output_file = os.path.join(input_dir, output_filename)

    progress("列インデックスの特定中...")
    if date_col is not None:
        dt_expr = pl.col(date_col).cast(pl.Date).dt
        dt_iso = cast(Any, dt_expr).week()
    else:
        dt_iso = pl.lit(None)

    lap_section_start_index = 0
    for i, c in enumerate(cols):
        if c and ("ラップテスト" in c or "ラップ" in c):
            lap_section_start_index = i
            break

    hw_section_start_index = len(cols)
    for i, c in enumerate(cols):
        if c and "宿題名" in c:
            hw_section_start_index = i
            break

    progress(
        f"インデックス特定完了: ラップ開始={lap_section_start_index}, 宿題開始={hw_section_start_index}"
    )

    test_section_column_names = cols[lap_section_start_index:hw_section_start_index]
    lap_count_column_names = [
        c for c in test_section_column_names if "ラップ" in c and "（分子）" in c
    ]
    confirmation_count_column_names = [
        c for c in test_section_column_names if "確認" in c and "（分子）" in c
    ]

    hw_page_exprs = []
    for i in range(hw_section_start_index, len(cols), 3):
        if i + 2 < len(cols):
            hw_name_col = cols[i]
            hw_start_col = cols[i + 1]
            hw_end_col = cols[i + 2]

            norm_col = f"{hw_name_col}__norm"
            hw_name_for_expr = norm_col if norm_col in cols else hw_name_col

            hw_page_exprs.append(
                build_hw_pages_expr(
                    hw_name_for_expr,
                    hw_start_col,
                    hw_end_col,
                    EXCLUDED_HW_TERMS,
                )
            )

    progress("データのクレンジングと指標の計算中...")

    missing_required_columns = []
    for req_col in ["出欠", "担当講師名", "担当講師N0", "教室"]:
        if req_col not in cols:
            missing_required_columns.append(req_col)

    if missing_required_columns:
        raise ValueError(
            f"必須の列が見つかりません: {', '.join(missing_required_columns)}"
        )

    attendance_metrics_df = (
        df.filter(pl.col("出欠") == "出席").with_columns(
            pl.col("担当講師名")
            .cast(pl.Utf8)
            .str.replace_all(r"[Ａ-Ｚ_]", "")
            .alias("担当講師名"),
            dt_iso.alias("week_num"),
            (
                pl.col(date_col).cast(pl.Date, strict=False)
                if date_col
                else pl.lit(None)
            ).alias("date_val"),
            (pl.sum_horizontal(hw_page_exprs) if hw_page_exprs else pl.lit(0)).alias(
                "hw_pages"
            ),
        )
        .with_columns(
            *build_test_presence_exprs(
                lap_count_column_names, confirmation_count_column_names
            )
        )
    )

    week_period_by_week_number = {}
    if "date_val" in attendance_metrics_df.columns:
        week_ranges_df = (
            attendance_metrics_df.filter(pl.col("date_val").is_not_null())
            .group_by("week_num")
            .agg(
                pl.col("date_val").min().alias("min_dt"),
                pl.col("date_val").max().alias("max_dt"),
            )
        )
        for row in week_ranges_df.iter_rows(named=True):
            w_num = row["week_num"]
            md = row["min_dt"]
            xd = row["max_dt"]
            if md and xd:
                week_period_by_week_number[str(w_num)] = (
                    f"{md.strftime('%Y/%m/%d')}({get_jp_weekday(md)}) ～ {xd.strftime('%Y/%m/%d')}({get_jp_weekday(xd)})"
                )

    def aggregate_and_pivot(attendance_metrics_local_df: pl.DataFrame):
        progress("週次集計中...")
        week_period_by_week_number_local = {}
        if "date_val" in attendance_metrics_local_df.columns:
            week_ranges_df_local = (
                attendance_metrics_local_df.filter(pl.col("date_val").is_not_null())
                .group_by("week_num")
                .agg(
                    pl.col("date_val").min().alias("min_dt"),
                    pl.col("date_val").max().alias("max_dt"),
                )
            )
            for row in week_ranges_df_local.iter_rows(named=True):
                w_num = row["week_num"]
                md = row["min_dt"]
                xd = row["max_dt"]
                if md and xd:
                    week_period_by_week_number_local[str(w_num)] = (
                        f"{md.strftime('%Y/%m/%d')}({get_jp_weekday(md)}) ～ {xd.strftime('%Y/%m/%d')}({get_jp_weekday(xd)})"
                    )

        base_aggregate_df_local = attendance_metrics_local_df.group_by(
            ["担当講師N0", "担当講師名", "教室", "week_num"]
        ).agg(
            pl.len().alias("授業数"),
            pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"),
            pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数"),
        )

        aggregated_df_local = base_aggregate_df_local.with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率"),
        ).select(
            [
                "担当講師N0",
                "担当講師名",
                "教室",
                "week_num",
                "授業数",
                "宿題平均",
                "ラップ回数",
                "ラップ率",
                "テスト回数",
                "テスト率",
            ]
        )

        progress("横持ちへのピボット処理...")
        pivot_df_local = aggregated_df_local.pivot(
            on="week_num",
            index=["担当講師N0", "担当講師名", "教室"],
            values=[
                "授業数",
                "宿題平均",
                "ラップ回数",
                "ラップ率",
                "テスト回数",
                "テスト率",
            ],
        )

        progress("希望の列順への並び替えとソート処理中...")
        week_numbers_local = sorted(
            list(set([c.split("_")[-1] for c in pivot_df_local.columns if "_" in c]))
        )

        class_count_columns_local = [f"授業数_{w}" for w in week_numbers_local]
        total_class_expr_local = pl.sum_horizontal(
            [
                pl.col(c)
                for c in class_count_columns_local
                if c in pivot_df_local.columns
            ]
        ).fill_null(0)
        pivot_df_local = pivot_df_local.with_columns(
            total_class_expr_local.alias("総授業数")
        )
        # 各週ごとに列の扱いを v3/v4 と同等にする:
        # - 授業数, ラップ回数, テスト回数 は欠損を0にする
        # - 宿題平均、ラップ率、テスト率 は授業数が0の場合は欠損のまま（後で '-' として表示される）
        for w in week_numbers_local:
            c_class = f"授業数_{w}"
            c_hw = f"宿題平均_{w}"
            c_lap_cnt = f"ラップ回数_{w}"
            c_lap_rate = f"ラップ率_{w}"
            c_test_cnt = f"テスト回数_{w}"
            c_test_rate = f"テスト率_{w}"

            if c_class in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(pl.col(c_class).fill_null(0).alias(c_class))
            if c_lap_cnt in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(pl.col(c_lap_cnt).fill_null(0).alias(c_lap_cnt))
            if c_test_cnt in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(pl.col(c_test_cnt).fill_null(0).alias(c_test_cnt))

            if c_hw in pivot_df_local.columns and c_class in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(
                    pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_hw)).alias(c_hw)
                )
            if c_lap_rate in pivot_df_local.columns and c_class in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(
                    pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_lap_rate)).alias(c_lap_rate)
                )
            if c_test_rate in pivot_df_local.columns and c_class in pivot_df_local.columns:
                pivot_df_local = pivot_df_local.with_columns(
                    pl.when(pl.col(c_class) == 0).then(pl.lit(None)).otherwise(pl.col(c_test_rate)).alias(c_test_rate)
                )

        desired_column_order_local = ["担当講師N0", "担当講師名", "教室", "総授業数"]
        rename_map_local = {
            "担当講師N0": "No",
            "担当講師名": "氏名",
            "教室": "教室",
            "総授業数": "総授業数",
        }

        week_period_labels_local = []
        for i, w in enumerate(week_numbers_local, start=1):
            c_class = f"授業数_{w}"
            c_hw = f"宿題平均_{w}"
            c_lap_cnt = f"ラップ回数_{w}"
            c_lap_rate = f"ラップ率_{w}"
            c_test_cnt = f"テスト回数_{w}"
            c_test_rate = f"テスト率_{w}"

            desired_column_order_local.extend(
                [c_class, c_hw, c_lap_cnt, c_lap_rate, c_test_cnt, c_test_rate]
            )

            rename_map_local[c_class] = f"授業{i}"
            rename_map_local[c_hw] = f"平均HW{i}"
            rename_map_local[c_lap_cnt] = f"Lap数{i}"
            rename_map_local[c_lap_rate] = f"Lap％{i}"
            rename_map_local[c_test_cnt] = f"テスト数{i}"
            rename_map_local[c_test_rate] = f"テスト％{i}"

            week_period_labels_local.append(
                week_period_by_week_number_local.get(str(w), f"Week {i}")
            )

        existing_desired_columns_local = [
            c for c in desired_column_order_local if c in pivot_df_local.columns
        ]

        pivot_df_local = (
            pivot_df_local.select(existing_desired_columns_local)
            .rename(
                {
                    k: v
                    for k, v in rename_map_local.items()
                    if k in existing_desired_columns_local
                }
            )
            .sort(["教室", "No"])
        )

        return pivot_df_local, week_period_labels_local

    pivot_df_all, week_period_labels_all = aggregate_and_pivot(attendance_metrics_df)

    subject_column_name = None
    for c in cols:
        if c in ("科目", "教科"):
            subject_column_name = c
            break

    classroom_exclusion_condition = (
        pl.col("教室").cast(pl.Utf8).str.contains(r"(高3|高３|高卒)").fill_null(False)
    )
    subject_exclusion_condition = (
        (
            pl.col(subject_column_name)
            .cast(pl.Utf8)
            .str.contains(r"国語")
            .fill_null(False)
        )
        if subject_column_name
        else pl.lit(False)
    )

    excluded_metrics_df = attendance_metrics_df.filter(
        ~classroom_exclusion_condition & ~subject_exclusion_condition
    )

    excluded_pivot_df, week_period_labels_excl = aggregate_and_pivot(
        excluded_metrics_df
    )

    classroom_summary_df = (
        attendance_metrics_df.group_by("教室")
        .agg(
            pl.len().alias("授業数"),
            pl.col("担当講師N0").n_unique().alias("講師数"),
            pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"),
            pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数"),
        )
        .with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率"),
        )
        .sort("教室")
    )

    excluded_classroom_summary_df = (
        excluded_metrics_df.group_by("教室")
        .agg(
            pl.len().alias("授業数"),
            pl.col("担当講師N0").n_unique().alias("講師数"),
            pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"),
            pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数"),
        )
        .with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率"),
        )
        .sort("教室")
    )

    return (
        pivot_df_all,
        excluded_pivot_df,
        classroom_summary_df,
        excluded_classroom_summary_df,
        output_file,
        date_range_str,
        sheet_name,
        week_period_labels_all,
        week_period_labels_excl,
    )


def format_excel_fast(
    all_pivot_df: pl.DataFrame,
    excluded_pivot_df: pl.DataFrame,
    classroom_summary_df: pl.DataFrame,
    excluded_classroom_summary_df: pl.DataFrame,
    output_file: str,
    date_range_str: str,
    sheet_name: str,
    week_period_labels_all: list[str],
    week_period_labels_excl: list[str],
):
    progress("xlsxwriterによる超高速書き出し・書式設定中...")

    # pandas の ExcelWriter を使ってデータを一括書き出しする（高速化）
    with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
        workbook = writer.book
        ud_font = "Noto Sans JP"

        fmt_title = workbook.add_format({
            "font_name": ud_font,
            "bold": True,
            "font_size": 11,
            "align": "left",
            "valign": "vcenter",
        })

        fmt_title_center = workbook.add_format({
            "font_name": ud_font,
            "bold": True,
            "font_size": 11,
            "align": "center",
            "valign": "vcenter",
            "bg_color": "#E8F6F3",
            "border": 1,
        })

        fmt_header = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#D3D3D3", "bold": True})
        fmt_white = workbook.add_format({"font_name": ud_font, "border": 1})
        fmt_gray = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#F5F5F5"})
        fmt_khaki = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#F0E68C"})
        fmt_lavender = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#E6E6FA"})

        fmt_white_pct = workbook.add_format({"font_name": ud_font, "border": 1, "num_format": "0.0%"})
        fmt_gray_pct = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#F5F5F5", "num_format": "0.0%"})

        fmt_white_center = workbook.add_format({"font_name": ud_font, "border": 1, "align": "center"})
        fmt_gray_center = workbook.add_format({"font_name": ud_font, "border": 1, "bg_color": "#F5F5F5", "align": "center"})

        fmt_cond_orange = workbook.add_format({"font_name": ud_font, "bg_color": "#FFA500", "font_color": "#000000"})
        fmt_cond_firebrick = workbook.add_format({"font_name": ud_font, "bg_color": "#B22222", "font_color": "#FFFFFF"})

        def make_safe_sheet_name(raw_name: str, used_names: set[str]) -> str:
            name = str(raw_name).strip() or "無名教室"
            translate_table = str.maketrans(
                {
                    "/": "／",
                    "\\": "＼",
                    "?": "？",
                    "*": "＊",
                    "[": "［",
                    "]": "］",
                    ":": "：",
                }
            )
            name = name.translate(translate_table)
            if name.startswith("'"):
                name = name[1:]
            if name.endswith("'"):
                name = name[:-1]
            name = name[:31] or "無名教室"

            base_name = name
            counter = 2
            while name in used_names:
                suffix = f"_{counter}"
                name = f"{base_name[:31 - len(suffix)]}{suffix}"
                counter += 1

            used_names.add(name)
            return name
        def write_sheet(
            writer,
            sheet_df: pl.DataFrame,
            week_period_labels: list[str],
            sheet_title: str,
            summary_text: str | None = None,
        ):
            progress(f"{sheet_title} の書き込みを開始します...")
            # Polarsベースで列を直接取り出して書き出す（pandas変換回避）
            headers = list(sheet_df.columns)
            max_col = len(headers) - 1

            header_row = 1
            if summary_text:
                header_row = 2

            data_start = header_row + 1

            nrows = sheet_df.height

            # 列値を Python リストで保持する辞書
            col_values: dict[str, list] = {}

            # helper to safely get series
            for col in headers:
                try:
                    s = sheet_df.get_column(col)
                except Exception:
                    s = pl.Series(col, [None] * nrows)

                # No の正規化（先頭ゼロ除去）
                if col == "No":
                    vals = [None if v is None else str(v).strip() for v in s.to_list()]
                    normed = []
                    for v in vals:
                        if v is None:
                            normed.append(None)
                            continue
                        tmp = v.lstrip("0")
                        normed.append(tmp if tmp != "" else v)
                    col_values[col] = normed
                    continue

                # counts-like columns
                if any(k in col for k in ("授業", "Lap数", "テスト数", "総授業数")):
                    try:
                        num = s.cast(pl.Float64, strict=False).fill_null(0)
                        col_values[col] = [None if v is None else (float(v) if v is not None else 0.0) for v in num.to_list()]
                    except Exception:
                        col_values[col] = [0.0] * nrows
                    continue

                # average / percent: keep None for missing
                if ("平均" in col) or ("％" in col) or ("率" in col):
                    try:
                        num = s.cast(pl.Float64, strict=False)
                        col_values[col] = [None if v is None else float(v) for v in num.to_list()]
                    except Exception:
                        col_values[col] = [None] * nrows
                    continue

                # fallback: try numeric, else strings
                try:
                    num = s.cast(pl.Float64, strict=False)
                    # Heuristic: if more than half values are numeric, treat as numeric
                    if (nrows == 0) or (num.null_count() < (nrows / 2)):
                        filled = num.fill_null(0)
                        col_values[col] = [None if v is None else (float(v) if v is not None else 0.0) for v in filled.to_list()]
                    else:
                        raw = s.to_list()
                        col_values[col] = [None if v is None else str(v) for v in raw]
                except Exception:
                    raw = s.to_list()
                    col_values[col] = [None if v is None else str(v) for v in raw]

            # 授業が0の期間については関連指標を空にする（v3/v4と一致）
            lesson_nums = set()
            for h in headers:
                m = re.search(r"授業(\d+)", str(h))
                if m:
                    lesson_nums.add(int(m.group(1)))

            for i in lesson_nums:
                lesson_col = f"授業{i}"
                if lesson_col not in col_values:
                    continue
                lesson_vals = col_values[lesson_col]
                related = [f"平均HW{i}", f"Lap数{i}", f"Lap％{i}", f"テスト数{i}", f"テスト％{i}"]
                for idx, lv in enumerate(lesson_vals):
                    if lv == 0 or lv == 0.0:
                        for c in related:
                            if c in col_values:
                                col_values[c][idx] = None

            # 最終的な欠損は '-' にして表示（文字列化）
            for col, vals in col_values.items():
                for j, v in enumerate(vals):
                    if v is None:
                        vals[j] = "-"

            # シート作成
            if sheet_title in writer.sheets:
                worksheet = writer.sheets[sheet_title]
            else:
                worksheet = workbook.add_worksheet(sheet_title)
                writer.sheets[sheet_title] = worksheet

            # データ列を一括書き出す
            for col_idx, col_name in enumerate(headers):
                arr = col_values.get(col_name, ["-"] * nrows)
                worksheet.write_column(data_start, col_idx, arr)

            # 列フォーマット: デフォルトを一括、パーセント列は連続範囲でまとめて適用
            if max_col >= 3:
                try:
                    worksheet.set_column(3, max_col, None, fmt_white)
                except Exception:
                    worksheet.set_column(3, max_col, 9, fmt_white)

            # collect percent/率 column indices and merge contiguous runs
            pct_idxs = [i for i, c in enumerate(headers) if ("％" in str(c) or "率" in str(c))]
            if pct_idxs:
                ranges = []
                start = prev = pct_idxs[0]
                for idx in pct_idxs[1:]:
                    if idx == prev + 1:
                        prev = idx
                        continue
                    ranges.append((start, prev))
                    start = prev = idx
                ranges.append((start, prev))

                for a, b in ranges:
                    try:
                        worksheet.set_column(a, b, None, fmt_white_pct)
                    except Exception:
                        worksheet.set_column(a, b, 9, fmt_white_pct)

            # タイトル／ヘッダー上書き（フォーマット適用）
            try:
                worksheet.merge_range(0, 0, 0, min(2, max_col), date_range_str, fmt_title)
            except Exception:
                worksheet.write(0, 0, date_range_str, fmt_title)
            worksheet.set_row(0, 20)

            # ヘッダーは一度に書く
            worksheet.write_row(header_row, 0, headers, fmt_header)

            # 列幅算出（先頭3列は日本語幅を考慮）
            max_a_width = 4
            max_b_width = 12
            max_c_width = 10
            for idx, col in enumerate(headers[:3]):
                vals = col_values.get(col, [""] * nrows)
                # estimate length from stringified values
                max_len = max((len(str(x)) for x in vals), default=0)
                if idx == 0:
                    max_a_width = max(max_a_width, float(max_len) * 1.9)
                elif idx == 1:
                    max_b_width = max(max_b_width, float(max_len) * 1.6)
                else:
                    max_c_width = max(max_c_width, float(max_len) * 1.6)

            worksheet.set_column(0, 0, max_a_width + 3)
            worksheet.set_column(1, 1, max_b_width + 3)
            worksheet.set_column(2, 2, max_c_width + 3)
            worksheet.set_column(3, 3, 9)
            if max_col >= 4:
                worksheet.set_column(4, max_col, 9)

            for i, period_str in enumerate(week_period_labels):
                start_col = 4 + i * 6
                if start_col + 5 <= max_col:
                    worksheet.merge_range(
                        0, start_col, 0, start_col + 5, period_str, fmt_title_center
                    )

            last_row = data_start + nrows - 1
            worksheet.autofilter(header_row, 0, last_row, max_col)

            # 条件付き書式は週ごとのブロックごとに3つだけ適用（既存と同様）
            for c_base in range(4, len(headers), 6):
                if c_base + 5 > max_col:
                    break

                col_hw = c_base + 1
                col_lap = c_base + 3
                col_test = c_base + 5

                worksheet.conditional_format(
                    data_start,
                    col_hw,
                    last_row,
                    col_hw,
                    {"type": "cell", "criteria": "<", "value": 6, "format": fmt_cond_orange},
                )

                worksheet.conditional_format(
                    data_start,
                    col_lap,
                    last_row,
                    col_lap,
                    {"type": "cell", "criteria": "<", "value": 0.7, "format": fmt_cond_firebrick},
                )
                worksheet.conditional_format(
                    data_start,
                    col_test,
                    last_row,
                    col_test,
                    {"type": "cell", "criteria": "<", "value": 0.7, "format": fmt_cond_firebrick},
                )

            worksheet.freeze_panes(data_start, 3)
            progress(f"{sheet_title} の書き込みが完了しました。")

        # シート書き出し（writer 経由）
        write_sheet(writer, all_pivot_df, week_period_labels_all, "全集計")
        write_sheet(writer, excluded_pivot_df, week_period_labels_excl, "除外集計")

        # 全教室サマリは後で結合して一度だけ書き出す（重複書き込みを避ける）

    def write_simple_sheet(
        worksheet, summary_df: pl.DataFrame, sheet_title: str, date_title: str
    ):
        # 列単位で一括書き出ししてフォーマットも列単位で適用する
        try:
            worksheet.merge_range(
                0, 0, 0, max(0, summary_df.width - 1), date_title, fmt_title
            )
        except Exception:
            worksheet.write(0, 0, date_title, fmt_title)
        worksheet.set_row(0, 20)

        # pandas を介して一括書き出し
        dfpd = summary_df.to_pandas()
        # ヘッダーは自分で書く
        for c, col_name in enumerate(dfpd.columns):
            worksheet.write(1, c, col_name, fmt_header)

        # データは to_excel を使って一括書き出しし、列単位でフォーマットを適用
        # write to a temporary area starting row 2
        tmp_start = 2
        dfpd.to_excel(writer, sheet_name=sheet_title, index=False, startrow=tmp_start, header=False)

        # 列ごとにフォーマット適用
        for c, colname in enumerate(dfpd.columns):
            if "率" in colname:
                worksheet.set_column(c, c, 12, fmt_white_pct)
            else:
                worksheet.set_column(c, c, 12, fmt_white)

        worksheet.autofilter(1, 0, 1 + dfpd.shape[0], max(0, dfpd.shape[1] - 1))
        worksheet.freeze_panes(2, 0)

    try:
        merged = classroom_summary_df.join(
            excluded_classroom_summary_df, on="教室", how="outer", suffix="_除外"
        )
    except Exception:
        merged = classroom_summary_df

    preferred = [
        "教室",
        "授業数",
        "講師数",
        "宿題平均",
        "ラップ率",
        "テスト率",
        "授業数_除外",
        "講師数_除外",
        "宿題平均_除外",
        "ラップ率_除外",
        "テスト率_除外",
    ]
    cols_to_show = [c for c in preferred if c in merged.columns]
    if "教室" not in cols_to_show:
        cols_to_show = merged.columns

    merged = merged.select(cols_to_show)

        # 全教室のサマリは上書きでフォーマット適用するため一度書いたシートを再利用
        # pandas で一括書き出した後、追加フォーマットは write_simple_sheet を使うため取得
        # ここは簡便化のためそのまま write_simple_sheet を呼ぶ代わりに write_sheet を使う
    write_sheet(writer, merged, [], "全教室")

    used_sheet_names = {"全集計", "除外集計"}
    classroom_values = []
    seen_classrooms = set()
    for value in all_pivot_df.get_column("教室").to_list():
        key = "" if value is None else str(value)
        if key in seen_classrooms:
            continue
        seen_classrooms.add(key)
        classroom_values.append(value)

    progress(f"教室別シートを作成中... 対象教室数={len(classroom_values)}")
    for classroom_value in classroom_values:
        classroom_label = (
            "無名教室"
            if classroom_value is None
            else str(classroom_value).strip() or "無名教室"
        )
        sheet_label = make_safe_sheet_name(classroom_label, used_sheet_names)

        if classroom_value is None:
            classroom_df = all_pivot_df.filter(pl.col("教室").is_null())
        else:
            classroom_df = all_pivot_df.filter(
                pl.col("教室").cast(pl.Utf8).str.strip_chars() == classroom_label
            )

        if classroom_df.height == 0:
            continue

        progress(f"教室シート作成: {classroom_label} -> {sheet_label}")
        summary_row = None
        try:
            if classroom_value is None:
                summary_row = classroom_summary_df.filter(pl.col("教室").is_null())
            else:
                summary_row = classroom_summary_df.filter(
                    pl.col("教室").cast(pl.Utf8).str.strip_chars() == classroom_label
                )
            if summary_row.height == 1:
                r = summary_row.row(0)
                total_classes = int(r[1]) if r[1] is not None else 0
                instructor_count = int(r[2]) if r[2] is not None else 0
                hw_avg = float(r[6]) if r[6] is not None else 0.0
                lap_rate = float(r[7]) if r[7] is not None else 0.0
                test_rate = float(r[8]) if r[8] is not None else 0.0
                top_summary = f"教室: {classroom_label}  総授業数: {total_classes}  講師数: {instructor_count}  宿題平均: {hw_avg:.1f}  Lap率: {lap_rate:.1%}  テスト率: {test_rate:.1%}"
            else:
                top_summary = f"教室: {classroom_label}"
        except Exception:
            top_summary = f"教室: {classroom_label}"

            # クラス別シートは書式付きの通常シートとして出力
        write_sheet(writer, classroom_df, week_period_labels_all, sheet_label, top_summary)
        progress("Excelファイルの保存が完了しました！")


def show_error_dialog(title: str, message: str, parent: tk.Tk | None = None) -> None:
    dlg = tk.Toplevel(parent) if parent else tk.Toplevel()
    dlg.title(title)

    try:
        if parent:
            scale = float(parent.tk.call("tk", "scaling"))
        else:
            scale = float(dlg.tk.call("tk", "scaling"))
    except Exception:
        scale = 1.0

    base_font = tkfont.nametofont("TkDefaultFont")
    try:
        fsize = max(int(base_font.cget("size") * scale), 10)
    except Exception:
        fsize = 11
    dlg_font = tkfont.Font(family=base_font.cget("family"), size=fsize)

    dlg.geometry(f"{int(700*scale)}x{int(360*scale)}")

    txt = tk.Text(dlg, wrap="word", font=dlg_font)
    txt.insert("1.0", message)
    txt.configure(state="disabled")
    txt.pack(expand=True, fill="both", padx=int(6 * scale), pady=int(6 * scale))

    frm = tk.Frame(dlg)
    frm.pack(fill="x", padx=int(6 * scale), pady=int(6 * scale))

    def _copy():
        try:
            dlg.clipboard_clear()
            dlg.clipboard_append(message)
        except Exception:
            pass

    btn_copy = tk.Button(
        frm, text="エラーメッセージをコピー", command=_copy, font=dlg_font
    )
    btn_copy.pack(side="left")

    btn_close = tk.Button(frm, text="閉じる", command=dlg.destroy, font=dlg_font)
    btn_close.pack(side="right")

    try:
        dlg.transient(parent)
        dlg.grab_set()
        dlg.wait_window()
    except Exception:
        pass


def main():
    enable_windows_dpi_awareness()
    root = tk.Tk()
    configure_tk_scaling(root)
    root.withdraw()
    try:
        root.attributes("-topmost", True)
        root.lift()
        root.focus_force()

        print("集計元のExcelファイルを選択してください...")
        input_file = filedialog.askopenfilename(
            title="集計元のExcelファイルを選択してください",
            filetypes=[
                ("Excelファイル", "*.xlsx *.xls *.xlsm"),
                ("すべてのファイル", "*.*"),
            ],
            parent=root,
        )

        root.attributes("-topmost", False)

        if not input_file:
            print("ファイルの選択がキャンセルされました。")
            return

        total_start_time = time.perf_counter()

        (
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
        ) = process_attendance_data(input_file)

        format_excel_fast(
            all_pivot_df,
            excluded_pivot_df,
            classroom_summary_df,
            excluded_classroom_summary_df,
            output_file,
            date_range_str,
            sheet_name,
            week_period_labels_all,
            week_period_labels_excl,
        )

        total_elapsed = time.perf_counter() - total_start_time

        result_msg = (
            f"処理が完了しました！\n"
            f"実行時間: {total_elapsed:.2f}秒\n\n"
            f"出力先:\n{output_file}"
        )

        root.attributes("-topmost", True)
        messagebox.showinfo("集計完了", result_msg, parent=root)
        root.attributes("-topmost", False)

        print(f"すべての処理が完了しました！ (実行時間: {total_elapsed:.2f}秒)")

    except KeyboardInterrupt:
        print("処理を中断しました。")
    except Exception as e:
        tb = traceback.format_exc()
        print("\n=== エラー詳細 ===")
        print(tb)
        print("==================\n")
        try:
            root.attributes("-topmost", True)
            show_error_dialog("処理中にエラーが発生しました", tb, parent=root)
        except Exception:
            messagebox.showerror(
                "エラー", f"処理中にエラーが発生しました:\n{e}", parent=root
            )
    finally:
        root.destroy()


if __name__ == "__main__":
    main()