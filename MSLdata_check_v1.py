import polars as pl
import tkinter as tk
from tkinter import filedialog, messagebox
import xlsxwriter
import openpyxl
import os
from datetime import datetime
from typing import Any, cast
import sys
import ctypes
import tkinter.font as tkfont
import traceback
import time
import math
import unicodedata


def enable_windows_dpi_awareness() -> None:
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


def configure_tk_scaling(root: tk.Tk) -> None:
    try:
        dpi = root.winfo_fpixels('1i')
        scale = float(dpi) / 96.0
        if scale <= 0:
            scale = 1.0
    except Exception:
        try:
            scale = float(root.tk.call('tk', 'scaling'))
        except Exception:
            scale = 1.0

    try:
        root.tk.call('tk', 'scaling', scale)
    except Exception:
        pass

    try:
        # GUI画面のフォントも見やすいUDフォントに変更
        base = tkfont.nametofont('TkDefaultFont')
        size = int(max(base.cget('size') * scale, 10))
        base.configure(family='Noto Sans JP', size=size)
    except Exception:
        pass


def progress(msg: str) -> None:
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def load_sheet_preserve_extra_columns(path: str, sheet_index: int = 1) -> pl.DataFrame:
    try:
        import fastexcel
        excel_reader = fastexcel.read_excel(path)
        sheet = excel_reader.load_sheet(sheet_index - 1, header_row=None)
        df_raw = sheet.to_polars()

        if df_raw.height == 0:
            return pl.DataFrame()

        progress("fastexcelエンジンによる読み込み完了。Polarsネイティブでヘッダー処理を開始します。")

        header = [str(x) if x is not None else "" for x in df_raw.row(0)]

        hw_base_idx = -1
        for i, h in enumerate(header):
            if "宿題名" in h:
                hw_base_idx = i
                break

        if hw_base_idx != -1:
            pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
            k = 2
            pattern_idx = 0
            for i in range(hw_base_idx + 1, len(header)):
                if header[i].strip() == "":
                    header[i] = f"{pattern[pattern_idx]}_{k}"
                    pattern_idx += 1
                    if pattern_idx >= 3:
                        pattern_idx = 0
                        k += 1
        else:
            idx = 1
            for i in range(len(header)):
                if header[i].strip() == "":
                    header[i] = f"追加列_{idx}"
                    idx += 1

        unique_headers = []
        seen = set()
        for h in header:
            h = h.strip()
            if h == "":
                h = "無名列"
            original_h = h
            counter = 1
            while h in seen:
                h = f"{original_h}_{counter}"
                counter += 1
            seen.add(h)
            unique_headers.append(h)

        df_data = df_raw.slice(1)
        df_data = df_data.rename(dict(zip(df_data.columns, unique_headers)))

        progress(f"データ構築完了: 行数={df_data.height}, 列数={df_data.width}")
        return df_data

    except Exception as e:
        progress(f"高速エンジンでの読み込み失敗。通常モードに切り替えます: {e}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[sheet_index - 1]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not rows:
            return pl.DataFrame()

        header = [str(c) if c is not None else "" for c in rows[0]]
        max_cols = max(len(r) for r in rows)

        if max_cols > len(header):
            try:
                base = header.index("宿題名")
                pattern = ["宿題名", "宿題開始ページ", "宿題終了ページ"]
                k = 2
                while len(header) < max_cols:
                    for p in pattern:
                        header.append(f"{p}_{k}")
                        if len(header) >= max_cols:
                            break
                    k += 1
            except ValueError:
                idx = 1
                while len(header) < max_cols:
                    header.append(f"追加列_{idx}")
                    idx += 1

        header_len = len(header)
        padded_rows = [tuple(row) + (None,) * (header_len - len(row)) if len(row) < header_len else tuple(row) for row in rows[1:]]

        df = pl.DataFrame(padded_rows, schema=header, orient="row")
        return df


def get_jp_weekday(dt_obj):
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    return weekdays[dt_obj.weekday()]


def process_attendance_data(input_file: str) -> tuple[pl.DataFrame, str, str, str, list[str]]:
    df = load_sheet_preserve_extra_columns(input_file, sheet_index=1)
    cols = df.columns

    possible_date_cols = ["日付", "授業日"]
    date_col: str | None = None
    for dc in possible_date_cols:
        if dc in cols:
            date_col = dc
            break

    progress("日付情報からファイル名・シート名を生成中...")
    if date_col is not None:
        valid_dates = df.select(
            pl.col(date_col).cast(pl.Date, strict=False)
        ).drop_nulls()
    else:
        valid_dates = pl.DataFrame()

    if valid_dates.height > 0:
        min_dt = valid_dates.min().item()
        max_dt = valid_dates.max().item()

        wd_min = get_jp_weekday(min_dt)
        wd_max = get_jp_weekday(max_dt)

        min_date_file = min_dt.strftime("%Y%m%d")
        max_date_file = max_dt.strftime("%Y%m%d")

        date_range_str = f"集計期間: {min_dt.strftime('%Y/%m/%d')}({wd_min}) ～ {max_dt.strftime('%Y/%m/%d')}({wd_max})"

        sheet_name = f"{min_dt.strftime('%Y.%m.%d')}({wd_min})～{max_dt.strftime('%Y.%m.%d')}({wd_max})"
        sheet_name = sheet_name[:31]
    else:
        min_date_file, max_date_file = "不明", "不明"
        date_range_str = "集計期間: 不明"
        sheet_name = "集計結果"

    now_str = datetime.now().strftime("%y%m%d-%H%M%S")
    input_dir = os.path.dirname(input_file)
    output_filename = f"集計結果_[{min_date_file}-{max_date_file}]_{now_str}.xlsx"
    output_file = os.path.join(input_dir, output_filename)

    progress("列インデックスの特定中...")
    if date_col is not None:
        dt_expr = pl.col(date_col).cast(pl.Date).dt
        dt_iso = cast(Any, dt_expr).week()
    else:
        dt_iso = pl.lit(None)

    lap_start_idx = 0
    for i, c in enumerate(cols):
        if c and ("ラップテスト" in c or "ラップ" in c):
            lap_start_idx = i
            break

    hw_start_idx = len(cols)
    for i, c in enumerate(cols):
        if c and "宿題名" in c:
            hw_start_idx = i
            break

    progress(f"インデックス特定完了: ラップ開始={lap_start_idx}, 宿題開始={hw_start_idx}")

    target_cols = cols[lap_start_idx:hw_start_idx]
    lap_cols = [c for c in target_cols if "ラップ" in c and "（分子）" in c]
    kaku_cols = [c for c in target_cols if "確認" in c and "（分子）" in c]

    hw_exprs = []
    for i in range(hw_start_idx, len(cols), 3):
        if i + 2 < len(cols):
            name_col = cols[i]
            start_col = cols[i+1]
            end_col = cols[i+2]

            start_val = pl.col(start_col).cast(pl.Utf8).str.extract(r"(\d+)").cast(pl.Float64).fill_null(0)
            end_val = pl.col(end_col).cast(pl.Utf8).str.extract(r"(\d+)").cast(pl.Float64).fill_null(0)

            calc = (
                pl.when(pl.col(name_col).is_not_null() & (pl.col(name_col).cast(pl.Utf8) != ""))
                .then(end_val - start_val + 1)
                .otherwise(0)
            )
            hw_exprs.append(calc)

    progress("データのクレンジングと指標の計算中...")

    missing_cols = []
    for req_col in ["出欠", "担当講師名", "担当講師N0", "教室"]:
        if req_col not in cols:
            missing_cols.append(req_col)

    if missing_cols:
        raise ValueError(f"必須の列が見つかりません: {', '.join(missing_cols)}")

    df_metrics = (
        df.filter(pl.col("出欠") == "出席")
        .with_columns(
            pl.col("担当講師名").cast(pl.Utf8).str.replace_all(r"[Ａ-Ｚ_]", "").alias("担当講師名"),
            dt_iso.alias("week_num"),
            (pl.col(date_col).cast(pl.Date, strict=False) if date_col else pl.lit(None)).alias("date_val"),
            (pl.sum_horizontal(hw_exprs) if hw_exprs else pl.lit(0)).alias("hw_pages"),
            (pl.any_horizontal([
                (pl.col(c).is_not_null() & (pl.col(c).cast(pl.Utf8) != "")) for c in lap_cols
            ]) if lap_cols else pl.lit(False)).fill_null(False).alias("has_lap"),
            (pl.any_horizontal([
                (pl.col(c).is_not_null() & (pl.col(c).cast(pl.Utf8) != "")) for c in kaku_cols
            ]) if kaku_cols else pl.lit(False)).fill_null(False).alias("has_kaku"),
        )
        .with_columns(
            (pl.col("has_lap") | pl.col("has_kaku")).alias("has_test")
        )
    )

    week_dict = {}
    if "date_val" in df_metrics.columns:
        week_ranges = (
            df_metrics.filter(pl.col("date_val").is_not_null())
            .group_by("week_num")
            .agg(
                pl.col("date_val").min().alias("min_dt"),
                pl.col("date_val").max().alias("max_dt")
            )
        )
        for row in week_ranges.iter_rows(named=True):
            w_num = row["week_num"]
            md = row["min_dt"]
            xd = row["max_dt"]
            if md and xd:
                week_dict[str(w_num)] = f"{md.strftime('%Y/%m/%d')}({get_jp_weekday(md)}) ～ {xd.strftime('%Y/%m/%d')}({get_jp_weekday(xd)})"

    progress("週次集計中...")
    df_agg_base = (
        df_metrics.group_by(["担当講師N0", "担当講師名", "教室", "week_num"])
        .agg(
            pl.len().alias("授業数"),
            pl.col("hw_pages").sum().alias("宿題合計"),
            pl.col("has_lap").cast(pl.Int32).sum().alias("ラップ回数"),
            pl.col("has_test").cast(pl.Int32).sum().alias("テスト回数")
        )
    )

    df_agg = (
        df_agg_base
        .with_columns(
            (pl.col("宿題合計") / pl.col("授業数")).round(1).alias("宿題平均"),
            (pl.col("ラップ回数") / pl.col("授業数")).alias("ラップ率"),
            (pl.col("テスト回数") / pl.col("授業数")).alias("テスト率"),
        )
        .select([
            "担当講師N0", "担当講師名", "教室", "week_num",
            "授業数", "宿題平均", "ラップ回数", "ラップ率", "テスト回数", "テスト率"
        ])
    )

    progress("横持ちへのピボット処理...")
    df_pivot = df_agg.pivot(
        on="week_num",
        index=["担当講師N0", "担当講師名", "教室"],
        values=["授業数", "宿題平均", "ラップ回数", "ラップ率", "テスト回数", "テスト率"]
    )

    progress("希望の列順への並び替えとソート処理中...")
    weeks = sorted(list(set([c.split("_")[-1] for c in df_pivot.columns if "_" in c])))

    # 総授業数を計算
    class_cols = [f"授業数_{w}" for w in weeks]
    total_class_expr = pl.sum_horizontal([pl.col(c) for c in class_cols if c in df_pivot.columns]).fill_null(0)
    df_pivot = df_pivot.with_columns(total_class_expr.alias("総授業数"))

    new_col_order = ["担当講師N0", "担当講師名", "教室", "総授業数"]
    rename_mapping = {
        "担当講師N0": "No",
        "担当講師名": "氏名",
        "教室": "教室",
        "総授業数": "総授業数"
    }

    week_periods = []
    for i, w in enumerate(weeks, start=1):
        c_class = f"授業数_{w}"
        c_hw = f"宿題平均_{w}"
        c_lap_cnt = f"ラップ回数_{w}"
        c_lap_rate = f"ラップ率_{w}"
        c_test_cnt = f"テスト回数_{w}"
        c_test_rate = f"テスト率_{w}"

        new_col_order.extend([c_class, c_hw, c_lap_cnt, c_lap_rate, c_test_cnt, c_test_rate])

        rename_mapping[c_class] = f"授業{i}"
        rename_mapping[c_hw] = f"平均HW{i}"
        rename_mapping[c_lap_cnt] = f"Lap数{i}"
        rename_mapping[c_lap_rate] = f"Lap％{i}"
        rename_mapping[c_test_cnt] = f"テスト数{i}"
        rename_mapping[c_test_rate] = f"テスト％{i}"

        week_periods.append(week_dict.get(str(w), f"Week {i}"))

    valid_new_col_order = [c for c in new_col_order if c in df_pivot.columns]

    df_pivot = (
        df_pivot.select(valid_new_col_order)
        .rename({k: v for k, v in rename_mapping.items() if k in valid_new_col_order})
        .sort(["教室", "No"])
    )

    return df_pivot, output_file, date_range_str, sheet_name, week_periods


def format_excel_fast(df: pl.DataFrame, output_file: str, date_range_str: str, sheet_name: str, week_periods: list[str]):
    progress("xlsxwriterによる超高速書き出し・書式設定中...")

    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet(sheet_name)

    # Windows標準の見やすいUDフォントを設定
    ud_font = 'Noto Sans JP'

    # 全ての書式に UDフォント を適用
    fmt_title = workbook.add_format({'font_name': ud_font, 'bold': True, 'font_size': 11, 'align': 'left', 'valign': 'vcenter'})
    fmt_title_center = workbook.add_format({'font_name': ud_font, 'bold': True, 'font_size': 11, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#E8F6F3', 'border': 1})

    fmt_header = workbook.add_format({'font_name': ud_font, 'border': 1, 'bg_color': '#D3D3D3', 'bold': True})
    fmt_white = workbook.add_format({'font_name': ud_font, 'border': 1})
    fmt_gray = workbook.add_format({'font_name': ud_font, 'border': 1, 'bg_color': '#F5F5F5'})
    fmt_khaki = workbook.add_format({'font_name': ud_font, 'border': 1, 'bg_color': '#F0E68C'})
    fmt_lavender = workbook.add_format({'font_name': ud_font, 'border': 1, 'bg_color': '#E6E6FA'})

    fmt_white_pct = workbook.add_format({'font_name': ud_font, 'border': 1, 'num_format': '0.0%'})
    fmt_gray_pct = workbook.add_format({'font_name': ud_font, 'border': 1, 'bg_color': '#F5F5F5', 'num_format': '0.0%'})

    fmt_white_center = workbook.add_format({'font_name': ud_font, 'border': 1, 'align': 'center'})
    fmt_gray_center = workbook.add_format({'font_name': ud_font, 'border': 1, 'bg_color': '#F5F5F5', 'align': 'center'})

    fmt_cond_orange = workbook.add_format({'font_name': ud_font, 'bg_color': '#FFA500', 'font_color': '#000000'})
    fmt_cond_firebrick = workbook.add_format({'font_name': ud_font, 'bg_color': '#B22222', 'font_color': '#FFFFFF'})

    headers = df.columns
    max_col = len(headers) - 1
    data_rows = df.rows()
    max_row_idx = len(data_rows) + 1

    # 1. 2行目（Row 1）にヘッダーを書き込む
    for col_num, col_name in enumerate(headers):
        worksheet.write(1, col_num, col_name, fmt_header)

    # カスタムオートフィット用の幅計算変数
    max_a_width = 4  # Noの最低幅
    max_b_width = 12 # 氏名の最低幅
    max_c_width = 10 # 教室の最低幅

    # 2. 3行目（Row 2）からデータ書き込み ＆ A〜C列の幅を計算
    for row_num, row_data in enumerate(data_rows, start=2):
        is_even_excel_row = ((row_num + 1) % 2 == 0)

        # A列(No), B列(氏名), C列(教室) の文字幅を計算して記録
        val_a = row_data[0]
        if val_a is not None:
            w_a = sum(2.2 if unicodedata.east_asian_width(c) in ('F', 'W', 'A') else 1.2 for c in str(val_a))
            if w_a > max_a_width: max_a_width = w_a

        val_b = row_data[1]
        if val_b is not None:
            w_b = sum(2.2 if unicodedata.east_asian_width(c) in ('F', 'W', 'A') else 1.2 for c in str(val_b))
            if w_b > max_b_width: max_b_width = w_b

        val_c = row_data[2]
        if val_c is not None:
            w_c = sum(2.2 if unicodedata.east_asian_width(c) in ('F', 'W', 'A') else 1.2 for c in str(val_c))
            if w_c > max_c_width: max_c_width = w_c

        for col_num, cell_value in enumerate(row_data):
            if col_num < 3:
                base_fmt = fmt_gray if is_even_excel_row else fmt_white
                if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
                    worksheet.write_blank(row_num, col_num, "", base_fmt)
                else:
                    worksheet.write(row_num, col_num, cell_value, base_fmt)
                continue
            
            # D列（総授業数）
            if col_num == 3:
                if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
                    worksheet.write_blank(row_num, col_num, "", fmt_lavender)
                else:
                    worksheet.write(row_num, col_num, cell_value, fmt_lavender)
                continue

            offset = (col_num - 4) % 6
            c_base = col_num - offset

            class_count = row_data[c_base]
            has_no_class = (class_count is None) or (isinstance(class_count, float) and math.isnan(class_count)) or (class_count == 0)

            if offset == 0:
                base_fmt = fmt_khaki
            elif offset == 3 or offset == 5:
                base_fmt = fmt_gray_pct if is_even_excel_row else fmt_white_pct
            else:
                base_fmt = fmt_gray if is_even_excel_row else fmt_white

            if has_no_class:
                if offset == 0:
                    worksheet.write(row_num, col_num, 0, base_fmt)
                else:
                    hyphen_fmt = fmt_gray_center if is_even_excel_row else fmt_white_center
                    worksheet.write(row_num, col_num, "-", hyphen_fmt)
            else:
                if cell_value is None or (isinstance(cell_value, float) and math.isnan(cell_value)):
                    worksheet.write_blank(row_num, col_num, "", base_fmt)
                else:
                    worksheet.write(row_num, col_num, cell_value, base_fmt)

    # 3. カスタムオートフィットの適用（UDフォント向けに少し余白を広めにとる）
    worksheet.set_column(0, 0, max_a_width + 3)  # No (計算した幅+余白)
    worksheet.set_column(1, 1, max_b_width + 3)  # 氏名 (計算した幅+余白)
    worksheet.set_column(2, 2, max_c_width + 3)  # 教室 (計算した幅+余白)
    worksheet.set_column(3, 3, 9)                # 総授業数
    if max_col >= 4:
        worksheet.set_column(4, max_col, 9)      # 授業数等の列はスッキリと固定幅

    # 4. 日付などのタイトルを入力（幅設定後に行うことで安全に結合可能）
    worksheet.merge_range(0, 0, 0, 2, date_range_str, fmt_title)
    worksheet.set_row(0, 20)

    for i, period_str in enumerate(week_periods):
        start_col = 4 + i * 6
        if start_col + 5 <= max_col:
            worksheet.merge_range(0, start_col, 0, start_col + 5, period_str, fmt_title_center)

    # 5. オートフィルターの設定
    worksheet.autofilter(1, 0, max_row_idx, max_col)

    # 6. 条件付き書式
    for c_base in range(4, len(headers), 6):
        if c_base + 5 > max_col:
            break

        col_hw = c_base + 1
        col_lap = c_base + 3
        col_test = c_base + 5

        worksheet.conditional_format(2, col_hw, max_row_idx, col_hw,
            {'type': 'cell', 'criteria': '<', 'value': 6, 'format': fmt_cond_orange})

        worksheet.conditional_format(2, col_lap, max_row_idx, col_lap,
            {'type': 'cell', 'criteria': '<', 'value': 0.7, 'format': fmt_cond_firebrick})
        worksheet.conditional_format(2, col_test, max_row_idx, col_test,
            {'type': 'cell', 'criteria': '<', 'value': 0.7, 'format': fmt_cond_firebrick})

    worksheet.freeze_panes(2, 3)
    workbook.close()
    progress("Excelファイルの保存が完了しました！")


def show_error_dialog(title: str, message: str, parent: tk.Tk | None = None) -> None:
    dlg = tk.Toplevel(parent) if parent else tk.Toplevel()
    dlg.title(title)

    try:
        if parent:
            scale = float(parent.tk.call("tk", "scaling"))
        else:
            scale = float(dlg.tk.call("tk", "scaling"))
    except Exception:
        scale = 1.0

    base_font = tkfont.nametofont("TkDefaultFont")
    try:
        fsize = max(int(base_font.cget("size") * scale), 10)
    except Exception:
        fsize = 11
    dlg_font = tkfont.Font(family=base_font.cget("family"), size=fsize)

    dlg.geometry(f"{int(700*scale)}x{int(360*scale)}")

    txt = tk.Text(dlg, wrap="word", font=dlg_font)
    txt.insert("1.0", message)
    txt.configure(state="disabled")
    txt.pack(expand=True, fill="both", padx=int(6*scale), pady=int(6*scale))

    frm = tk.Frame(dlg)
    frm.pack(fill="x", padx=int(6*scale), pady=int(6*scale))

    def _copy():
        try:
            dlg.clipboard_clear()
            dlg.clipboard_append(message)
        except Exception:
            pass

    btn_copy = tk.Button(frm, text="エラーメッセージをコピー", command=_copy, font=dlg_font)
    btn_copy.pack(side="left")

    btn_close = tk.Button(frm, text="閉じる", command=dlg.destroy, font=dlg_font)
    btn_close.pack(side="right")

    try:
        dlg.transient(parent)
        dlg.grab_set()
        dlg.wait_window()
    except Exception:
        pass


def main():
    enable_windows_dpi_awareness()
    root = tk.Tk()
    configure_tk_scaling(root)
    root.withdraw()

    # ダイアログが背面に隠れないように強制的に最前面に持ってくる
    root.attributes("-topmost", True)
    root.lift()
    root.focus_force()

    print("集計元のExcelファイルを選択してください...")
    input_file = filedialog.askopenfilename(
        title="集計元のExcelファイルを選択してください",
        filetypes=[("Excelファイル", "*.xlsx *.xls *.xlsm"), ("すべてのファイル", "*.*")],
        parent=root
    )

    root.attributes("-topmost", False)

    if not input_file:
        print("ファイルの選択がキャンセルされました。")
        return

    try:
        total_start_time = time.perf_counter()

        df_pivot, output_file, date_range_str, sheet_name, week_periods = process_attendance_data(input_file)

        format_excel_fast(df_pivot, output_file, date_range_str, sheet_name, week_periods)

        total_elapsed = time.perf_counter() - total_start_time

        result_msg = (
            f"処理が完了しました！\n"
            f"実行時間: {total_elapsed:.2f}秒\n\n"
            f"出力先:\n{output_file}"
        )

        root.attributes("-topmost", True)
        messagebox.showinfo("集計完了", result_msg, parent=root)
        root.attributes("-topmost", False)

        print(f"すべての処理が完了しました！ (実行時間: {total_elapsed:.2f}秒)")

    except Exception as e:
        tb = traceback.format_exc()
        print("\n=== エラー詳細 ===")
        print(tb)
        print("==================\n")
        try:
            root.attributes("-topmost", True)
            show_error_dialog("処理中にエラーが発生しました", tb, parent=root)
        except Exception:
            messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}", parent=root)
    finally:
        root.destroy()

if __name__ == "__main__":
    main()